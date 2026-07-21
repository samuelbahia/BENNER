#!/usr/bin/env python3
"""
CADASTRO DE PASTAS NO BENNER (PREVI JURÍDICO) - Versão Python/Selenium
=======================================================================
Automatiza cadastro via +Novo > Cadastro rápido de pasta (Categoria: Cível)
Usa IDs exatos dos campos ASP.NET conforme Mapa de Campos.

URL: https://previ.bennercloud.com.br/JURIDICO/jur/e/PREVI.aspx?i=K9_INICIOPREVI&m=MAIN

PREFIXO COMUM: ctl00_Main_WIDGET_CADASTRO_RAPIDO_PageControl_GERAL_GERAL_

REQUISITOS:
  pip install selenium openpyxl
  Driver do navegador (Chrome/Edge) no PATH
"""

import random
import re
import time
from datetime import date
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)

# ==============================================================================
# CONSTANTES DO SISTEMA
# ==============================================================================
URL_BENNER = "https://previ.bennercloud.com.br/JURIDICO/jur/e/PREVI.aspx?i=K9_INICIOPREVI&m=MAIN"
URL_PASTAS = "https://previ.bennercloud.com.br/JURIDICO/jur/e/PREVI.aspx?i=K9_INICIOPREVI&m=PASTAS"

CATEGORIA = "Cível"
TIPO_PASTA = "Cobrança"
CAUSA_PEDIR = "Previdencial"
CAUSA_RAIZ = "Produto"
PROCESSO = "Cobrança"
ANDAMENTO = "PEDIDO DE AJUIZAMENTO DE AÇÃO"
PEDIDO = "Dívida Previdenciária"
RITO = "Ordinário"
TIPO_PROCESSO = "Ativo"

# === IDs dos campos ASP.NET (Mapa de Campos Benner) ===
PFX = "ctl00_Main_WIDGET_CADASTRO_RAPIDO_PageControl_GERAL_GERAL_"

# Seção Pasta
ID_FILIAL_SELECT = PFX + "ctl11_ctl01_select"
ID_FILIAL_VALUE = PFX + "ctl11_FILIAL_VALUE"
ID_GERENCIA_SELECT = PFX + "ctl22_ctl01_select"
ID_GERENCIA_VALUE = PFX + "ctl22_DIVISAO_VALUE"
ID_CAUSA_PEDIR_SELECT = PFX + "ctl34_ctl01_select"
ID_CAUSA_PEDIR_VALUE = PFX + "ctl34_ASSUNTO_VALUE"
ID_CAUSA_RAIZ_SELECT = PFX + "ctl43_ctl01_select"
ID_CAUSA_RAIZ_VALUE = PFX + "ctl43_CAUSARAIZ_VALUE"
ID_TIPO_PROCESSO_ATIVO = PFX + "GroupRadioButton_TIPOPROCESSO_1"
ID_TIPO_PROCESSO_PASSIVO = PFX + "GroupRadioButton_TIPOPROCESSO_2"

# Seção Processo
ID_PROCESSO_SELECT = PFX + "ctl79_ctl01_select"
ID_PROCESSO_VALUE = PFX + "ctl79_DESDOBRAMENTO_VALUE"
ID_RITO_SELECT = PFX + "ctl87_ctl01_select"
ID_RITO_VALUE = PFX + "ctl87_RITO_VALUE"
ID_ORGAO_SELECT = PFX + "ctl95_ctl01_select"
ID_ORGAO_VALUE = PFX + "ctl95_ORGAO_VALUE"
ID_UF_SELECT = PFX + "ctl99_ctl01_select"
ID_UF_VALUE = PFX + "ctl99_UF_VALUE"

# Distribuição
ID_DATA_DISTRIBUICAO = PFX + "DISTRIBUIDO_1_DATADISTRIBUICAO_DATE"
ID_TIPO_DOC_PROCESSO_SELECT = PFX + "DISTRIBUIDO_1_ctl10_ctl01_select"
ID_TIPO_DOC_PROCESSO_VALUE = PFX + "DISTRIBUIDO_1_ctl10_TIPODOCUMENTO_VALUE"

# Número único
ID_NUMERO = PFX + "NUMEROUNICO_1_NUMERODISTRIBUICAO"

# Andamentos
ID_ANDAMENTO_SELECT = PFX + "ctl122_ctl01_select"
ID_ANDAMENTO_VALUE = PFX + "ctl122_EVENTO1_VALUE"
ID_DATA_ANDAMENTO = PFX + "DATAANDAMENTO1_DATE"

# Participantes
ID_ADVERSO_NAO = PFX + "POSSUIPESSOAADVERSO_ctl03"
ID_ADVERSO_SIM = PFX + "POSSUIPESSOAADVERSO_ctl05"
ID_PARTICIPANTE1_SELECT = PFX + "POSSUIPESSOAADVERSO_2_ctl04_ctl01_select"
ID_PARTICIPANTE1_VALUE = PFX + "POSSUIPESSOAADVERSO_2_ctl04_PARTICIPANTE1_VALUE"
ID_CONDICAO1_SELECT = PFX + "POSSUIPESSOAADVERSO_2_ctl13_ctl01_select"
ID_CONDICAO1_VALUE = PFX + "POSSUIPESSOAADVERSO_2_ctl13_CONDICAO1_VALUE"
ID_ADV_INTERNO_SELECT = PFX + "ctl202_ctl01_select"
ID_ADV_INTERNO_VALUE = PFX + "ctl202_ADVOGADOINTERNO_VALUE"
ID_ADV_EXTERNO_SELECT = PFX + "ctl206_ctl01_select"
ID_ADV_EXTERNO_VALUE = PFX + "ctl206_ADVOGADOEXTERNO_VALUE"

# Pedidos
ID_PEDIDO_SELECT = PFX + "ctl213_ctl01_select"
ID_PEDIDO_VALUE = PFX + "ctl213_PEDIDO1_VALUE"

# Documentos (para limpar)
ID_TIPO_DOC_ARQ_SELECT = PFX + "ctl256_ctl01_select"
ID_TIPO_DOC_ARQ_VALUE = PFX + "ctl256_TIPODOCUMENTOARQUIVO1_VALUE"
ID_NOME_ARQUIVO = PFX + "NOMEARQUIVO1"

# Colunas da planilha (0-indexed for openpyxl row tuple access)
COL_PLANO = 1          # A
COL_NOME = 4           # D
COL_CONTRATO = 6       # F
COL_VALOR_DIVIDA = 15  # O
COL_GERENCIA = 17      # Q
COL_UF = 20            # T
COL_CPF = 23           # W
COL_BENNER = 28        # AB
COL_ANALISE = 29       # AC
COL_STATUS = 30        # AD
COL_CNJ = 31           # AE
COL_PLANO_DESC = 32    # AF
COL_PESQUISA_BENNER = 33  # AG
COL_ID_PASTA = 34      # AH

# Advogados internos (seleção aleatória)
ADVOGADOS_INTERNOS = [
    "EDSON EDUARDO AGUIAR AVELAR",
    "MICHELLE CERQUEIRA NUNEZ",
    "DOMINIQUE DE SOUZA MACHADO",
]

# Advogados/escritórios externos (seleção aleatória)
ADVOGADOS_EXTERNOS = [
    "Aldrigues Cândido Advocacia",
    "Bicudo, Matos, e Moraes Sociedade de Advogados",
    "Dannemann Siemsen Advogados",
    "Queiroga, Vieira, Queiroz & Ramos Advocacia",
    "Wambier, Yamasaki, Bevervanço & Lobo Advocacia",
]

# Timeout padrão para aguardar elementos (segundos)
WAIT_TIMEOUT = 30
WAIT_AFTER_CLICK = 2
WAIT_BETWEEN_RECORDS = 3


# ==============================================================================
# CLASSE PRINCIPAL
# ==============================================================================
class CadastroPastasBenner:
    """Automatiza cadastro de pastas no Benner via Selenium."""

    def __init__(self, arquivo_excel: str, sheet_name: str = "Planilha1"):
        self.arquivo_excel = Path(arquivo_excel)
        self.sheet_name = sheet_name
        self.driver = None
        self.wait = None
        self.wb = None
        self.ws = None

    # ==========================================================================
    # INICIALIZAÇÃO
    # ==========================================================================
    def iniciar_navegador(self):
        """Inicia o navegador Chrome (ou Edge como fallback)."""
        try:
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            self.driver = webdriver.Chrome(options=options)
        except Exception:
            try:
                options = webdriver.EdgeOptions()
                options.add_argument("--start-maximized")
                self.driver = webdriver.Edge(options=options)
            except Exception as e:
                raise RuntimeError(f"Não foi possível iniciar o navegador: {e}")
        self.wait = WebDriverWait(self.driver, WAIT_TIMEOUT)
        print("Navegador iniciado.")

    def carregar_planilha(self):
        """Carrega a planilha Excel."""
        self.wb = openpyxl.load_workbook(str(self.arquivo_excel))
        self.ws = self.wb[self.sheet_name]
        print(f"Planilha carregada: {self.arquivo_excel} [{self.sheet_name}]")

    def salvar_planilha(self):
        """Salva a planilha Excel."""
        self.wb.save(str(self.arquivo_excel))
        print("Planilha salva.")

    def fechar(self):
        """Fecha navegador e salva planilha."""
        if self.wb:
            self.salvar_planilha()
        if self.driver:
            self.driver.quit()
            self.driver = None
        print("Encerrado.")

    # ==========================================================================
    # ETAPA 1 - ANÁLISE LOCAL DE DUPLICIDADES
    # ==========================================================================
    def analise_previa_duplicidades(self):
        """Analisa duplicidades locais na planilha."""
        self.carregar_planilha()
        ws = self.ws

        last_row = ws.max_row

        # Cabeçalhos das colunas auxiliares
        ws.cell(1, COL_ANALISE, "ANÁLISE DUPLICIDADE")
        ws.cell(1, COL_STATUS, "STATUS CADASTRO")
        ws.cell(1, COL_CNJ, "NÚMERO CNJ")
        ws.cell(1, COL_PLANO_DESC, "PLANO DESCRIÇÃO")
        ws.cell(1, COL_PESQUISA_BENNER, "PESQUISA BENNER")
        ws.cell(1, COL_ID_PASTA, "ID PASTA BENNER")

        # Limpar dados anteriores
        for row in range(2, last_row + 1):
            for col in range(COL_ANALISE, COL_ID_PASTA + 1):
                ws.cell(row, col, None)

        # Contar ocorrências de nomes
        nomes_count: dict[str, int] = {}
        for row in range(2, last_row + 1):
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            nomes_count[nome] = nomes_count.get(nome, 0) + 1

        # Detectar duplicatas exatas e preencher campos auxiliares
        chaves_vistas: set[str] = set()
        for row in range(2, last_row + 1):
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            contrato = str(ws.cell(row, COL_CONTRATO).value or "")
            valor = float(ws.cell(row, COL_VALOR_DIVIDA).value or 0)
            chave = f"{nome}|{contrato}|{valor}"

            # Número CNJ
            ws.cell(row, COL_CNJ, f"DP{contrato}")

            # Plano descrição
            plano = ws.cell(row, COL_PLANO).value
            if plano == 1:
                ws.cell(row, COL_PLANO_DESC, "Plano de Benefícios 1")
            elif plano == 2:
                ws.cell(row, COL_PLANO_DESC, "Plano PREVI Futuro")

            # Análise de duplicidade
            if chave in chaves_vistas:
                ws.cell(row, COL_ANALISE, "DUPLICATA EXATA - REMOVER")
                ws.cell(row, COL_STATUS, "NÃO CADASTRAR")
            elif nomes_count.get(nome, 0) > 1:
                ws.cell(row, COL_ANALISE, f"MESMO PARTICIPANTE - {nomes_count[nome]} OPERAÇÕES")
                ws.cell(row, COL_STATUS, "VERIFICAR")
            else:
                ws.cell(row, COL_ANALISE, "OK")
                ws.cell(row, COL_STATUS, "PENDENTE")
            chaves_vistas.add(chave)

            # Verificar se já está no Benner
            benner_val = str(ws.cell(row, COL_BENNER).value or "").strip()
            if benner_val:
                analise_atual = str(ws.cell(row, COL_ANALISE).value or "")
                ws.cell(row, COL_ANALISE, f"{analise_atual} | JÁ NO BENNER ({benner_val})")
                ws.cell(row, COL_STATUS, "JÁ CADASTRADO")

        self.salvar_planilha()
        total = last_row - 1
        print(f"Etapa 1 concluída - Análise local. Total: {total} operações.")
        return total

    # ==========================================================================
    # ETAPA 2 - PESQUISA NO BENNER (Pastas > Parte Pasta)
    # ==========================================================================
    def verificar_no_benner(self):
        """Pesquisa cada participante no Benner para verificar existência."""
        self.carregar_planilha()
        ws = self.ws

        if str(ws.cell(1, COL_STATUS).value or "") != "STATUS CADASTRO":
            print("ERRO: Execute primeiro a Etapa 1!")
            return

        self.iniciar_navegador()
        self.driver.get(URL_PASTAS)
        self._aguardar_carregamento()

        last_row = ws.max_row
        pesquisados = 0
        ja_existentes = 0

        for row in range(2, last_row + 1):
            status = str(ws.cell(row, COL_STATUS).value or "").strip().upper()
            if status not in ("PENDENTE", "VERIFICAR"):
                continue

            nome = str(ws.cell(row, COL_NOME).value or "").strip()
            if not nome:
                continue

            resultado = self._pesquisar_parte_pasta(nome)
            ws.cell(row, COL_PESQUISA_BENNER, resultado)
            pesquisados += 1

            resultado_upper = resultado.upper()
            if "ENCONTRADA" in resultado_upper:
                if "DÍVIDA PREVIDENCIÁRIA" in resultado_upper or "DIVIDA PREVIDENCIARIA" in resultado_upper:
                    ws.cell(row, COL_STATUS, "JÁ CADASTRADO NO BENNER")
                    ja_existentes += 1
                else:
                    analise_atual = str(ws.cell(row, COL_ANALISE).value or "")
                    ws.cell(row, COL_ANALISE, f"{analise_atual} | PASTA EXISTENTE OUTRO OBJETO")
            elif "NÃO ENCONTRADA" in resultado_upper:
                if status == "VERIFICAR":
                    ws.cell(row, COL_STATUS, "PENDENTE")

            time.sleep(WAIT_AFTER_CLICK)
            print(f"  Pesquisando... {pesquisados}/{last_row - 1}")

        self.salvar_planilha()
        print(f"Etapa 2 concluída. Pesquisados: {pesquisados}, Já existentes: {ja_existentes}")

    # ==========================================================================
    # ETAPA 3 - CADASTRO VIA +NOVO > CADASTRO RÁPIDO DE PASTA (CÍVEL)
    # ==========================================================================
    def cadastrar_pastas(self):
        """Cadastra pastas pendentes no Benner."""
        self.carregar_planilha()
        ws = self.ws

        if str(ws.cell(1, COL_STATUS).value or "") != "STATUS CADASTRO":
            print("ERRO: Execute primeiro as Etapas 1 e 2!")
            return

        last_row = ws.max_row
        pendentes = []
        for row in range(2, last_row + 1):
            if str(ws.cell(row, COL_STATUS).value or "").strip().upper() == "PENDENTE":
                pendentes.append(row)

        if not pendentes:
            print("Nenhuma operação PENDENTE.")
            return

        print(f"Cadastrando {len(pendentes)} pastas...")
        self.iniciar_navegador()
        self.driver.get(URL_BENNER)
        self._aguardar_carregamento()

        cadastrados = 0
        erros = 0

        for row in pendentes:
            nome = str(ws.cell(row, COL_NOME).value or "").strip()
            contrato = str(ws.cell(row, COL_CONTRATO).value or "")
            valor_divida = float(ws.cell(row, COL_VALOR_DIVIDA).value or 0)
            gerencia = str(ws.cell(row, COL_GERENCIA).value or "").strip()
            uf = str(ws.cell(row, COL_UF).value or "").strip()
            cpf = self._formatar_cpf(str(ws.cell(row, COL_CPF).value or ""))
            filial = str(ws.cell(row, COL_PLANO_DESC).value or "")
            numero_cnj = f"DP{contrato}"

            adv_interno = random.choice(ADVOGADOS_INTERNOS)
            adv_externo = random.choice(ADVOGADOS_EXTERNOS)

            resultado = self._cadastrar_pasta_civel(
                nome, contrato, valor_divida, gerencia, uf, cpf,
                filial, numero_cnj, adv_interno, adv_externo
            )

            if resultado.startswith("OK"):
                ws.cell(row, COL_STATUS, "CADASTRADO + ANDAMENTO")
                if "|" in resultado:
                    ws.cell(row, COL_ID_PASTA, resultado.split("|", 1)[1])
                cadastrados += 1
            else:
                ws.cell(row, COL_STATUS, f"ERRO: {resultado}")
                erros += 1

            self.salvar_planilha()
            print(f"  Cadastrando... {cadastrados + erros}/{len(pendentes)} "
                  f"(OK: {cadastrados}, Erros: {erros})")
            time.sleep(WAIT_BETWEEN_RECORDS)

        print(f"Concluído! Sucesso: {cadastrados}, Erros: {erros}")

    # ==========================================================================
    # FUNÇÃO PRINCIPAL - CADASTRAR PASTA CÍVEL (IDs exatos)
    # ==========================================================================
    def _cadastrar_pasta_civel(
        self, nome: str, contrato: str, valor_divida: float,
        gerencia: str, uf: str, cpf: str, filial: str,
        numero_cnj: str, adv_interno: str, adv_externo: str
    ) -> str:
        """Executa o cadastro de uma pasta Cível. Retorna 'OK|id' ou mensagem de erro."""
        try:
            # === PASSO 1: +Novo > Cadastro rápido de pasta ===
            btn_novo = self._buscar_elemento_por_texto("a", "+Novo")
            if not btn_novo:
                btn_novo = self._buscar_elemento_por_texto("span", "Novo")
            if not btn_novo:
                btn_novo = self._buscar_elemento_por_texto("button", "Novo")
            if not btn_novo:
                return "Botão +Novo não encontrado"

            btn_novo.click()
            self._aguardar_carregamento()

            link_cad_rapido = self._buscar_elemento_por_texto("a", "Cadastro rápido de pasta")
            if not link_cad_rapido:
                link_cad_rapido = self._buscar_elemento_por_texto("span", "Cadastro rápido")
            if not link_cad_rapido:
                return "Cadastro rápido não encontrado"

            link_cad_rapido.click()
            self._aguardar_carregamento()

            # === PASSO 2: Selecionar Categoria Cível ===
            self._preencher_campo_por_label("Categoria", CATEGORIA)
            self._aguardar_carregamento()

            # === PASSO 3: Preencher campos com IDs exatos ===
            # Filial
            self._selecionar_lookup(ID_FILIAL_SELECT, ID_FILIAL_VALUE, filial)
            # Gerência
            self._selecionar_lookup(ID_GERENCIA_SELECT, ID_GERENCIA_VALUE, gerencia)
            # Causa de Pedir
            self._selecionar_lookup(ID_CAUSA_PEDIR_SELECT, ID_CAUSA_PEDIR_VALUE, CAUSA_PEDIR)
            # Causa Raiz
            self._selecionar_lookup(ID_CAUSA_RAIZ_SELECT, ID_CAUSA_RAIZ_VALUE, CAUSA_RAIZ)
            # Tipo Processo: Ativo
            self._clicar_radio(ID_TIPO_PROCESSO_ATIVO)
            # Processo
            self._selecionar_lookup(ID_PROCESSO_SELECT, ID_PROCESSO_VALUE, PROCESSO)
            # Rito
            self._selecionar_lookup(ID_RITO_SELECT, ID_RITO_VALUE, RITO)
            # Órgão
            self._selecionar_lookup(ID_ORGAO_SELECT, ID_ORGAO_VALUE, "Tribunal de Justiça")
            # UF
            self._selecionar_lookup(ID_UF_SELECT, ID_UF_VALUE, uf)

            # Data distribuição: hoje
            data_hoje = date.today().strftime("%d/%m/%Y")
            self._preencher_texto(ID_DATA_DISTRIBUICAO, data_hoje)
            # Número
            self._preencher_texto(ID_NUMERO, numero_cnj)
            # Andamento
            self._selecionar_lookup(ID_ANDAMENTO_SELECT, ID_ANDAMENTO_VALUE, ANDAMENTO)
            # Data andamento: hoje
            self._preencher_texto(ID_DATA_ANDAMENTO, data_hoje)

            # === PASSO 4: Participantes ===
            self._clicar_radio(ID_ADVERSO_SIM)
            time.sleep(1)
            # Participante 1 (adverso)
            self._selecionar_lookup(ID_PARTICIPANTE1_SELECT, ID_PARTICIPANTE1_VALUE, nome)
            # Condição 1: Réu
            self._selecionar_lookup(ID_CONDICAO1_SELECT, ID_CONDICAO1_VALUE, "Réu")
            # Advogado interno
            self._selecionar_lookup(ID_ADV_INTERNO_SELECT, ID_ADV_INTERNO_VALUE, adv_interno)
            # Advogado externo
            self._selecionar_lookup(ID_ADV_EXTERNO_SELECT, ID_ADV_EXTERNO_VALUE, adv_externo)

            # === PASSO 5: Pedido ===
            self._selecionar_lookup(ID_PEDIDO_SELECT, ID_PEDIDO_VALUE, PEDIDO)

            # === PASSO 6: Documentos - Limpar ===
            self._limpar_campo(ID_TIPO_DOC_ARQ_VALUE)
            self._limpar_campo(ID_NOME_ARQUIVO)
            self._limpar_select(ID_TIPO_DOC_ARQ_SELECT)

            # === PASSO 7: Salvar ===
            btn_salvar = self._buscar_elemento_por_texto("a", "Salvar")
            if not btn_salvar:
                btn_salvar = self._buscar_elemento_por_texto("button", "Salvar")
            if not btn_salvar:
                btn_salvar = self._buscar_elemento_por_texto("span", "Salvar")
            if not btn_salvar:
                return "Botão Salvar não encontrado"

            btn_salvar.click()
            self._aguardar_carregamento()

            # === PASSO 8: Capturar ID da pasta ===
            id_pasta = self._capturar_id_pasta()
            if id_pasta:
                return f"OK|{id_pasta}"
            return "OK"

        except Exception as e:
            return str(e)

    # ==========================================================================
    # FUNÇÕES DE PREENCHIMENTO POR ID EXATO
    # ==========================================================================
    def _selecionar_lookup(self, id_select: str, id_value: str, texto: str):
        """
        Campos lookup do Benner: SELECT visível + INPUT hidden _VALUE.
        Seleciona a opção pelo texto e seta o valor oculto.
        """
        try:
            select_elem = self.driver.find_element(By.ID, id_select)
            select_obj = Select(select_elem)
            # Buscar opção que contém o texto (case-insensitive)
            for option in select_obj.options:
                if texto.lower() in option.text.lower():
                    select_obj.select_by_visible_text(option.text)
                    # Setar hidden value
                    try:
                        hidden = self.driver.find_element(By.ID, id_value)
                        self.driver.execute_script(
                            "arguments[0].value = arguments[1];", hidden, option.get_attribute("value")
                        )
                    except NoSuchElementException:
                        pass
                    # Disparar change
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                        select_elem
                    )
                    return
        except NoSuchElementException:
            pass
        except Exception:
            pass

    def _preencher_texto(self, id_campo: str, valor: str):
        """Preenche campo de texto pelo ID e dispara eventos."""
        try:
            elem = self.driver.find_element(By.ID, id_campo)
            elem.clear()
            elem.send_keys(valor)
            self.driver.execute_script("""
                var e = arguments[0];
                e.dispatchEvent(new Event('change', {bubbles: true}));
                e.dispatchEvent(new Event('input', {bubbles: true}));
                e.dispatchEvent(new Event('blur', {bubbles: true}));
            """, elem)
        except NoSuchElementException:
            pass

    def _clicar_radio(self, id_radio: str):
        """Clica em radio button pelo ID."""
        try:
            elem = self.driver.find_element(By.ID, id_radio)
            elem.click()
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", elem
            )
        except NoSuchElementException:
            pass

    def _limpar_campo(self, id_campo: str):
        """Limpa o valor de um campo."""
        try:
            elem = self.driver.find_element(By.ID, id_campo)
            self.driver.execute_script("arguments[0].value = '';", elem)
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", elem
            )
        except NoSuchElementException:
            pass

    def _limpar_select(self, id_select: str):
        """Reseta um select para o primeiro índice."""
        try:
            elem = self.driver.find_element(By.ID, id_select)
            Select(elem).select_by_index(0)
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", elem
            )
        except NoSuchElementException:
            pass

    # ==========================================================================
    # PESQUISA NO BENNER
    # ==========================================================================
    def _pesquisar_parte_pasta(self, nome: str) -> str:
        """Pesquisa participante em Pastas > Parte Pasta."""
        try:
            # Navegar para Pastas se necessário
            menu_pastas = self._buscar_elemento_por_texto("a", "Pastas")
            if not menu_pastas:
                menu_pastas = self._buscar_elemento_por_texto("span", "Pastas")
            if menu_pastas:
                menu_pastas.click()
                self._aguardar_carregamento()

            # Buscar campo Parte Pasta
            campo_parte = self._buscar_campo_por_label("Parte Pasta")
            if not campo_parte:
                campo_parte = self._buscar_input_por_atributo("placeholder", "Parte")
            if not campo_parte:
                campo_parte = self._buscar_input_por_atributo("title", "Parte")
            if not campo_parte:
                return "ERRO: Campo não encontrado"

            campo_parte.clear()
            campo_parte.send_keys(nome)
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", campo_parte
            )
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", campo_parte
            )

            # Pesquisar
            btn_pesquisar = self._buscar_elemento_por_texto("button", "Pesquisar")
            if not btn_pesquisar:
                btn_pesquisar = self._buscar_elemento_por_texto("a", "Pesquisar")
            if btn_pesquisar:
                btn_pesquisar.click()
            else:
                campo_parte.send_keys(Keys.ENTER)
            self._aguardar_carregamento()
            time.sleep(2)

            # Ler resultados
            resultado = self._ler_resultados_pesquisa(nome)

            # Limpar campo
            try:
                campo_parte = self._buscar_campo_por_label("Parte Pasta")
                if campo_parte:
                    campo_parte.clear()
            except Exception:
                pass

            return resultado

        except Exception as e:
            return f"ERRO: {e}"

    def _ler_resultados_pesquisa(self, nome_pesquisado: str) -> str:
        """Interpreta resultados da pesquisa de Parte Pasta."""
        try:
            body_text = self.driver.find_element(By.TAG_NAME, "body").text.upper()

            # Verificar se há tabelas com resultados
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            encontrou = False
            objetos = ""

            for table in tables:
                try:
                    rows = table.find_elements(By.TAG_NAME, "tr")
                    for tr in rows:
                        row_text = tr.text.upper()
                        if nome_pesquisado.upper() in row_text:
                            encontrou = True
                            if "DÍVIDA PREVIDENCIÁRIA" in row_text or "DIVIDA PREVIDENCIARIA" in row_text:
                                objetos += "DÍVIDA PREVIDENCIÁRIA; "
                            else:
                                cells = tr.find_elements(By.TAG_NAME, "td")
                                if len(cells) > 1:
                                    objetos += cells[1].text[:50] + "; "
                except StaleElementReferenceException:
                    continue

            if encontrou:
                if "DÍVIDA PREVIDENCIÁRIA" in objetos.upper() or "DIVIDA PREVIDENCIARIA" in objetos.upper():
                    return "ENCONTRADA - MESMO OBJETO (DÍVIDA PREVIDENCIÁRIA)"
                elif objetos:
                    return f"ENCONTRADA - OUTRO OBJETO: {objetos[:100]}"
                else:
                    return "ENCONTRADA - objeto não identificado"
            else:
                if "NENHUM REGISTRO" in body_text or "NÃO ENCONTR" in body_text:
                    return "NÃO ENCONTRADA - OK para cadastrar"
                else:
                    return "NÃO ENCONTRADA - verificar manualmente"

        except Exception as e:
            return f"ERRO ao ler resultados: {e}"

    # ==========================================================================
    # CAPTURAR ID DA PASTA
    # ==========================================================================
    def _capturar_id_pasta(self) -> str:
        """Tenta capturar o ID da pasta recém-criada."""
        try:
            current_url = self.driver.current_url
            match = re.search(r"id=(\d+)", current_url)
            if match:
                return match.group(1)
        except Exception:
            pass

        try:
            campo_codigo = self._buscar_campo_por_label("Código")
            if campo_codigo:
                return campo_codigo.get_attribute("value") or ""
        except Exception:
            pass

        return ""

    # ==========================================================================
    # FUNÇÕES AUXILIARES
    # ==========================================================================
    def _aguardar_carregamento(self):
        """Aguarda a página terminar de carregar."""
        try:
            self.wait.until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except TimeoutException:
            pass
        time.sleep(WAIT_AFTER_CLICK)

    def _buscar_elemento_por_texto(self, tag: str, texto: str):
        """Busca elemento por tag e texto contido (case-insensitive)."""
        try:
            elements = self.driver.find_elements(By.TAG_NAME, tag)
            for elem in elements:
                try:
                    if texto.lower() in elem.text.lower():
                        return elem
                except StaleElementReferenceException:
                    continue
        except Exception:
            pass
        return None

    def _buscar_campo_por_label(self, label_text: str):
        """Busca campo de formulário associado a um label."""
        try:
            labels = self.driver.find_elements(By.TAG_NAME, "label")
            for label in labels:
                try:
                    if label_text.lower() in label.text.lower():
                        for_id = label.get_attribute("for")
                        if for_id:
                            try:
                                return self.driver.find_element(By.ID, for_id)
                            except NoSuchElementException:
                                pass
                        # Buscar input/select no elemento pai
                        parent = label.find_element(By.XPATH, "..")
                        inputs = parent.find_elements(By.TAG_NAME, "input")
                        if inputs:
                            return inputs[0]
                        selects = parent.find_elements(By.TAG_NAME, "select")
                        if selects:
                            return selects[0]
                except StaleElementReferenceException:
                    continue
        except Exception:
            pass
        return None

    def _buscar_input_por_atributo(self, atributo: str, valor: str):
        """Busca input por atributo (placeholder, title, etc.)."""
        try:
            inputs = self.driver.find_elements(By.TAG_NAME, "input")
            for inp in inputs:
                try:
                    attr_val = inp.get_attribute(atributo) or ""
                    if valor.lower() in attr_val.lower():
                        return inp
                except StaleElementReferenceException:
                    continue
        except Exception:
            pass
        return None

    def _preencher_campo_por_label(self, label_text: str, valor: str):
        """Preenche campo localizado via label."""
        campo = self._buscar_campo_por_label(label_text)
        if not campo:
            return
        tag = campo.tag_name.lower()
        if tag == "select":
            select_obj = Select(campo)
            for option in select_obj.options:
                if valor.lower() in option.text.lower():
                    select_obj.select_by_visible_text(option.text)
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", campo
                    )
                    return
        else:
            campo.clear()
            campo.send_keys(valor)
            self.driver.execute_script("""
                var e = arguments[0];
                e.dispatchEvent(new Event('change', {bubbles: true}));
                e.dispatchEvent(new Event('input', {bubbles: true}));
                e.dispatchEvent(new Event('blur', {bubbles: true}));
            """, campo)

    @staticmethod
    def _formatar_cpf(cpf_raw: str) -> str:
        """Formata CPF: 000.000.000-00"""
        cpf = re.sub(r"[.\-\s]", "", cpf_raw.strip())
        cpf = cpf.zfill(11)
        return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:11]}"

    # ==========================================================================
    # RELATÓRIO
    # ==========================================================================
    def gerar_relatorio(self):
        """Imprime relatório de status das operações."""
        self.carregar_planilha()
        ws = self.ws
        last_row = ws.max_row

        stats = {"pendentes": 0, "cadastrados": 0, "erros": 0,
                 "duplicatas": 0, "ja_cadastrados": 0, "verificar": 0}

        for row in range(2, last_row + 1):
            st = str(ws.cell(row, COL_STATUS).value or "").strip().upper()
            if st == "PENDENTE":
                stats["pendentes"] += 1
            elif st in ("CADASTRADO", "CADASTRADO + ANDAMENTO"):
                stats["cadastrados"] += 1
            elif st == "NÃO CADASTRAR":
                stats["duplicatas"] += 1
            elif st in ("JÁ CADASTRADO", "JÁ CADASTRADO NO BENNER"):
                stats["ja_cadastrados"] += 1
            elif st == "VERIFICAR":
                stats["verificar"] += 1
            elif "ERRO" in st:
                stats["erros"] += 1

        total = last_row - 1
        print(f"""
{'='*50}
 RELATÓRIO DE STATUS
{'='*50}
 Total:          {total}
 Pendentes:      {stats['pendentes']}
 Cadastradas:    {stats['cadastrados']}
 Duplicatas:     {stats['duplicatas']}
 Já no Benner:   {stats['ja_cadastrados']}
 Verificar:      {stats['verificar']}
 Erros:          {stats['erros']}
{'='*50}
""")


# ==============================================================================
# EXECUÇÃO PRINCIPAL
# ==============================================================================
def main():
    """Ponto de entrada - executa as 3 etapas em sequência."""
    import sys

    # Caminho padrão da planilha (pode ser passado como argumento)
    arquivo = sys.argv[1] if len(sys.argv) > 1 else "Ajuizamento+2024+2+parte+ (2) -Planilha original.xlsx"

    if not Path(arquivo).exists():
        print(f"ERRO: Arquivo não encontrado: {arquivo}")
        sys.exit(1)

    cadastro = CadastroPastasBenner(arquivo)

    print("\n" + "=" * 60)
    print(" CADASTRO DE PASTAS BENNER - PREVI JURÍDICO (Python/Selenium)")
    print("=" * 60)

    # Menu interativo
    while True:
        print("""
Opções:
  1 - Etapa 1: Análise prévia de duplicidades (local)
  2 - Etapa 2: Verificar no Benner (Pastas > Parte Pasta)
  3 - Etapa 3: Cadastrar pastas (Cadastro rápido > Cível)
  4 - Gerar relatório de status
  5 - Executar tudo (Etapas 1 + 2 + 3)
  0 - Sair
""")
        opcao = input("Escolha: ").strip()

        if opcao == "1":
            cadastro.analise_previa_duplicidades()
        elif opcao == "2":
            print("\nCertifique-se de estar LOGADO no Benner antes de continuar.")
            input("Pressione Enter para continuar...")
            cadastro.verificar_no_benner()
            cadastro.fechar()
        elif opcao == "3":
            print("\nCertifique-se de estar LOGADO no Benner antes de continuar.")
            input("Pressione Enter para continuar...")
            cadastro.cadastrar_pastas()
            cadastro.fechar()
        elif opcao == "4":
            cadastro.gerar_relatorio()
        elif opcao == "5":
            cadastro.analise_previa_duplicidades()
            print("\nCertifique-se de estar LOGADO no Benner antes de continuar.")
            input("Pressione Enter para continuar...")
            cadastro.verificar_no_benner()
            cadastro.cadastrar_pastas()
            cadastro.fechar()
        elif opcao == "0":
            cadastro.fechar()
            break
        else:
            print("Opção inválida.")


if __name__ == "__main__":
    main()
