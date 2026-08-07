#!/usr/bin/env python3
"""
CADASTRO DE PASTAS NO BENNER (PREVI JURÍDICO) - Versão Python/Selenium
=======================================================================
Parecer PAR.0000871/26 - Ajuizamento dívidas prev. 2024 Parte 2

ETAPA 0 (gravação/diagnóstico manual - baseada no diagnostico_busca_global.py):
  Modo 100% MANUAL. O robô NÃO clica/digita nada no Benner: abre o navegador,
  você opera a tela manualmente e, sempre que digitar um RÓTULO (nome do passo)
  no terminal, o script grava screenshot (.png) + DOM (.html) + resumo (.json)
  daquele instante. Digite 'sair' para terminar.

FLUXO ETAPA 3 (elementos reais confirmados via console F12):
  Abrir "Cadastro rápido de pasta"
   -> PASSO 1: seletor de Categoria -> selecionar 'Cível' + OK
        (__doPostBack('ctl00$Main$TV_CADASTRORAPIDOMANUAL_FORM','Save'))
   -> PASSO 2: aguardar form real montar (integração tribunal ~28s; até 120s)
   -> PASSO 3: preencher campos por LABEL+OCORRÊNCIA (select2 AJAX, clique JS)
        PREVI = Autor (Condição 1ª); pessoa da planilha = Réu (2ª)
   -> PASSO 4: salvar (__doPostBack('ctl00$Main$WIDGET_CADASTRO_RAPIDO','Save'))
   -> capturar pst= da URL / IDENTIFICADOR

Robustez: clique de opção SEMPRE via JS (dropdown abre sobre a barra de menu
-> click nativo era interceptado). Termo de busca curto p/ AJAX. Retry.
Validação: NÃO salva se campo essencial falhar.
"""

import json
import random
import re
import time
from datetime import date, datetime
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)


class ErroBennerParalisar(Exception):
    pass


# ==============================================================================
URL_BENNER = "https://previ.bennercloud.com.br/JURIDICO/jur/e/PREVI.aspx?i=K9_INICIOPREVI&m=MAIN"

CATEGORIA = "Cível"
CAUSA_PEDIR = "Previdencial"
CAUSA_RAIZ = "Produto"
RITO = "Ordinário"
ANDAMENTO = "PEDIDO DE AJUIZAMENTO DE AÇÃO"
PEDIDO = "Dívida Previdenciária"
RISCO = "Possível"
GERENCIA_FALLBACK = "GEPAB"
COND_AUTOR = "Autor"
COND_REU = "Réu"
OBSERVACAO = "Cadastrado conforme Parecer PAR.0000871/26 - Ajuizamento dívidas previdenciárias (Ano 2024 Parte 2)"

LBL_CATEGORIA = "Categoria"
LBL_FILIAL = "Filial"
LBL_GERENCIA = "Gerência"
LBL_CAUSA_PEDIR = "Causa de Pedir"
LBL_CAUSA_RAIZ = "Causa Raiz"
LBL_RITO = "Rito"
LBL_ORGAO = "Orgão"
LBL_UF = "UF"
LBL_ANDAMENTO = "Andamento"
LBL_PARTICIPANTE = "Participante"
LBL_CONDICAO = "Condição"
LBL_ADV_INTERNO = "Advogado interno"
LBL_ADV_EXTERNO = "Advogado externo"
LBL_PEDIDO = "Pedido"
LBL_RISCO = "Risco"
LBL_OBSERVACOES = "Observações"
LBL_NUMERO = "Número"
LBL_DATA_ANDAMENTO = "Data andamento"
LBL_VALOR_PEDIDO = "Valor pedido"

# ------------------------------------------------------------------------------
# MAPA DE CAMPOS DO FORMULÁRIO REAL (PR_CADASTRORAPIDOPASTA)
# IDs confirmados via diagnóstico (Etapa 0) do HTML/JSON reais do Benner.
# Preencher por ID é MUITO mais confiável do que por label+ocorrência, porque
# o form tem vários campos com o MESMO label (Andamento x2, Participante x3,
# Condição x5, Pedido x2, Risco x2). Se o layout mudar e os IDs "ctlNN" mudarem,
# o código cai automaticamente no fallback por label.
# ------------------------------------------------------------------------------
_PREF = "ctl00_Main_WIDGET_CADASTRO_RAPIDO_PageControl_GERAL_GERAL_"
FIELD_IDS = {
    "FILIAL":        _PREF + "ctl11_ctl01_select",
    "DIRETORIA":     _PREF + "ctl18_ctl01_select",
    "GERENCIA":      _PREF + "ctl22_ctl01_select",
    "TIPO":          _PREF + "ctl30_ctl01_select",
    "CAUSA_PEDIR":   _PREF + "ctl34_ctl01_select",
    "CAUSA_RAIZ":    _PREF + "ctl43_ctl01_select",
    "RITO":          _PREF + "ctl87_ctl01_select",
    "ORGAO":         _PREF + "ctl95_ctl01_select",
    "UF":            _PREF + "ctl99_ctl01_select",
    "ANDAMENTO_1":   _PREF + "ctl122_ctl01_select",
    # Participantes: bloco RÉU (pessoa da planilha) e bloco AUTOR (PREVI)
    "PART_REU":      _PREF + "POSSUIPESSOAADVERSO_2_ctl04_ctl01_select",
    "COND_REU":      _PREF + "POSSUIPESSOAADVERSO_2_ctl13_ctl01_select",
    "PART_AUTOR":    _PREF + "ctl163_ctl01_select",
    "COND_AUTOR":    _PREF + "ctl172_ctl01_select",
    "ADV_INTERNO":   _PREF + "ctl218_ctl01_select",
    "ADV_EXTERNO":   _PREF + "ctl222_ctl01_select",
    "PEDIDO_1":      _PREF + "ctl229_ctl01_select",
    "RISCO_1":       _PREF + "ctl244_ctl01_select",
    # Campos de texto (IDs semânticos estáveis)
    "NUMERO":           _PREF + "NUMEROUNICO_1_NUMERODISTRIBUICAO",
    "DATA_ANDAMENTO_1": _PREF + "DATAANDAMENTO1_DATE",
    "VALOR_PEDIDO_1":   _PREF + "VALORPEDIDO1",
    "OBSERVACOES":      _PREF + "OBSERVACOES",
}

# ------------------------------------------------------------------------------
# MAPA POR data-fieldname (ESTÁVEL) — descoberto no diagnóstico do form real.
# Os IDs "ctlNN" MUDAM entre sessões do Benner; já o data-fieldname (e o
# data-field dos spans de texto) é ESTÁVEL. Este é o mapeamento oficial.
# ------------------------------------------------------------------------------
# SELECTS (localizados por select[data-fieldname="..."]):
FN = {
    "FILIAL": "FILIAL",
    "DIRETORIA": "DEPARTAMENTO",
    "GERENCIA": "DIVISAO",
    "TIPO": "TIPO",
    "CAUSA_PEDIR": "ASSUNTO",           # cascata: só popula após TIPO
    "CAUSA_RAIZ": "CAUSARAIZ",
    "RITO": "RITO",
    "ORGAO": "ORGAO",
    "UF": "UF",
    "ANDAMENTO_1": "EVENTO1",
    # Participantes (CONFIRMADO pelo usuário - Opção A):
    #   Bloco 1 = pessoa da planilha (Réu) ; Bloco 2 = PREVI (Autor)
    "PART_REU": "PARTICIPANTE1",
    "COND_REU": "CONDICAO1",
    "PART_AUTOR": "PARTICIPANTE2",
    "COND_AUTOR": "CONDICAO2",
    "ADV_INTERNO": "ADVOGADOINTERNO",
    "ADV_EXTERNO": "ADVOGADOEXTERNO",
    "PEDIDO_1": "PEDIDO1",
    "RISCO_1": "RISCOPEDIDO1",
}
# CAMPOS DE TEXTO (localizados por span[data-field="..."] input/textarea):
TEXT_FN = {
    "NUMERO": "NUMERODISTRIBUICAO",
    "DATA_ANDAMENTO_1": "DATAANDAMENTO1",
    "VALOR_PEDIDO_1": "VALORPEDIDO1",
    "OBSERVACOES": "OBSERVACOES",
}

# ------------------------------------------------------------------------------
# IDs INTERNOS FIXOS (via DevTools). Campos com valor IGUAL em todas as pastas.
# ------------------------------------------------------------------------------
# Padrões dos campos que o usuário definiu (Instância/Fase/Distribuição):
INSTANCIA_PADRAO = "1º Grau"
FASE_PADRAO = "Preliminar"      # confirmado pelo usuário (NÃO é "Conhecimento")

IDS_FIXOS = {
    "TIPO":         (978,  "COBRANÇA"),
    "DESDOBRAMENTO": (139, "Cobrança"),   # campo "Processo" = Cobrança
    "INSTANCIA":    (1,    "1º grau"),     # Instância = 1º Grau (id fixo)
    "ASSUNTO":      (5447, "PREVIDENCIAL"),
    "CAUSARAIZ":    (2,    "Produto"),
    "RITO":         (7,    "Ordinário"),
    "EVENTO1":      (4415, "Pedido de ajuizamento de ação"),
    "PEDIDO1":      (411,  "DÍVIDA PREVIDENCIÁRIA"),
    "RISCOPEDIDO1": (1,    "Possível"),
}
COND_ID = {"Réu": 2, "Reu": 2, "Autor": 1}
PREVI_ID = 137650
PREVI_TEXTO = "PREVI - CAIXA DE PREVIDÊNCIA DOS FUNCIONÁRIOS DO BANCO DO BRASIL"

# Cascatas conhecidas: campo dependente -> (fieldname do PAI)
CASCATA_PAI = {
    "DIVISAO": "DEPARTAMENTO",   # Gerência depende de Diretoria
}
# UF (ESTADOS) - tabela de IDs (do diagnóstico /api/search)
UF_ID = {
    "AC":8,"AL":9,"AM":10,"AP":11,"BA":12,"CE":13,"DF":14,"ES":15,"GO":16,
    "MA":17,"MG":18,"MS":19,"MT":20,"PA":21,"PB":22,"PE":23,"PI":24,"PR":3,
    "RJ":25,"RN":26,"RO":27,"RR":28,"RS":7,"SC":1,"SE":29,"SP":6,"TO":30,
}
# FILIAL (id interno) - tabela fixa (a API de FILIAL não responde à busca):
FILIAL_ID = {
    "PLANO DE BENEFICIOS 1": 3,
    # "PLANO PREVI FUTURO": <id> (se aparecer, o robô tenta a API/opções embutidas)
}

# Fallback por LABEL (usado só se o ID exato não existir mais no DOM):
_FALLBACK_LABEL = {
    "FILIAL": LBL_FILIAL, "DIRETORIA": "Diretoria", "GERENCIA": LBL_GERENCIA,
    "TIPO": "Tipo", "CAUSA_PEDIR": LBL_CAUSA_PEDIR, "CAUSA_RAIZ": LBL_CAUSA_RAIZ,
    "RITO": LBL_RITO, "ORGAO": LBL_ORGAO, "UF": LBL_UF,
    "ANDAMENTO_1": LBL_ANDAMENTO,
    "PART_REU": LBL_PARTICIPANTE, "COND_REU": LBL_CONDICAO,
    "PART_AUTOR": LBL_PARTICIPANTE, "COND_AUTOR": LBL_CONDICAO,
    "ADV_INTERNO": LBL_ADV_INTERNO, "ADV_EXTERNO": LBL_ADV_EXTERNO,
    "PEDIDO_1": LBL_PEDIDO, "RISCO_1": LBL_RISCO,
    "NUMERO": LBL_NUMERO, "DATA_ANDAMENTO_1": LBL_DATA_ANDAMENTO,
    "VALOR_PEDIDO_1": LBL_VALOR_PEDIDO, "OBSERVACOES": LBL_OBSERVACOES,
}
# Ocorrência do label no fallback (bloco Réu = 1ª; Autor = 2ª).
_FALLBACK_OCC = {
    "PART_REU": 1, "COND_REU": 1, "PART_AUTOR": 2, "COND_AUTOR": 2,
    "PEDIDO_1": 1, "RISCO_1": 1, "ANDAMENTO_1": 1,
    "DATA_ANDAMENTO_1": 1, "VALOR_PEDIDO_1": 1,
}

# Valores fixos confirmados no cadastro manual (ground truth) e pelo usuário:
TIPO = "Cobrança"           # campo "Tipo" = COBRANÇA (vale p/ todas as linhas)
PREVI_AUTOR = "PREVI"       # participante Autor (PREVI - CAIXA DE PREVIDÊNCIA...)

BTN_OK_CATEGORIA_JS = "__doPostBack('ctl00$Main$TV_CADASTRORAPIDOMANUAL_FORM','Save')"
BTN_SALVAR_JS = "__doPostBack('ctl00$Main$WIDGET_CADASTRO_RAPIDO','Save')"
ID_PASTA_HANDLE = "ctl00_Main_PASTA_VIEW_HANDLE_HiddenField"

COL_PLANO = 1
COL_NOME = 4
COL_CONTRATO = 6
COL_VALOR_DIVIDA = 15
COL_GERENCIA = 17
COL_UF = 20
COL_CPF = 23
COL_BENNER_FLAG = 28
COL_ANALISE = 30
COL_STATUS = 31
COL_CNJ = 32
COL_PLANO_DESC = 33
COL_PESQUISA_BENNER = 34
COL_ID_PASTA = 35
COL_VALOR_PEDIDO = 36

ADVOGADOS_INTERNOS = [
    "EDSON EDUARDO AGUIAR AVELAR",
    "MICHELLE CERQUEIRA NUNEZ",
    "DOMINIQUE DE SOUZA MACHADO",
]
ADVOGADOS_EXTERNOS = [
    "Aldrigues Cândido Advocacia",
    "Bicudo, Matos, e Moraes Sociedade de Advogados",
    "Dannemann Siemsen Advogados",
    "Queiroga, Vieira, Queiroz & Ramos Advocacia",
    "Wambier, Yamasaki, Bevervanço & Lobo Advocacia",
]
UF_ORGAO = {
    "AC": "do Acre", "AL": "de Alagoas", "AP": "do Amapá", "AM": "do Amazonas",
    "BA": "da Bahia", "CE": "do Ceará", "DF": "do Distrito Federal", "ES": "do Espírito Santo",
    "GO": "de Goiás", "MA": "do Maranhão", "MT": "de Mato Grosso", "MS": "de Mato Grosso do Sul",
    "MG": "de Minas Gerais", "PA": "do Pará", "PB": "da Paraíba", "PR": "do Paraná",
    "PE": "de Pernambuco", "PI": "do Piauí", "RJ": "do Rio de Janeiro",
    "RN": "do Rio Grande do Norte", "RS": "do Rio Grande do Sul", "RO": "de Rondônia",
    "RR": "de Roraima", "SC": "de Santa Catarina", "SP": "de São Paulo",
    "SE": "de Sergipe", "TO": "do Tocantins",
}

WAIT_TIMEOUT = 30
WAIT_AFTER_CLICK = 2
WAIT_BETWEEN_RECORDS = 3

# ------------------------------------------------------------------------------
# CONFIGURAÇÃO DO NAVEGADOR (ambiente corporativo travado - EDR/DLP)
# ------------------------------------------------------------------------------
# Em máquinas com antivírus/EDR corporativo, o Selenium muitas vezes NÃO consegue
# ABRIR o navegador sozinho (selenium-manager.exe é bloqueado / Chrome é morto na
# inicialização -> "DevToolsActivePort file doesn't exist"). A forma mais estável
# é você ABRIR o Edge/Chrome MANUALMENTE em modo debug e o script apenas SE
# CONECTAR a essa janela (attach via debuggerAddress).
#
# COMO USAR O MODO ATTACH (recomendado):
#   1) Feche todas as janelas do Edge.
#   2) Rode o arquivo abrir_edge_debug.bat (gerado junto deste script) OU no
#      terminal:
#        "C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe" ^
#          --remote-debugging-port=9222 ^
#          --user-data-dir="K:\BennerData\CadastraPastas\EdgeDebugProfile"
#   3) Faça o login no Benner nessa janela.
#   4) Rode este script normalmente; ele detecta a porta 9222 e se conecta.
#
# Se o selenium-manager estiver bloqueado, baixe o msedgedriver.exe da versão do
# seu Edge (edge://settings/help mostra a versão) e aponte o caminho na variável
# de ambiente BENNER_DRIVER_PATH (ou coloque msedgedriver.exe na pasta de trabalho).
# ------------------------------------------------------------------------------
import os as _os

# Navegador preferido: "edge" (padrão) ou "chrome".
BROWSER = _os.environ.get("BENNER_BROWSER", "edge").strip().lower()
# Porta de debug para o modo attach (0/"" desativa o attach automático).
DEBUG_PORT = _os.environ.get("BENNER_DEBUG_PORT", "9222").strip()
# Caminho manual do driver (msedgedriver.exe/chromedriver.exe), se o selenium
# manager estiver bloqueado pelo antivírus. Vazio = deixa o Selenium resolver.
DRIVER_PATH = _os.environ.get("BENNER_DRIVER_PATH", "").strip()
# Caminho do binário do navegador (opcional).
BROWSER_BINARY = _os.environ.get("BENNER_BROWSER_BINARY", "").strip()
# Perfil dedicado para o modo launch (evita conflito com o perfil pessoal).
USER_DATA_DIR = _os.environ.get(
    "BENNER_USER_DATA_DIR",
    r"K:\BennerData\CadastraPastas\EdgeDebugProfile")


def _parse_valor_br(txt) -> float:
    if txt is None or txt == "":
        return 0.0
    if isinstance(txt, (int, float)):
        return float(txt)
    s = str(txt).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _so_digitos(s) -> str:
    return re.sub(r"\D", "", str(s or ""))


# ==============================================================================
class CadastroPastasBenner:
    def __init__(self, arquivo_excel: str, sheet_name: str = None):
        self.arquivo_excel = Path(arquivo_excel)
        self.sheet_name = sheet_name
        self.driver = None
        self.wait = None
        self.wb = None
        self.ws = None
        self._log_path = str(self.arquivo_excel.parent /
                             f"log_execucao_{time.strftime('%Y%m%d_%H%M%S')}.txt")

    def _log(self, msg):
        linha = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
        print(linha)
        try:
            with open(self._log_path, "a", encoding="utf-8") as f:
                f.write(linha + "\n")
        except Exception:
            pass

    def iniciar_navegador(self):
        """Inicia (ou se conecta a) um navegador, resistente a bloqueios de EDR.

        Ordem de tentativa:
          1) ATTACH: se houver um navegador em 127.0.0.1:DEBUG_PORT (aberto por
             você em modo debug), o script apenas se conecta a ele. É o modo
             mais confiável em máquina corporativa.
          2) LAUNCH: tenta abrir o navegador com flags robustas
             (--no-sandbox, --disable-dev-shm-usage, --remote-debugging-port,
             --user-data-dir dedicado).
        Tenta primeiro o navegador BROWSER e depois o outro como fallback.
        """
        preferidos = ["edge", "chrome"] if BROWSER != "chrome" else ["chrome", "edge"]

        ultimo_erro = None
        # 1) ATTACH ------------------------------------------------------------
        if DEBUG_PORT and DEBUG_PORT != "0":
            for nav in preferidos:
                try:
                    self.driver = self._attach_navegador(nav, DEBUG_PORT)
                    if self.driver:
                        self.wait = WebDriverWait(self.driver, WAIT_TIMEOUT)
                        self._log(f"Conectado ao {nav.upper()} já aberto "
                                  f"(127.0.0.1:{DEBUG_PORT}).")
                        return
                except Exception as e:
                    ultimo_erro = e
                    self._log(f"    Attach ao {nav} falhou: {e}")

        # 2) LAUNCH ------------------------------------------------------------
        for nav in preferidos:
            try:
                self.driver = self._launch_navegador(nav)
                if self.driver:
                    self.wait = WebDriverWait(self.driver, WAIT_TIMEOUT)
                    self._log(f"Navegador {nav.upper()} iniciado (modo launch).")
                    return
            except Exception as e:
                ultimo_erro = e
                self._log(f"    Launch do {nav} falhou: {e}")

        raise RuntimeError(
            "Não foi possível iniciar NEM conectar a um navegador.\n"
            "Provável bloqueio do antivírus/EDR corporativo.\n"
            "SOLUÇÃO: abra o Edge MANUALMENTE em modo debug (use o arquivo\n"
            "abrir_edge_debug.bat) e rode este script de novo — ele se conecta\n"
            f"à janela aberta na porta {DEBUG_PORT}.\n"
            f"Último erro técnico: {ultimo_erro}")

    def _opcoes_navegador(self, nav):
        if nav == "chrome":
            options = webdriver.ChromeOptions()
        else:
            options = webdriver.EdgeOptions()
        if BROWSER_BINARY:
            options.binary_location = BROWSER_BINARY
        return options

    def _service_navegador(self, nav):
        """Retorna um Service com caminho manual do driver, se configurado."""
        try:
            if nav == "chrome":
                from selenium.webdriver.chrome.service import Service as Svc
            else:
                from selenium.webdriver.edge.service import Service as Svc
        except Exception:
            return None
        if DRIVER_PATH:
            return Svc(executable_path=DRIVER_PATH)
        return None

    def _attach_navegador(self, nav, porta):
        """Conecta a um navegador já aberto em 127.0.0.1:porta (modo debug)."""
        options = self._opcoes_navegador(nav)
        options.add_experimental_option("debuggerAddress", f"127.0.0.1:{porta}")
        service = self._service_navegador(nav)
        if nav == "chrome":
            return (webdriver.Chrome(service=service, options=options)
                    if service else webdriver.Chrome(options=options))
        return (webdriver.Edge(service=service, options=options)
                if service else webdriver.Edge(options=options))

    def _launch_navegador(self, nav):
        """Abre o navegador com flags robustas para ambiente travado."""
        options = self._opcoes_navegador(nav)
        for arg in (
            "--start-maximized",
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
            "--disable-extensions",
            "--remote-debugging-port=9222",
            f"--user-data-dir={USER_DATA_DIR}",
        ):
            options.add_argument(arg)
        service = self._service_navegador(nav)
        if nav == "chrome":
            return (webdriver.Chrome(service=service, options=options)
                    if service else webdriver.Chrome(options=options))
        return (webdriver.Edge(service=service, options=options)
                if service else webdriver.Edge(options=options))

    def carregar_planilha(self):
        self.wb = openpyxl.load_workbook(str(self.arquivo_excel))
        if self.sheet_name and self.sheet_name in self.wb.sheetnames:
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb[self.wb.sheetnames[0]]
        self._log(f"Planilha carregada: {self.arquivo_excel} [{self.ws.title}]")

    def salvar_planilha(self):
        self.wb.save(str(self.arquivo_excel))
        self._log("Planilha salva.")

    def fechar(self):
        if self.wb:
            self.salvar_planilha()
        if self.driver:
            self.driver.quit()
            self.driver = None
        self._log("Encerrado.")

    # ==========================================================================
    # ETAPA 0 - GRAVAÇÃO / DIAGNÓSTICO MANUAL (você nomeia cada passo)
    # ==========================================================================
    # Baseada no diagnostico_busca_global.py: modo 100% MANUAL. O script NÃO
    # clica nem digita nada na tela do Benner; apenas abre o navegador, espera
    # você operar manualmente e, sempre que você digitar um RÓTULO (nome do
    # passo) no terminal, grava o estado atual da tela:
    #   - screenshot (.png)  -> <pasta_da_planilha>/diagnostico_screenshots/
    #   - DOM completo (.html) + resumo estruturado (.json) -> .../diagnostico_dom/
    # Digite 'sair' (ou s/q/parar/exit/quit) no lugar do rótulo para terminar.
    # ==========================================================================
    def etapa_zero_gravacao_manual(self):
        self.iniciar_navegador()
        self.driver.get(URL_BENNER)
        self._aguardar_carregamento()
        self._instalar_console_recorder()
        try:
            self._habilitar_cdp_logs()
        except Exception:
            pass

        self._log("=" * 70)
        self._log("ETAPA 0 - GRAVAÇÃO/DIAGNÓSTICO MANUAL")
        self._log("Nenhuma automação de clique/digitação será feita pelo robô.")
        self._log("Faça o LOGIN e opere tudo manualmente na janela do navegador")
        self._log("(abrir a busca com Ctrl+Espaço ou na lupa, digitar o nome,")
        self._log("esperar os resultados, clicar em 'Pastas', abrir a pasta, etc.).")
        self._log("Quando quiser gravar o estado atual da tela, volte ao terminal,")
        self._log("digite um RÓTULO curto (ex.: 'antes', 'modal_aberto', 'resultados',")
        self._log("'pasta_aberta') e pressione ENTER. Digite 'sair' para terminar.")
        self._log("=" * 70)

        contador = 1
        try:
            while True:
                try:
                    rotulo = input(
                        f"\n>>> [{contador:02d}] Rótulo para gravar agora "
                        f"(ou 'sair' para terminar): "
                    ).strip()
                except EOFError:
                    self._log("Entrada não interativa (EOF). Encerrando gravação.")
                    break

                if rotulo.lower() in ("sair", "s", "q", "parar", "exit", "quit"):
                    self._log("Encerrando a Etapa 0 a pedido do usuário.")
                    break

                if not rotulo:
                    rotulo = f"passo_{contador:02d}"

                self._gravar_screenshot_passo(rotulo)
                self._gravar_dom_passo(rotulo)
                contador += 1
        except KeyboardInterrupt:
            self._log("ETAPA 0 ENCERRADA (Ctrl+C).")

        self._log("=" * 70)
        self._log("FIM DA ETAPA 0. Nenhum registro foi criado/alterado no Benner.")
        self._log(f"Screenshots: {self.arquivo_excel.parent / 'diagnostico_screenshots'}")
        self._log(f"DOM/JSON:    {self.arquivo_excel.parent / 'diagnostico_dom'}")
        self._log("=" * 70)

    def _instalar_console_recorder(self):
        """Injeta um gravador de console/erros no navegador (funciona no modo
        attach, sem precisar de capabilities). Guarda em window.__diag."""
        js = r"""
            (function(){
              if(window.__diag && window.__diag.__installed) return 'ja';
              window.__diag = {logs:[], errors:[], __installed:true};
              var push=function(kind,args){
                try{
                  var parts=[];
                  for(var i=0;i<args.length;i++){
                    var a=args[i];
                    try{ parts.push(typeof a==='object'?JSON.stringify(a):String(a)); }
                    catch(e){ parts.push(String(a)); }
                  }
                  window.__diag.logs.push({t:Date.now(),kind:kind,msg:parts.join(' ')});
                  if(window.__diag.logs.length>1000) window.__diag.logs.shift();
                }catch(e){}
              };
              ['log','info','warn','error','debug'].forEach(function(k){
                var orig=console[k];
                console[k]=function(){ push(k, arguments); try{orig.apply(console,arguments);}catch(e){} };
              });
              window.addEventListener('error', function(ev){
                try{ window.__diag.errors.push({t:Date.now(),
                     msg:(ev.message||'')+' @ '+(ev.filename||'')+':'+(ev.lineno||'')}); }catch(e){}
              });
              window.addEventListener('unhandledrejection', function(ev){
                try{ window.__diag.errors.push({t:Date.now(),
                     msg:'unhandledrejection: '+(ev.reason&&ev.reason.message?ev.reason.message:String(ev.reason))}); }catch(e){}
              });
              return 'ok';
            })();
        """
        try:
            r = self.driver.execute_script(js)
            self._log(f"    [DevTools] console recorder: {r}")
        except Exception as e:
            self._log(f"    [DevTools] falha ao instalar recorder: {e}")

    def _habilitar_cdp_logs(self):
        """Best-effort: habilita coleta de rede via CDP (Edge/Chromium)."""
        try:
            self.driver.execute_cdp_cmd("Network.enable", {})
            self._log("    [DevTools] CDP Network.enable OK.")
        except Exception as e:
            self._log(f"    [DevTools] CDP indisponível: {e}")

    def _coletar_devtools(self):
        """Coleta console/erros (window.__diag), rede (performance resource
        timing) e, se possível, logs do driver. Retorna um dict."""
        dados = {"console": [], "errors": [], "network": [], "browser_log": []}
        # reinstalar recorder (pode ter havido navegação/postback que resetou a página)
        self._instalar_console_recorder()
        try:
            diag = self.driver.execute_script("return window.__diag ? window.__diag : null;")
            if diag:
                dados["console"] = diag.get("logs", [])[-300:]
                dados["errors"] = diag.get("errors", [])[-100:]
        except Exception:
            pass
        # rede via Performance API (URLs, tipo, duração, tamanho)
        try:
            dados["network"] = self.driver.execute_script("""
                try{
                  return performance.getEntriesByType('resource').slice(-120).map(function(e){
                    return {name:e.name, type:e.initiatorType,
                            dur:Math.round(e.duration),
                            size:(e.transferSize||0)};
                  });
                }catch(err){ return []; }
            """) or []
        except Exception:
            pass
        # logs do driver (só funciona se logging estiver habilitado)
        for tipo in ("browser", "performance"):
            try:
                dados["browser_log"] += [
                    {"level": l.get("level"), "message": (l.get("message") or "")[:500]}
                    for l in self.driver.get_log(tipo)[-100:]
                ]
            except Exception:
                pass
        return dados

    def _gravar_screenshot_passo(self, nome_passo):
        try:
            pasta = self.arquivo_excel.parent / "diagnostico_screenshots"
            pasta.mkdir(parents=True, exist_ok=True)
            nome_seguro = re.sub(r"[^A-Za-z0-9_-]+", "_", nome_passo).strip("_") or "passo"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            caminho = pasta / f"{timestamp}_{nome_seguro}.png"
            self.driver.save_screenshot(str(caminho))
            self._log(f"    [GRAVAÇÃO] Screenshot salvo: {caminho.name}")
            return caminho
        except Exception as e:
            self._log(f"    [GRAVAÇÃO] Falha ao salvar screenshot ('{nome_passo}'): {e}")
            return None

    def _gravar_dom_passo(self, nome_passo):
        """Grava o DOM (HTML completo) da tela atual e um resumo em JSON de
        todos os campos visíveis (input, select, textarea, label, button),
        incluindo opções de <select> e os painéis de resultado da busca global
        (#searcher-results-items / #searcher-results-entities, se existirem)."""
        try:
            pasta = self.arquivo_excel.parent / "diagnostico_dom"
            pasta.mkdir(parents=True, exist_ok=True)
            nome_seguro = re.sub(r"[^A-Za-z0-9_-]+", "_", nome_passo).strip("_") or "passo"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            caminho_html = pasta / f"{timestamp}_{nome_seguro}.html"
            try:
                caminho_html.write_text(self.driver.page_source, encoding="utf-8")
            except Exception as e:
                self._log(f"    [DOM] Falha ao capturar HTML ('{nome_passo}'): {e}")
                caminho_html = None

            resumo = []
            for tipo in ("input", "select", "textarea", "label", "button"):
                try:
                    elementos = self.driver.find_elements(By.TAG_NAME, tipo)
                except Exception:
                    continue
                for el in elementos:
                    try:
                        item = {
                            "tag": tipo,
                            "id": el.get_attribute("id"),
                            "name": el.get_attribute("name"),
                            "class": el.get_attribute("class"),
                            "type": el.get_attribute("type"),
                            "texto": (el.text or "").strip()[:150],
                            "value": (el.get_attribute("value") or "")[:150],
                            "placeholder": el.get_attribute("placeholder"),
                            "aria_label": el.get_attribute("aria-label"),
                            "visivel": el.is_displayed(),
                            "habilitado": el.is_enabled(),
                        }
                        if tipo == "select":
                            try:
                                item["opcoes"] = [
                                    {"value": o.get_attribute("value"), "texto": o.text}
                                    for o in el.find_elements(By.TAG_NAME, "option")
                                ]
                            except Exception:
                                item["opcoes"] = []
                        if item["id"] or item["name"] or item["texto"] or \
                                item["placeholder"] or item["class"]:
                            resumo.append(item)
                    except StaleElementReferenceException:
                        continue
                    except Exception:
                        continue

            # Elementos li/a/div cuja classe pareça ser de resultado de busca.
            resultados_busca = []
            for tipo in ("li", "a", "div"):
                try:
                    elementos = self.driver.find_elements(By.TAG_NAME, tipo)
                except Exception:
                    continue
                for el in elementos:
                    try:
                        classe = el.get_attribute("class") or ""
                        texto = (el.text or "").strip()
                        if not texto or len(texto) > 200:
                            continue
                        classe_norm = classe.lower()
                        if any(t in classe_norm for t in
                               ("search", "result", "busca", "pasta", "pesquisa")):
                            resultados_busca.append({
                                "tag": tipo,
                                "id": el.get_attribute("id"),
                                "class": classe,
                                "texto": texto[:200],
                                "href": el.get_attribute("href") if tipo == "a" else None,
                                "visivel": el.is_displayed(),
                            })
                    except StaleElementReferenceException:
                        continue
                    except Exception:
                        continue

            # Painéis de resultado da busca global (se abertos na tela).
            paineis_busca = {}
            for painel_id in ("searcher-results-items", "searcher-results-entities"):
                try:
                    el = self.driver.find_element(By.ID, painel_id)
                    paineis_busca[painel_id] = {
                        "innerHTML": self.driver.execute_script(
                            "return arguments[0].innerHTML;", el),
                        "texto": el.text,
                        "visivel": el.is_displayed(),
                    }
                except Exception:
                    pass  # painel pode não existir se a busca não estiver aberta

            caminho_json = pasta / f"{timestamp}_{nome_seguro}.json"
            try:
                caminho_json.write_text(
                    json.dumps({
                        "campos_formulario": resumo,
                        "possiveis_resultados_busca": resultados_busca,
                        "paineis_busca_global": paineis_busca,
                        "url_atual": self.driver.current_url,
                    }, ensure_ascii=False, indent=2),
                    encoding="utf-8")
            except Exception as e:
                self._log(f"    [DOM] Falha ao salvar resumo JSON ('{nome_passo}'): {e}")
                caminho_json = None

            # DevTools: console + erros + rede (o que você pediu)
            caminho_dev = pasta / f"{timestamp}_{nome_seguro}_devtools.json"
            try:
                dev = self._coletar_devtools()
                caminho_dev.write_text(
                    json.dumps(dev, ensure_ascii=False, indent=2), encoding="utf-8")
                dev_resumo = (f"console={len(dev.get('console',[]))} "
                              f"erros={len(dev.get('errors',[]))} "
                              f"rede={len(dev.get('network',[]))}")
            except Exception as e:
                caminho_dev = None
                dev_resumo = f"devtools=falhou({e})"

            self._log(
                f"    [DOM] '{nome_passo}': "
                f"{('HTML=' + caminho_html.name) if caminho_html else 'HTML=falhou'} | "
                f"{('JSON=' + caminho_json.name) if caminho_json else 'JSON=falhou'} | "
                f"{('DEV=' + caminho_dev.name) if caminho_dev else 'DEV=falhou'} "
                f"({len(resumo)} campos, {len(resultados_busca)} result., "
                f"{len(paineis_busca)} painel(is); {dev_resumo})")
            return caminho_html, caminho_json
        except Exception as e:
            self._log(f"    [DOM] Falha geral ao gravar DOM ('{nome_passo}'): {e}")
            return None, None

    # ==========================================================================
    # ETAPA 1
    # ==========================================================================
    def analise_previa_duplicidades(self):
        self.carregar_planilha()
        ws = self.ws
        last_row = ws.max_row

        ws.cell(1, COL_ANALISE, "ANÁLISE DUPLICIDADE / PESQUISA BENNER (Pasta)")
        ws.cell(1, COL_STATUS, "STATUS CADASTRO")
        ws.cell(1, COL_CNJ, "NÚMERO CNJ")
        ws.cell(1, COL_PLANO_DESC, "PLANO DESCRIÇÃO (Filial)")
        ws.cell(1, COL_PESQUISA_BENNER, "PESQUISA BENNER")
        ws.cell(1, COL_ID_PASTA, "ID PASTA BENNER")
        ws.cell(1, COL_VALOR_PEDIDO, "VALOR PEDIDO")

        for row in range(2, last_row + 1):
            for col in (COL_ANALISE, COL_STATUS, COL_CNJ, COL_PLANO_DESC,
                        COL_PESQUISA_BENNER, COL_ID_PASTA, COL_VALOR_PEDIDO):
                ws.cell(row, col, None)

        def _linha_vazia(r):
            return not str(ws.cell(r, COL_NOME).value or "").strip()

        participante_linhas = {}
        for row in range(2, last_row + 1):
            if _linha_vazia(row):
                continue
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            participante_linhas.setdefault(nome, []).append(row)

        chaves_vistas = set()
        duplicatas_exatas = set()
        for row in range(2, last_row + 1):
            if _linha_vazia(row):
                continue
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            contrato = str(ws.cell(row, COL_CONTRATO).value or "")
            valor = _parse_valor_br(ws.cell(row, COL_VALOR_DIVIDA).value)
            chave = f"{nome}|{contrato}|{valor}"
            if chave in chaves_vistas:
                duplicatas_exatas.add(row)
            chaves_vistas.add(chave)

        total = 0
        for row in range(2, last_row + 1):
            if _linha_vazia(row):
                continue
            total += 1
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            contrato = str(ws.cell(row, COL_CONTRATO).value or "")

            plano = ws.cell(row, COL_PLANO).value
            ws.cell(row, COL_PLANO_DESC,
                    "PLANO PREVI FUTURO" if str(plano).strip() == "2" else "PLANO DE BENEFICIOS 1")

            benner_flag = str(ws.cell(row, COL_BENNER_FLAG).value or "").strip()

            if row in duplicatas_exatas:
                ws.cell(row, COL_ANALISE, "DUPLICATA EXATA - REMOVER")
                ws.cell(row, COL_STATUS, "NÃO CADASTRAR")
                ws.cell(row, COL_CNJ, f"DP{contrato}")
                continue

            if benner_flag:
                ws.cell(row, COL_ANALISE, f"ANÁLISE MANUAL: {benner_flag}")

            linhas_participante = participante_linhas.get(nome, [])
            linhas_validas = [r for r in linhas_participante if r not in duplicatas_exatas]

            if len(linhas_validas) > 1:
                contratos = [str(ws.cell(r, COL_CONTRATO).value or "") for r in linhas_validas]
                contratos_unicos = list(dict.fromkeys(contratos))
                numero_combinado = "DP" + "/".join(contratos_unicos)
                valor_total = sum(_parse_valor_br(ws.cell(r, COL_VALOR_DIVIDA).value) for r in linhas_validas)

                if row == linhas_validas[0]:
                    ws.cell(row, COL_ANALISE,
                            (str(ws.cell(row, COL_ANALISE).value or "") +
                             f" | MESMO PARTICIPANTE - {len(linhas_validas)} OPERAÇÕES (AGRUPADO)").strip(" |"))
                    ws.cell(row, COL_STATUS, "PENDENTE")
                    ws.cell(row, COL_CNJ, numero_combinado)
                    ws.cell(row, COL_VALOR_PEDIDO, round(valor_total, 2))
                else:
                    ws.cell(row, COL_ANALISE,
                            (str(ws.cell(row, COL_ANALISE).value or "") +
                             f" | AGRUPADO COM LINHA {linhas_validas[0]} - PASTA ÚNICA").strip(" |"))
                    ws.cell(row, COL_STATUS, "AGRUPADO")
                    ws.cell(row, COL_CNJ, numero_combinado)
            else:
                if not ws.cell(row, COL_ANALISE).value:
                    ws.cell(row, COL_ANALISE, "OK")
                ws.cell(row, COL_STATUS, "PENDENTE")
                ws.cell(row, COL_CNJ, f"DP{contrato}")
                ws.cell(row, COL_VALOR_PEDIDO,
                        round(_parse_valor_br(ws.cell(row, COL_VALOR_DIVIDA).value), 2))

        self.salvar_planilha()
        self._log(f"Etapa 1 concluída. Operações válidas: {total}.")
        return total

    # ==========================================================================
    # ETAPA 2
    # ==========================================================================
    def verificar_no_benner(self):
        self.carregar_planilha()
        ws = self.ws
        if str(ws.cell(1, COL_STATUS).value or "") != "STATUS CADASTRO":
            self._log("ERRO: Execute primeiro a Etapa 1!")
            return

        self.iniciar_navegador()
        self.driver.get(URL_BENNER)
        self._aguardar_carregamento()
        input(">>> Faça LOGIN no Benner e pressione ENTER para continuar...")
        self._log("=== INÍCIO ETAPA 2 ===")

        last_row = ws.max_row
        pesquisados = ja_existentes = 0
        try:
            for row in range(2, last_row + 1):
                status = str(ws.cell(row, COL_STATUS).value or "").strip().upper()
                if status not in ("PENDENTE", "VERIFICAR"):
                    continue
                nome = str(ws.cell(row, COL_NOME).value or "").strip()
                cpf = self._formatar_cpf(str(ws.cell(row, COL_CPF).value or ""))
                if not nome:
                    continue

                resultado = self._pesquisar_pessoa_por_nome(nome, cpf)
                if resultado.upper().startswith("ERRO"):
                    shot = self._screenshot_erro("erro_etapa2")
                    self.salvar_planilha()
                    raise ErroBennerParalisar(
                        f"Falha na linha {row} ({nome}, CPF {cpf}).\nDetalhe: {resultado}\n"
                        f"Screenshot: {shot or 'não gerado'}\nLog: {self._log_path}")

                ws.cell(row, COL_PESQUISA_BENNER, resultado)
                analise_atual = str(ws.cell(row, COL_ANALISE).value or "")
                if resultado and resultado not in analise_atual:
                    ws.cell(row, COL_ANALISE, f"{analise_atual} | {resultado}".strip(" |"))
                pesquisados += 1

                up = resultado.upper()
                if "ENCONTRADA" in up and "NÃO ENCONTRADA" not in up:
                    m = re.search(r"PASTA:(\S+)", resultado)
                    if m:
                        idp = m.group(1).strip("]").rstrip(";,")
                        ws.cell(row, COL_ID_PASTA, idp)
                        if "DÍVIDA PREVIDENCIÁRIA" in up or "DIVIDA PREVIDENCIARIA" in up:
                            ws.cell(row, COL_STATUS, f"JÁ CADASTRADO NO BENNER (PASTA:{idp})")
                            ja_existentes += 1

                self._log(f"  Pesquisando... {pesquisados} (linha {row}: {nome}) - {resultado}")
                time.sleep(WAIT_AFTER_CLICK)
        except ErroBennerParalisar as e:
            self._exibir_erro_e_parar(str(e))
            return

        self.salvar_planilha()
        self._log(f"Etapa 2 concluída. Pesquisados: {pesquisados}, Já cadastrados: {ja_existentes}")

    def _pesquisar_pessoa_por_nome(self, nome, cpf=""):
        try:
            if not self._abrir_searcher():
                return "ERRO: Não foi possível abrir a pesquisa (lupa)"
            campo = self._esperar_visivel(By.CSS_SELECTOR, "input.searcher-criteria", 10)
            if not campo:
                return "ERRO: Campo de pesquisa não encontrado"
            self._preencher_input_seguro(campo, nome)
            campo.send_keys(Keys.ENTER)
            time.sleep(3)
            btn = self._botao_pastas_da_pessoa(nome, cpf)
            if not btn:
                corpo = self.driver.find_element(By.TAG_NAME, "body").text.upper()
                if nome.upper() not in corpo:
                    return "PESSOA NÃO ENCONTRADA - OK para cadastrar"
                return "PESSOA ENCONTRADA - SEM PASTAS (OK para cadastrar)"
            self._click_seguro(btn)
            self._aguardar_carregamento()
            time.sleep(2)
            return self._checar_pastas_da_pessoa(nome)
        except Exception as e:
            return f"ERRO: {e}"

    def _abrir_searcher(self):
        try:
            for e in self.driver.find_elements(By.CSS_SELECTOR, "input.searcher-criteria"):
                if e.is_displayed():
                    return True
        except Exception:
            pass
        for xp in ("//li[@id='searcher']/a", "//li[@id='searcher']"):
            try:
                el = self.driver.find_element(By.XPATH, xp)
                self._click_seguro(el)
                time.sleep(1.5)
                for e in self.driver.find_elements(By.CSS_SELECTOR, "input.searcher-criteria"):
                    if e.is_displayed():
                        return True
            except Exception:
                continue
        return False

    def _botao_pastas_da_pessoa(self, nome, cpf=""):
        cpf_d = _so_digitos(cpf)
        try:
            cands = self.driver.find_elements(
                By.XPATH, "//a[contains(@class,'blue-hoki') and normalize-space()='Pastas']")
        except Exception:
            cands = []
        if not cands:
            try:
                todos = self.driver.find_elements(By.XPATH, "//a[normalize-space()='Pastas']")
                cands = [c for c in todos if (c.get_attribute("href") or "") != "javascript:;" and c.is_displayed()]
            except Exception:
                cands = []
        if cpf_d:
            for c in cands:
                try:
                    b = c
                    for _ in range(8):
                        b = b.find_element(By.XPATH, "..")
                        if cpf_d in _so_digitos(b.text):
                            return c
                except Exception:
                    continue
        for c in cands:
            try:
                b = c
                for _ in range(8):
                    b = b.find_element(By.XPATH, "..")
                    if nome.upper() in (b.text or "").upper():
                        return c
            except Exception:
                continue
        for c in cands:
            try:
                if c.is_displayed():
                    return c
            except Exception:
                continue
        return None

    def _checar_pastas_da_pessoa(self, nome):
        try:
            self._esperar_visivel(By.CSS_SELECTOR, "table.simple-grid", 10)
        except Exception:
            pass
        pastas = []
        try:
            linhas = self.driver.find_elements(
                By.XPATH, "//table[contains(@class,'simple-grid')]//tbody/tr[@handle]")
            for tr in linhas:
                ident = target = arg = ""
                try:
                    td = tr.find_element(By.CSS_SELECTOR, "td[data-field='IDENTIFICADOR']")
                    ident = (td.text or "").strip()
                    a = td.find_element(By.XPATH, ".//a")
                    oc = a.get_attribute("onclick") or ""
                    m = re.search(r"__doPostBack\('([^']+)'\s*,\s*'(ViewEntity\$[^']+)'\)", oc)
                    if m:
                        target, arg = m.group(1), m.group(2)
                except Exception:
                    pass
                if ident and target and arg:
                    pastas.append((ident, target, arg))
        except Exception:
            pass
        if not pastas:
            corpo = self.driver.find_element(By.TAG_NAME, "body").text.upper()
            if "NENHUM REGISTRO" in corpo:
                return "PESSOA ENCONTRADA - SEM PASTAS (OK para cadastrar)"
            return "PESSOA ENCONTRADA - PASTAS NÃO LIDAS (verificar manualmente)"
        vistos = set(); unicas = []
        for i, t, a in pastas:
            if i not in vistos:
                vistos.add(i); unicas.append((i, t, a))
        self._log(f"    Pastas de {nome}: {[p[0] for p in unicas]}")
        grid_url = self.driver.current_url
        com_divida = []
        for ident, target, arg in unicas:
            if self._abrir_resumo_por_postback(ident, target, arg):
                com_divida.append(ident)
            try:
                self.driver.get(grid_url)
                self._aguardar_carregamento()
                self._esperar_visivel(By.CSS_SELECTOR, "table.simple-grid", 8)
            except Exception:
                pass
        if com_divida:
            return (f"ENCONTRADA - MESMO OBJETO (DÍVIDA PREVIDENCIÁRIA) "
                    f"| PASTA:{com_divida[0]} | TODAS:[{'; '.join(com_divida)}]")
        return f"PESSOA ENCONTRADA - PASTAS SEM DÍVIDA PREVIDENCIÁRIA | PASTAS:[{'; '.join(p[0] for p in unicas)}]"

    def _abrir_resumo_por_postback(self, identificador, target, arg):
        try:
            arg_js = arg.replace("\\", "\\\\").replace("'", "\\'")
            tgt_js = target.replace("\\", "\\\\").replace("'", "\\'")
            self.driver.execute_script(f"__doPostBack('{tgt_js}','{arg_js}');")
            self._aguardar_carregamento()
            time.sleep(2)
            achou = self._pasta_tem_divida_previdenciaria()
            self._log(f"    Pasta {identificador}: "
                      f"{'DÍVIDA PREVIDENCIÁRIA ENCONTRADA' if achou else 'sem Dívida Previdenciária'}.")
            return achou
        except Exception as e:
            self._log(f"    Pasta {identificador}: erro: {e}")
            return False

    def _pasta_tem_divida_previdenciaria(self):
        try:
            self._esperar_visivel(By.CSS_SELECTOR, "td[data-field='PEDIDO']", 6)
        except Exception:
            pass
        try:
            for c in self.driver.find_elements(By.CSS_SELECTOR, "td[data-field='PEDIDO']"):
                t = (c.text or "").upper()
                if "DÍVIDA PREVIDENCIÁRIA" in t or "DIVIDA PREVIDENCIARIA" in t:
                    return True
        except Exception:
            pass
        try:
            obs = self.driver.find_element(By.XPATH, "//span[@data-field='OBSERVACOES']")
            t = (obs.text or "").upper()
            if "DÍVIDA PREVIDENCIÁRIA" in t or "DIVIDA PREVIDENCIARIA" in t:
                return True
        except Exception:
            pass
        return False

    def _esperar_visivel(self, by, seletor, timeout=10):
        try:
            return WebDriverWait(self.driver, timeout).until(
                EC.visibility_of_element_located((by, seletor)))
        except TimeoutException:
            return None

    def _click_seguro(self, elem):
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
            time.sleep(0.3)
        except Exception:
            pass
        try:
            elem.click(); return
        except Exception:
            pass
        try:
            self.driver.execute_script("arguments[0].click();", elem); return
        except Exception:
            pass
        try:
            attr = elem.get_attribute("onclick") or elem.get_attribute("href") or ""
            m = re.search(r"__doPostBack\('([^']+)'\s*,\s*'([^']*)'\)", attr)
            if m:
                self.driver.execute_script(f"__doPostBack('{m.group(1)}','{m.group(2)}');")
        except Exception:
            pass

    def _preencher_input_seguro(self, elem, valor):
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
            time.sleep(0.3)
            try:
                elem.click()
            except Exception:
                self.driver.execute_script("arguments[0].focus();", elem)
            elem.clear()
            elem.send_keys(valor)
        except Exception:
            self.driver.execute_script("""
                var e=arguments[0],v=arguments[1];e.value='';e.value=v;
                e.dispatchEvent(new Event('input',{bubbles:true}));
                e.dispatchEvent(new Event('change',{bubbles:true}));""", elem, valor)

    # ==========================================================================
    # ETAPA 3
    # ==========================================================================
    def cadastrar_pastas(self):
        self.carregar_planilha()
        ws = self.ws
        if str(ws.cell(1, COL_STATUS).value or "") != "STATUS CADASTRO":
            self._log("ERRO: Execute primeiro as Etapas 1 e 2!")
            return

        last_row = ws.max_row
        pendentes = [r for r in range(2, last_row + 1)
                     if str(ws.cell(r, COL_STATUS).value or "").strip().upper() == "PENDENTE"]
        if not pendentes:
            self._log("Nenhuma operação PENDENTE.")
            return

        self._log(f"Cadastrando {len(pendentes)} pastas...")
        self.iniciar_navegador()
        self.driver.get(URL_BENNER)
        self._aguardar_carregamento()
        input(">>> Faça LOGIN no Benner e pressione ENTER para continuar...")
        self._log("=== INÍCIO ETAPA 3 ===")
        # Instalar gravador de console/erros + hook de rede (fetch/XHR) para que a
        # DEPURAÇÃO PROFUNDA pós-save capture o POST do Salvar e erros JS.
        try:
            self._instalar_console_recorder()
            self._habilitar_cdp_logs()
        except Exception:
            pass

        cadastrados = erros = 0
        escritorio_por_participante = {}

        try:
            for row in pendentes:
                nome = str(ws.cell(row, COL_NOME).value or "").strip()
                contrato = str(ws.cell(row, COL_CONTRATO).value or "")
                valor_pedido = ws.cell(row, COL_VALOR_PEDIDO).value
                valor_pedido = _parse_valor_br(valor_pedido) if valor_pedido not in (None, "") \
                    else _parse_valor_br(ws.cell(row, COL_VALOR_DIVIDA).value)
                uf = str(ws.cell(row, COL_UF).value or "").strip().upper()
                filial = str(ws.cell(row, COL_PLANO_DESC).value or "PLANO DE BENEFICIOS 1")
                numero_cnj = str(ws.cell(row, COL_CNJ).value or f"DP{contrato}")
                gerencia = str(ws.cell(row, COL_GERENCIA).value or "").strip() or GERENCIA_FALLBACK

                adv_interno = random.choice(ADVOGADOS_INTERNOS)
                nome_upper = nome.upper()
                if nome_upper in escritorio_por_participante:
                    adv_externo = escritorio_por_participante[nome_upper]
                else:
                    adv_externo = random.choice(ADVOGADOS_EXTERNOS)
                    escritorio_por_participante[nome_upper] = adv_externo

                resultado = self._cadastrar_pasta_civel(
                    nome, valor_pedido, uf, filial, numero_cnj,
                    adv_interno, adv_externo, gerencia)

                if resultado.startswith("OK"):
                    ws.cell(row, COL_STATUS, "CADASTRADO + ANDAMENTO")
                    id_pasta = resultado.split("|", 1)[1] if "|" in resultado else ""
                    if id_pasta:
                        ws.cell(row, COL_ID_PASTA, id_pasta)
                    cadastrados += 1
                    for r2 in range(2, last_row + 1):
                        if r2 == row:
                            continue
                        if str(ws.cell(r2, COL_STATUS).value or "").strip().upper() == "AGRUPADO" \
                                and str(ws.cell(r2, COL_NOME).value or "").strip().upper() == nome_upper:
                            ws.cell(r2, COL_STATUS, "CADASTRADO (AGRUPADO)")
                            if id_pasta:
                                ws.cell(r2, COL_ID_PASTA, id_pasta)
                    self._log(f"  Cadastrada (linha {row}: {nome}) -> {resultado}")
                else:
                    shot = self._screenshot_erro("erro_etapa3")
                    ws.cell(row, COL_STATUS, f"ERRO: {resultado}")
                    self.salvar_planilha()
                    raise ErroBennerParalisar(
                        f"Falha ao CADASTRAR na linha {row} ({nome}).\nDetalhe: {resultado}\n"
                        f"Screenshot: {shot or 'não gerado'}\nLog: {self._log_path}\n"
                        f"Cadastradas até aqui: {cadastrados}.")

                self.salvar_planilha()
                self._log(f"  Progresso... {cadastrados + erros}/{len(pendentes)} (OK:{cadastrados} Err:{erros})")
                self._fechar_janelas_extras()  # fecha a aba do formulário e volta à principal
                time.sleep(WAIT_BETWEEN_RECORDS)
        except ErroBennerParalisar as e:
            self._exibir_erro_e_parar(str(e))
            return

        self._log(f"Concluído! Sucesso: {cadastrados}, Erros: {erros}")

    def _cadastrar_pasta_civel(self, nome, valor_pedido, uf, filial,
                               numero_cnj, adv_interno, adv_externo, gerencia) -> str:
        try:
            # Lembrar a janela PRINCIPAL (Benner) na primeira vez.
            if not getattr(self, "_janela_principal", None):
                try:
                    self._janela_principal = self.driver.current_window_handle
                except Exception:
                    self._janela_principal = None
            try:
                if getattr(self, "_janela_principal", None):
                    self.driver.switch_to.window(self._janela_principal)
            except Exception:
                pass
            self.driver.switch_to.default_content()

            # Abrir "Cadastro rápido de pasta"
            try:
                self.driver.find_element(By.ID, "sidebar_novoItem").click()
                time.sleep(1)
            except Exception:
                pass
            link = self._buscar_elemento_por_texto("span", "Cadastro rápido de pasta") or \
                   self._buscar_elemento_por_texto("a", "Cadastro rápido de pasta")
            if link:
                self._click_seguro(link)
            self._aguardar_carregamento()
            time.sleep(2)

            # O seletor de categoria pode abrir em NOVA ABA -> trocar para ela.
            self._trocar_para_janela_do_formulario(timeout=30)

            # PASSO 1: seletor de categoria (Cível + OK)
            if not self._selecionar_categoria_civel():
                return "FALHA no seletor de categoria (Cível + OK)."

            # Após o OK da categoria, o FORM REAL (PR_CADASTRORAPIDOPASTA, ~33
            # campos) abre - possivelmente em OUTRA ABA. Ir para a aba do form real.
            self._ir_para_janela_form_real(timeout=120)

            # PASSO 2: aguardar form real (integração ~28s)
            if not self._aguardar_formulario_carregar(timeout=120, min_selects=10):
                return ("FORMULARIO NAO CARREGOU COMPLETAMENTE (poucos selects). "
                        "PARANDO para não salvar vazio.")

            # PASSO 3: preencher (SOLUÇÃO DEFINITIVA via API /api/search)
            self._estabilizar_form(timeout=30, quieto=2.5)
            # Reinstalar o recorder (o postback da categoria recriou a página):
            try:
                self._instalar_console_recorder()
            except Exception:
                pass
            # >>> Instalar o HOOK DE REDE (fetch/XHR) no form real <<<
            try:
                self._instalar_hook_rede()
            except Exception:
                pass
            ok = {}
            # ---- Variáveis via API (determinístico): Diretoria -> Gerência ----
            # Gerência depende de Diretoria (cascata). Descobrir a Diretoria:
            depend_ger = None
            try:
                it_dir = self._buscar_item_api("DEPARTAMENTO", "Diretoria de Seguridade")
                if it_dir and it_dir.get("id") is not None:
                    self._set_valor_por_id("DEPARTAMENTO", it_dir["id"],
                                           it_dir.get("text", "Diretoria de Seguridade"))
                    depend_ger = f"DEPARTAMENTO={it_dir['id']}"
                    self._log(f"    [cascata] Diretoria id={it_dir['id']} -> depend p/ Gerência.")
                else:
                    self._log("    [cascata] Diretoria não encontrada via API.")
            except Exception as e:
                self._log(f"    [cascata] Diretoria falhou: {e}")
            # FILIAL: usar tabela de IDs (a API de FILIAL não responde à busca).
            # Fallbacks: opção embutida no <select> -> API.
            if filial in FILIAL_ID:
                ok["Filial"] = self._set_valor_por_id(FN["FILIAL"], FILIAL_ID[filial], filial)
            else:
                ok["Filial"] = self._selecionar_por_opcao_embutida(FN["FILIAL"], filial) \
                    or self._selecionar_via_api(FN["FILIAL"], filial)
            ok["Gerência"] = self._selecionar_via_api(FN["GERENCIA"], gerencia, depend=depend_ger)

            # ---- Fixos determinísticos ----
            ok["Tipo"] = self._selecionar_fixo("TIPO")
            time.sleep(0.8)
            ok["Causa Pedir"] = self._selecionar_fixo("ASSUNTO")
            self._selecionar_fixo("CAUSARAIZ")
            self._preencher_texto_por_fieldname(TEXT_FN["OBSERVACOES"], OBSERVACAO)
            ok["Rito"] = self._selecionar_fixo("RITO")
            self._selecionar_fixo("DESDOBRAMENTO")   # campo "Processo" = Cobrança (139)

            # ---- Instância = 1º Grau (fixo) ; Fase = Preliminar (via API) ----
            self._selecionar_fixo("INSTANCIA")               # 1º grau (id 1)
            self._selecionar_via_api("FASE", FASE_PADRAO)    # Preliminar

            # ---- "Já distribuído judicialmente" = Não  (desobriga Juízo/Data) ----
            # Confirmado pelo usuário: marcar NÃO remove a obrigatoriedade de
            # Juízo/Data de distribuição. Tentamos o grupo por nomes prováveis e,
            # como reforço, preenchemos a Data com hoje caso o campo persista.
            self._marcar_ja_distribuido_nao()
            try:
                self._preencher_texto_por_fieldname(
                    "DATADISTRIBUICAO", date.today().strftime("%d/%m/%Y"))
            except Exception:
                pass

            # ---- Variáveis via API: Órgão e UF ----
            orgao_txt = f"Tribunal de Justiça do Estado {UF_ORGAO.get(uf, '')}".strip()
            ok["Orgão"] = self._selecionar_via_api(FN["ORGAO"], orgao_txt)
            # UF: usar tabela fixa (id direto) e cair p/ API se faltar
            if uf in UF_ID:
                ok["UF"] = self._set_valor_por_id(FN["UF"], UF_ID[uf], uf)
            else:
                ok["UF"] = self._selecionar_via_api(FN["UF"], uf)

            self._preencher_texto_por_fieldname(TEXT_FN["NUMERO"], numero_cnj)

            ok["Andamento"] = self._selecionar_fixo("EVENTO1")
            self._preencher_texto_por_fieldname(
                TEXT_FN["DATA_ANDAMENTO_1"], date.today().strftime("%d/%m/%Y"))

            # ---- "Adverso já cadastrado" = Sim ANTES dos participantes ----
            # DESCOBERTA (log): marcar este radio RECONSTRÓI a seção Participantes
            # (os blocos passam de 2/3 para 1/2). Por isso marcamos PRIMEIRO,
            # estabilizamos, e SÓ ENTÃO detectamos/preenchemos os participantes.
            self._marcar_radio_por_pergunta(["adverso", "cadastrado"], "Sim")
            self._marcar_radio_por_pergunta(["advogado", "adverso"], "Sim")
            try:
                self._estabilizar_form(timeout=15, quieto=1.5)
            except Exception:
                pass
            time.sleep(1.0)

            # ---- PARTICIPANTES (detecção DINÂMICA após o Adverso=Sim) ----
            #   1º bloco = pessoa da planilha -> Réu ; 2º bloco = PREVI -> Autor
            parts, conds = self._blocos_participantes()
            self._log(f"    [participantes] blocos detectados: parts={parts} conds={conds}")
            if len(parts) >= 1:
                it_pessoa = self._buscar_item_api(parts[0], nome)
                if it_pessoa and it_pessoa.get("id") is not None:
                    ok["Participante réu"] = self._set_valor_por_id(
                        parts[0], it_pessoa["id"], it_pessoa.get("text", nome))
                if len(conds) >= 1:
                    self._set_valor_por_id(conds[0], COND_ID["Réu"], "Réu")
            if len(parts) >= 2:
                self._set_valor_por_id(parts[1], PREVI_ID, PREVI_TEXTO)
                if len(conds) >= 2:
                    self._set_valor_por_id(conds[1], COND_ID["Autor"], "Autor")

            # ---- Advogados via API ----
            self._selecionar_via_api(FN["ADV_INTERNO"], adv_interno)
            self._selecionar_via_api(FN["ADV_EXTERNO"], adv_externo)

            # ---- Fixos: Pedido e Risco ----
            ok["Pedido"] = self._selecionar_fixo("PEDIDO1")
            if valor_pedido and valor_pedido > 0:
                self._preencher_texto_por_fieldname(
                    TEXT_FN["VALOR_PEDIDO_1"], f"{valor_pedido:.2f}".replace(".", ","))
            ok["Risco"] = self._selecionar_fixo("RISCOPEDIDO1")

            # ---- AJUSTES FINAIS obrigatórios p/ o Benner permitir salvar ----
            # (a) Grupos de radio (confirmar estado correto):
            self._marcar_radio_grupo("NUMEROUNICO", "Não")       # Número único = Não
            self._marcar_radio_grupo("TIPOPROCESSO", "Ativo")    # Tipo processo = Ativo
            self._marcar_radio_grupo("PROCESSORELEVANTE", "Não") # Processo relevante = Não
            # Localização (Física/Digital) - obrigatório detectado no log -> Física
            self._marcar_radio_por_pergunta(["localiza"], "Física") \
                or self._marcar_radio_grupo("LOCALIZACAO", "Física")
            # (b) Reforçar Condição = Autor da PREVI (2º bloco de participante).
            try:
                _p, _c = self._blocos_participantes()
                if len(_c) >= 2:
                    self._set_valor_por_id(_c[1], COND_ID["Autor"], "Autor")
            except Exception:
                pass
            # (c) LIMPAR os campos de DOCUMENTOS que vêm com "Inicial" (o Benner
            #     não deixa salvar se Tipo/Nome/Data de documento vierem preenchidos).
            self._limpar_documentos_inicial()

            # (d) REFORÇO: os postbacks dos radios APAGAM número/datas -> re-preencher
            #     agora (Número, Data andamento, Data distribuição). Corrige o
            #     "NUMERODISTRIBUICAO vazio" e as datas obrigatórias do log.
            self._reforcar_campos_criticos_pre_save(numero_cnj)

            # (e) Data do ANDAMENTO (obrigatória) por id, caso o fieldname não pegue.
            try:
                self._preencher_texto_por_fieldname(
                    "DATAANDAMENTO1", date.today().strftime("%d/%m/%Y"))
            except Exception:
                pass

            # (f) TIPODOCUMENTO da DISTRIBUIÇÃO voltou a ser OBRIGATÓRIO (log):
            #     com "Já distribuído=Não" o Benner exige Tipo documento = Inicial.
            #     Setar via ID interno (22) de forma segura.
            try:
                self._set_valor_por_id("TIPODOCUMENTO", 22, "Inicial")
            except Exception:
                pass

            # VALIDAÇÃO
            essenciais = ["Filial", "Tipo", "Rito", "Orgão", "Andamento", "Participante réu", "Pedido"]
            faltando = [k for k in essenciais if not ok.get(k)]
            if faltando:
                return f"CAMPOS ESSENCIAIS NAO PREENCHIDOS: {', '.join(faltando)} - NÃO salvo."

            # PASSO 4: SALVAR (transacional definitivo)
            ok_tx, ident, detalhe = self._save_transacional()
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

            if ok_tx:
                return f"OK|{ident}" if ident else "OK"
            return f"NAO SALVO - {detalhe}"
        except Exception as e:
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass
            return str(e)

    def _categoria_confirmada(self):
        """True se a categoria 'Cível' foi REALMENTE commitada. Descoberto via
        diagnóstico do HTML real: o Benner lê um campo OCULTO (CATEGORIA_VALUE,
        apontado por data-inputhiddenid). Só quando esse hidden tem valor é que
        o OK monta o formulário. Checa: (1) hidden preenchido; (2) select nativo
        com Cível/value=2; (3) texto renderizado do select2."""
        try:
            return bool(self.driver.execute_script("""
                var sel = document.querySelector('select[data-fieldname="CATEGORIA"]')
                       || document.querySelector('select.benner-search');
                if(!sel) return false;
                // 1) hidden CATEGORIA_VALUE preenchido = commit REAL
                var hid = sel.getAttribute('data-inputhiddenid');
                if(hid){
                    var h=document.getElementById(hid);
                    if(h && h.value && (''+h.value).trim()!=='') return true;
                }
                // 2) <select> nativo com Cível (value=2) selecionado
                if(sel.value && (''+sel.value).trim()!==''){
                    if((''+sel.value)==='2') return true;
                    var op=sel.options[sel.selectedIndex];
                    var ot=(op&&op.text)?op.text.toLowerCase():'';
                    if(ot.indexOf('cível')>=0||ot.indexOf('civel')>=0) return true;
                }
                // 3) texto renderizado pelo select2
                var n=sel.nextElementSibling, ren=null;
                while(n){ if(n.classList&&n.classList.contains('select2')){
                    ren=n.querySelector('.select2-selection__rendered'); break;}
                    n=n.nextElementSibling; }
                var t = ren ? (ren.innerText||ren.textContent||'') : '';
                if (t.toLowerCase().indexOf('cível')>=0 ||
                    t.toLowerCase().indexOf('civel')>=0) return true;
                return false;"""))
        except Exception:
            return False

    def _forcar_categoria_civel_js(self):
        """Correção da CAUSA RAIZ (erro do servidor: SelectedItems[0].id = null).

        O Benner liga um handler ao evento select2 'select2:select' e lê
        e.params.data.id para montar SelectedItems. Antes disparávamos o evento
        SEM params.data -> id vinha null. Aqui disparamos o evento COM o objeto
        params.data = {id, text} que o Benner espera, além de atualizar o
        <select>, o hidden CATEGORIA_VALUE e a API do select2."""
        try:
            return bool(self.driver.execute_script("""
                var sel = document.querySelector('select[data-fieldname="CATEGORIA"]')
                       || document.querySelector('select.benner-search');
                if(!sel) return false;
                // localizar opção Cível (value=2 confirmado no HTML)
                var val=null, txtOpt='Cível';
                for(var i=0;i<sel.options.length;i++){
                    var t=(sel.options[i].text||'');
                    if(t.toLowerCase().indexOf('cível')>=0 ||
                       t.toLowerCase().indexOf('civel')>=0){
                        val=sel.options[i].value; txtOpt=t; break;
                    }
                }
                if(val===null) val='2';
                sel.value=val;
                // hidden CATEGORIA_VALUE
                var hid=sel.getAttribute('data-inputhiddenid');
                if(hid){var h=document.getElementById(hid);
                    if(h){h.value=val;
                        h.dispatchEvent(new Event('input',{bubbles:true}));
                        h.dispatchEvent(new Event('change',{bubbles:true}));}}
                sel.dispatchEvent(new Event('input',{bubbles:true}));
                sel.dispatchEvent(new Event('change',{bubbles:true}));
                if(window.jQuery){
                    var $=window.jQuery, $s=$(sel);
                    var data={id:val, text:txtOpt};
                    try{ $s.val(val).trigger('change'); }catch(e){}
                    // >>> A CHAVE: disparar select2:select COM params.data <<<
                    try{ $s.trigger({type:'select2:select', params:{data:data}}); }catch(e){}
                    try{ $s.trigger({type:'select2:selecting', params:{args:{data:data}}}); }catch(e){}
                    // se o objeto select2 expõe a seleção interna, popular também
                    try{
                        var inst = $s.data('select2');
                        if(inst){
                            if(inst.$container){}
                            if(typeof inst.trigger==='function'){
                                inst.trigger('select',{data:data});
                            }
                        }
                    }catch(e){}
                }
                return true;"""))
        except Exception:
            return False

    def _selecionar_categoria_click_real(self):
        """Clique REAL (ActionChains, nível do navegador) na opção 'Cível'.
        Diferente de JS/eventos sintéticos, gera eventos 'trusted' que os
        handlers do select2 processam integralmente -> popula SelectedItems."""
        try:
            sel = self.driver.find_element(
                By.CSS_SELECTOR, 'select[data-fieldname="CATEGORIA"], select.benner-search')
        except Exception:
            return False
        # abrir o dropdown clicando no container renderizado do select2
        try:
            cont = self.driver.execute_script("""
                var s=arguments[0],n=s.nextElementSibling;
                while(n){if(n.classList&&n.classList.contains('select2'))return n;
                    n=n.nextElementSibling;}
                return null;""", sel)
            if cont is None:
                return False
            self.driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", cont)
            time.sleep(0.3)
            alvo = cont.find_element(By.CSS_SELECTOR, ".select2-selection")
            ActionChains(self.driver).move_to_element(alvo).pause(0.2).click().perform()
            time.sleep(0.8)
        except Exception as e:
            self._log(f"    (click real) falha ao abrir dropdown: {e}")
            return False
        # localizar e clicar de verdade na opção Cível
        try:
            opt = None
            fim = time.time() + 8
            while time.time() < fim and opt is None:
                for o in self.driver.find_elements(
                        By.CSS_SELECTOR, "li.select2-results__option"):
                    low = (o.text or "").strip().lower()
                    if "cível" in low or "civel" in low:
                        opt = o
                        break
                if opt is None:
                    time.sleep(0.4)
            if opt is None:
                self._log("    (click real) opção Cível não apareceu no dropdown.")
                return False
            self.driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", opt)
            time.sleep(0.2)
            ActionChains(self.driver).move_to_element(opt).pause(0.2).click().perform()
            time.sleep(0.8)
            return True
        except Exception as e:
            self._log(f"    (click real) falha ao clicar em Cível: {e}")
            return False

    def _form_real_carregou(self):
        """True se o formulário real já montou (URL PR_CADASTRORAPIDOPASTA ou
        muitos selects na página)."""
        try:
            url = (self.driver.current_url or "").lower()
        except Exception:
            url = ""
        if "cadastrorapidopasta" in url:
            return True
        try:
            n = self.driver.execute_script(
                "return document.querySelectorAll("
                "'select.select2-hidden-accessible, select.benner-search').length;")
        except Exception:
            n = 0
        return n >= 10

    def _clicar_ok_categoria(self):
        """Clica no botão OK do seletor de categoria. Tenta o botão VISÍVEL
        primeiro (mais fiel ao fluxo manual) e usa o postback como fallback."""
        # 1) Botão visível "Ok" (link/button/input)
        try:
            candidatos = self.driver.find_elements(
                By.XPATH,
                "//a[normalize-space()='Ok'] | //button[normalize-space()='Ok'] | "
                "//input[@type='submit' and (@value='Ok' or @value='OK')] | "
                "//span[normalize-space()='Ok']/ancestor::a[1] | "
                "//span[normalize-space()='Ok']/ancestor::button[1]")
            for b in candidatos:
                try:
                    if b.is_displayed():
                        self._click_seguro(b)
                        return True
                except Exception:
                    continue
        except Exception:
            pass
        # 2) Fallback: postback direto
        try:
            self.driver.execute_script(f"{BTN_OK_CATEGORIA_JS};")
            return True
        except Exception as e:
            self._log(f"    Falha ao clicar OK da categoria: {e}")
            return False

    def _selecionar_categoria_civel(self):
        """AUTOMÁTICO com fallback manual mínimo.

        Ordem:
          1) CLIQUE REAL (ActionChains) na opção 'Cível' -> gera eventos trusted
             que populam o modelo interno do select2 (SelectedItems).
          2) Se não confirmar, FORÇA via JS disparando select2:select COM
             params.data={id,text} (corrige o erro SelectedItems[0].id=null).
          3) Clica OK e espera o form real.
          4) SÓ se tudo falhar, pede 1 clique manual (Cível+OK) e segue.
        Assim, no caso normal, roda 100% sozinho para as 100+ pastas."""
        if self._form_real_carregou():
            return True

        self._log("    Seletor de categoria: selecionando 'Cível' (automático)...")
        confirmou = False
        for tentativa in range(3):
            # (1) clique REAL via ActionChains
            self._selecionar_categoria_click_real()
            time.sleep(0.6)
            if self._categoria_confirmada():
                confirmou = True
                break
            # (2) reforço: força select2:select com params.data
            self._forcar_categoria_civel_js()
            time.sleep(0.6)
            if self._categoria_confirmada():
                confirmou = True
                break
            self._log(f"    (tentativa {tentativa+1}) Cível ainda não commitou; repetindo...")
            self._fechar_dropdown()
            time.sleep(0.6)

        if confirmou:
            self._log("    'Cível' confirmado. Clicando OK...")
            time.sleep(0.4)
            self._clicar_ok_categoria()
            self._log("    OK (categoria) acionado. Aguardando form real...")
            self._aguardar_carregamento()
            fim = time.time() + 60
            while time.time() < fim:
                self._ir_para_janela_form_real(timeout=3)
                if self._form_real_carregou():
                    self._log("    Form real detectado após OK da categoria.")
                    return True
                time.sleep(2)
            self._log("    OK enviado mas o form real não abriu; tentando fallback manual.")

        # (4) Fallback manual mínimo (só se a automação falhar nesta pasta).
        print("\n" + "=" * 66)
        print("  A automacao da CATEGORIA falhou nesta pasta.")
        print("  Faca MANUALMENTE: selecione CIVEL e clique OK.")
        print("  Quando o formulario completo abrir, pressione ENTER.")
        print("=" * 66)
        try:
            input("  [ENTER apos Civel + OK] ")
        except EOFError:
            pass
        self._ir_para_janela_form_real(timeout=60)
        if self._form_real_carregou():
            self._log("    Form real detectado (fallback manual).")
            return True
        self._log("    Form real ainda não detectado.")
        return False

    def _aguardar_formulario_carregar(self, timeout=120, min_selects=10):
        """Aguarda a integração montar o form (~28s; até timeout). Loga progresso."""
        fim = time.time() + timeout
        ultimo_log = 0
        while time.time() < fim:
            try:
                n = self.driver.execute_script(
                    "return document.querySelectorAll("
                    "'select.select2-hidden-accessible, select.benner-search').length;")
            except Exception:
                n = 0
            if n >= min_selects:
                time.sleep(0.8)
                try:
                    n2 = self.driver.execute_script(
                        "return document.querySelectorAll("
                        "'select.select2-hidden-accessible, select.benner-search').length;")
                except Exception:
                    n2 = n
                if n2 >= min_selects:
                    self._log(f"    Formulário carregado: {n2} selects.")
                    time.sleep(1.0)
                    return True
            agora = time.time()
            if agora - ultimo_log >= 5:
                self._log(f"    Aguardando formulário... selects={n}")
                ultimo_log = agora
            time.sleep(1.5)
        self._log(f"    Timeout ({timeout}s) aguardando formulário.")
        return False

    def _achar_select2_por_label(self, label_text, ocorrencia=1):
        try:
            selects = self.driver.find_elements(
                By.CSS_SELECTOR, "select.select2-hidden-accessible, select.benner-search")
        except Exception:
            selects = []
        alvo = label_text.strip().lower().rstrip("*").strip()
        cont = 0
        for s in selects:
            try:
                lbl = self.driver.execute_script("""
                    var s=arguments[0];
                    var p=s.closest('.form-group,.field,.control-group,tr,.row');
                    if(p){var l=p.querySelector('label'); if(l) return (l.innerText||'').trim();}
                    var node=s;
                    for(var k=0;k<8&&node;k++){
                        var q=node.querySelector&&node.querySelector('label,span.caption,.control-label');
                        if(q&&(q.innerText||'').trim()) return (q.innerText||'').trim();
                        node=node.parentElement;
                    }
                    return '';""", s) or ""
                ln = lbl.strip().lower().rstrip("*").strip()
                casa = False
                if ln == alvo:
                    casa = True
                elif alvo in ln:
                    if alvo == "participante" and ln.startswith("participantes"):
                        casa = False
                    elif alvo == "condição" and ln != "condição":
                        casa = False
                    else:
                        casa = True
                if casa:
                    cont += 1
                    if cont == ocorrencia:
                        return s
            except Exception:
                continue
        return None

    def _achar_input_por_label(self, label_text, ocorrencia=1):
        alvo = label_text.strip().lower().rstrip("*").strip()
        try:
            campos = self.driver.find_elements(
                By.CSS_SELECTOR, "input.form-control, textarea.form-control")
        except Exception:
            campos = []
        cont = 0
        for c in campos:
            try:
                if c.get_attribute("type") == "hidden":
                    continue
                lbl = self.driver.execute_script("""
                    var e=arguments[0];
                    var p=e.closest('.form-group,.field,.control-group,tr,.row');
                    if(p){var l=p.querySelector('label'); if(l) return (l.innerText||'').trim();}
                    return '';""", c) or ""
                ln = lbl.strip().lower().rstrip("*").strip()
                if ln == alvo or (alvo and alvo in ln):
                    cont += 1
                    if cont == ocorrencia:
                        return c
            except Exception:
                continue
        return None

    def _termo_busca(self, texto):
        t = texto.strip()
        especiais = {
            "pedido de ajuizamento de ação": "ajuizamento",
            "dívida previdenciária": "previden",
            "plano de beneficios 1": "beneficios 1",
            "plano previ futuro": "previ futuro",
            "previdencial": "previden",
            "produto": "produto",
            "ordinário": "ordin",
            "possível": "poss",
            "réu": "réu",
            "autor": "autor",
        }
        low = t.lower()
        if low in especiais:
            return especiais[low]
        # termo curto (8 chars) costuma trazer mais resultados no AJAX do Benner
        return t[:8]

    def _aguardar_opcao(self, texto, timeout=15):
        fim = time.time() + timeout
        alvo = texto.lower().strip()
        # tokens úteis do alvo (ignora palavras muito curtas)
        tokens = [w for w in alvo.replace("-", " ").split() if len(w) >= 3]
        while time.time() < fim:
            try:
                opts = self.driver.find_elements(By.CSS_SELECTOR, "li.select2-results__option")
            except Exception:
                opts = []
            validas = []
            loading = False
            for o in opts:
                cls = (o.get_attribute("class") or "")
                low = (o.text or "").strip().lower()
                if ("loading" in cls or "procurando" in low or "searching" in low
                        or "carregando" in low):
                    loading = True
                    continue
                # ignorar mensagens de "nenhum resultado"
                if ("nenhum resultado" in low or "no results" in low
                        or "sem resultados" in low):
                    continue
                if low:
                    validas.append((o, low))
            # 1) match exato/contido
            for o, low in validas:
                if alvo and (alvo in low or low in alvo):
                    return o
            # 2) match por TODOS os tokens do alvo
            if tokens:
                for o, low in validas:
                    if all(tok in low for tok in tokens):
                        return o
            # 3) match por prefixo do alvo (8 chars)
            pref = alvo[:8]
            if pref:
                for o, low in validas:
                    if pref in low:
                        return o
            # 4) se só existe UMA opção válida, assume ela
            if len(validas) == 1 and not loading:
                return validas[0][0]
            time.sleep(0.5)
        return None

    def _fechar_dropdown(self):
        try:
            self.driver.find_element(By.CSS_SELECTOR, "input.select2-search__field").send_keys(Keys.ESCAPE)
        except Exception:
            try:
                self.driver.execute_script("document.activeElement.blur();")
            except Exception:
                pass

    def _dropdown_fechou(self):
        """True se o dropdown do select2 fechou (opção assumida)."""
        try:
            campos = self.driver.find_elements(
                By.CSS_SELECTOR, "input.select2-search__field")
            for c in campos:
                if c.is_displayed():
                    return False
            return True
        except Exception:
            return True

    def _forcar_select_por_texto(self, s, texto, descricao=""):
        """TÉCNICA VENCEDORA generalizada (a mesma que resolveu a Categoria).
        Em vez de digitar/AJAX, seleciona a opção diretamente no <select>:
          1) acha a <option> cujo texto casa com 'texto' (case-insensitive,
             por contido OU por todos os tokens);
          2) seta value no <select> e no hidden (data-inputhiddenid), se houver;
          3) dispara change nativo + select2:select COM params.data={id,text}
             (é isso que o handler do Benner lê -> evita SelectedItems null).
        Só funciona se a <option> já existir no <select> (para campos em
        cascata, chame DEPOIS de commitar o campo pai)."""
        if s is None or not texto:
            return False
        try:
            ok = self.driver.execute_script(r"""
                var sel=arguments[0], alvo=(arguments[1]||'').toLowerCase().trim();
                if(!sel||!sel.options) return false;
                var toks=alvo.replace(/-/g,' ').split(/\s+/).filter(function(w){return w.length>=3;});
                function casa(t){
                    t=(t||'').toLowerCase();
                    if(!t) return false;
                    if(t.indexOf(alvo)>=0 || alvo.indexOf(t)>=0) return true;
                    if(toks.length){ for(var k=0;k<toks.length;k++){ if(t.indexOf(toks[k])<0) return false;} return true; }
                    return false;
                }
                var achouVal=null, achouTxt=null;
                for(var i=0;i<sel.options.length;i++){
                    var o=sel.options[i];
                    if(!o.value && !o.text) continue;
                    if(casa(o.text)){ achouVal=o.value; achouTxt=o.text; sel.selectedIndex=i; break; }
                }
                if(achouVal===null) return false;
                sel.value=achouVal;
                var hid=sel.getAttribute('data-inputhiddenid');
                if(hid){var h=document.getElementById(hid);
                    if(h){h.value=achouVal;
                        h.dispatchEvent(new Event('input',{bubbles:true}));
                        h.dispatchEvent(new Event('change',{bubbles:true}));}}
                sel.dispatchEvent(new Event('input',{bubbles:true}));
                sel.dispatchEvent(new Event('change',{bubbles:true}));
                if(window.jQuery){
                    var $=window.jQuery,$s=$(sel),data={id:achouVal,text:achouTxt};
                    try{$s.val(achouVal).trigger('change');}catch(e){}
                    try{$s.trigger({type:'select2:select',params:{data:data}});}catch(e){}
                }
                return true;""", s, texto)
            if ok:
                self._log(f"    [forçar] '{texto}' selecionado em {descricao}.")
            return bool(ok)
        except Exception as e:
            self._log(f"    [forçar] falha em {descricao}: {e}")
            return False

    def _operar_select2(self, s, texto, descricao=""):
        """Núcleo do select2 AJAX robusto (mesmas armas que venceram a categoria):
          - abre o dropdown com clique REAL (ActionChains);
          - digita o termo e DISPARA os eventos (input/keyup) p/ acionar o AJAX;
          - aguarda as opções (ignora 'nenhum resultado'/loading);
          - seleciona com ENTER (opção destacada) e, se preciso, clique REAL."""
        if s is None:
            return False
        termo_busca = self._termo_busca(texto)
        for tentativa in range(3):
            # (A) abrir o dropdown
            try:
                cont = self.driver.execute_script("""
                    var s=arguments[0],n=s.nextElementSibling;
                    while(n){if(n.classList&&n.classList.contains('select2'))return n;n=n.nextElementSibling;}
                    return null;""", s)
                if cont is None:
                    return False
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center',behavior:'instant'});", cont)
                time.sleep(0.3)
                sel_part = cont.find_element(By.CSS_SELECTOR, ".select2-selection")
                try:
                    ActionChains(self.driver).move_to_element(sel_part).pause(0.15).click().perform()
                except Exception:
                    try:
                        sel_part.click()
                    except Exception:
                        self.driver.execute_script("arguments[0].click();", sel_part)
                time.sleep(0.8)
            except Exception as e:
                self._log(f"    Falha abrir select2 {descricao}: {e}")
                return False

            # (B) digitar o termo e DISPARAR eventos para o AJAX carregar
            try:
                busca = self._esperar_visivel(By.CSS_SELECTOR, "input.select2-search__field", 6)
                if busca:
                    busca.clear()
                    # digitar char a char aciona melhor o keyup do select2
                    for ch in termo_busca:
                        busca.send_keys(ch)
                        time.sleep(0.05)
                    # reforço: disparar input/keyup manualmente
                    try:
                        self.driver.execute_script("""
                            var e=arguments[0];
                            e.dispatchEvent(new Event('input',{bubbles:true}));
                            e.dispatchEvent(new KeyboardEvent('keyup',{bubbles:true}));
                        """, busca)
                    except Exception:
                        pass
                    time.sleep(2.5)  # aguarda resposta do servidor (AJAX)
            except Exception:
                pass

            # (C) aguardar a opção correta aparecer
            opt = self._aguardar_opcao(texto, timeout=15)
            if opt:
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center',behavior:'instant'});", opt)
                time.sleep(0.2)
                selecionou = False

                # (C1) destacar + ENTER (select2 assume a opção destacada)
                try:
                    self.driver.execute_script("""
                        var o=arguments[0];
                        ['mouseenter','mouseover'].forEach(function(t){
                            o.dispatchEvent(new MouseEvent(t,{bubbles:true,view:window}));});
                        var lis=document.querySelectorAll('li.select2-results__option');
                        for(var i=0;i<lis.length;i++)
                            lis[i].classList.remove('select2-results__option--highlighted');
                        o.classList.add('select2-results__option--highlighted');
                        o.setAttribute('aria-selected','true');""", opt)
                    busca2 = self.driver.find_element(
                        By.CSS_SELECTOR, "input.select2-search__field")
                    busca2.send_keys(Keys.ENTER)
                    time.sleep(0.6)
                    selecionou = self._dropdown_fechou()
                except Exception:
                    pass

                # (C2) clique REAL (ActionChains) na opção
                if not selecionou:
                    try:
                        ActionChains(self.driver).move_to_element(opt).pause(0.15).click().perform()
                        time.sleep(0.6)
                        selecionou = self._dropdown_fechou()
                    except Exception:
                        pass

                # (C3) último recurso: clique JS
                if not selecionou:
                    try:
                        self.driver.execute_script("arguments[0].click();", opt)
                        time.sleep(0.6)
                    except Exception:
                        pass

                time.sleep(0.4)
                return True

            self._fechar_dropdown()
            self._log(f"    (tentativa {tentativa+1}) Sem opção para '{texto}' em {descricao}.")
            time.sleep(1)
        return False

    def _selecionar_lookup(self, label_text, texto, ocorrencia=1):
        """SELECT2 AJAX por LABEL (fallback)."""
        if not texto:
            return True
        s = self._achar_select2_por_label(label_text, ocorrencia)
        if s is None:
            self._log(f"    Select (label='{label_text}' #{ocorrencia}) não encontrado.")
            return False
        return self._operar_select2(s, texto, descricao=f"label='{label_text}' #{ocorrencia}")

    def _achar_select_por_fieldname(self, fieldname):
        """Localiza o <select> pelo atributo ESTÁVEL data-fieldname."""
        try:
            return self.driver.find_element(
                By.CSS_SELECTOR, f'select[data-fieldname="{fieldname}"]')
        except Exception:
            return None

    def _contar_opcoes_fn(self, fieldname):
        """Conta as <option> reais (com value) de um <select> por data-fieldname."""
        try:
            return int(self.driver.execute_script("""
                var s=document.querySelector('select[data-fieldname="'+arguments[0]+'"]');
                if(!s||!s.options) return 0;
                var c=0;
                for(var i=0;i<s.options.length;i++){
                    var o=s.options[i];
                    if(o.value && (''+o.value).trim()!=='') c++;
                }
                return c;""", fieldname) or 0)
        except Exception:
            return 0

    def _aguardar_cascata_fn(self, fieldname, timeout=20, minimo=1):
        """Após commitar um campo pai, aguarda o campo dependente (cascata)
        ganhar opções via o postback AJAX do Benner (ex.: TIPO -> CAUSA_PEDIR)."""
        fim = time.time() + timeout
        ultimo = 0
        while time.time() < fim:
            n = self._contar_opcoes_fn(fieldname)
            if n >= minimo:
                self._log(f"    Cascata OK: fieldname='{fieldname}' com {n} opções.")
                return True
            ag = time.time()
            if ag - ultimo >= 4:
                self._log(f"    Aguardando cascata popular '{fieldname}' (opções={n})...")
                ultimo = ag
            time.sleep(1.0)
        self._log(f"    Cascata NÃO populou '{fieldname}' em {timeout}s.")
        return False

    def _click_real_opcao_fieldname(self, fieldname, texto, timeout_ajax=15):
        """MÉTODO VENCEDOR (o mesmo que resolveu a Categoria), generalizado para
        QUALQUER campo por data-fieldname. Todos os 30 campos do form real são
        select2 AJAX (data-searchcontext) com ZERO opções embutidas -> a única
        forma que funciona é: abrir o dropdown com CLIQUE REAL (ActionChains),
        digitar o termo p/ disparar o AJAX, aguardar as opções chegarem e CLICAR
        no <li> com clique REAL (eventos trusted que o select2 processa)."""
        # IMPORTANTE (fix stale): NÃO guardamos referência Selenium do <select>
        # (o Benner reconstrói o DOM e a referência "envelhece"). Localizamos o
        # container select2 e o alvo de clique SEMPRE por data-fieldname dentro
        # do próprio JS, obtendo um elemento FRESCO imediatamente antes do uso.
        alvo = None
        for _tent_open in range(3):
            try:
                alvo = self.driver.execute_script("""
                    var fn=arguments[0];
                    var s=document.querySelector('select[data-fieldname="'+fn+'"]');
                    if(!s) return null;
                    var n=s.nextElementSibling, cont=null;
                    while(n){ if(n.classList&&n.classList.contains('select2')){cont=n;break;}
                        n=n.nextElementSibling; }
                    if(!cont) return null;
                    cont.scrollIntoView({block:'center'});
                    return cont.querySelector('.select2-selection');""", fieldname)
                if alvo is not None:
                    break
            except StaleElementReferenceException:
                pass
            except Exception:
                pass
            time.sleep(0.5)
        if alvo is None:
            self._log(f"    (click real fn={fieldname}) container select2 não achado.")
            return False
        try:
            time.sleep(0.2)
            try:
                ActionChains(self.driver).move_to_element(alvo).pause(0.15).click().perform()
            except StaleElementReferenceException:
                # re-obter fresco e clicar via JS como reforço
                alvo = self.driver.execute_script("""
                    var fn=arguments[0];
                    var s=document.querySelector('select[data-fieldname="'+fn+'"]');
                    if(!s) return null;
                    var n=s.nextElementSibling, cont=null;
                    while(n){ if(n.classList&&n.classList.contains('select2')){cont=n;break;}
                        n=n.nextElementSibling; }
                    return cont?cont.querySelector('.select2-selection'):null;""", fieldname)
                if alvo is not None:
                    self.driver.execute_script("arguments[0].click();", alvo)
            time.sleep(0.7)
        except Exception as e:
            self._log(f"    (click real fn={fieldname}) falha ao abrir: {e}")
            return False

        # digitar o termo no campo de busca do select2 (dispara o AJAX)
        termo = self._termo_busca(texto)
        try:
            busca = self._esperar_visivel(By.CSS_SELECTOR, "input.select2-search__field", 6)
            if busca:
                busca.clear()
                for ch in termo:
                    busca.send_keys(ch)
                    time.sleep(0.05)
                try:
                    self.driver.execute_script("""
                        var e=arguments[0];
                        e.dispatchEvent(new Event('input',{bubbles:true}));
                        e.dispatchEvent(new KeyboardEvent('keyup',{bubbles:true}));""", busca)
                except Exception:
                    pass
        except Exception:
            pass

        # aguardar as opções do AJAX chegarem e localizar a correta
        alvo_low = texto.lower().strip()
        toks = [w for w in alvo_low.replace("-", " ").split() if len(w) >= 3]
        opt = None
        fim = time.time() + timeout_ajax
        while time.time() < fim and opt is None:
            try:
                lis = self.driver.find_elements(By.CSS_SELECTOR, "li.select2-results__option")
            except Exception:
                lis = []
            validas = []
            carregando = False
            for o in lis:
                cls = (o.get_attribute("class") or "")
                low = (o.text or "").strip().lower()
                if ("loading" in cls or "procurando" in low or "searching" in low
                        or "carregando" in low):
                    carregando = True
                    continue
                if "nenhum resultado" in low or "no results" in low:
                    continue
                if low:
                    validas.append((o, low))
            # match exato/contido
            for o, low in validas:
                if alvo_low and (alvo_low in low or low in alvo_low):
                    opt = o; break
            # match por todos os tokens
            if opt is None and toks:
                for o, low in validas:
                    if all(t in low for t in toks):
                        opt = o; break
            # match por prefixo
            if opt is None and alvo_low:
                pref = alvo_low[:8]
                for o, low in validas:
                    if pref and pref in low:
                        opt = o; break
            # única opção disponível
            if opt is None and len(validas) == 1 and not carregando:
                opt = validas[0][0]
            if opt is None:
                time.sleep(0.4)

        if opt is None:
            self._log(f"    (click real fn={fieldname}) opção '{texto}' não apareceu (AJAX).")
            self._fechar_dropdown()
            return False

        # CLICAR no <li> com clique REAL (ActionChains) + ENTER de reforço
        try:
            self.driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", opt)
            time.sleep(0.15)
            ActionChains(self.driver).move_to_element(opt).pause(0.15).click().perform()
            time.sleep(0.6)
            if self._dropdown_fechou():
                self._log(f"    [click real] '{texto}' selecionado em fn={fieldname}.")
                return True
            # reforço: ENTER na opção destacada
            try:
                b2 = self.driver.find_element(By.CSS_SELECTOR, "input.select2-search__field")
                b2.send_keys(Keys.ENTER)
                time.sleep(0.5)
            except Exception:
                pass
            if self._dropdown_fechou():
                self._log(f"    [click real+ENTER] '{texto}' selecionado em fn={fieldname}.")
                return True
        except Exception as e:
            self._log(f"    (click real fn={fieldname}) falha ao clicar opção: {e}")
        return False

    def _estabilizar_form(self, timeout=25, quieto=2.0):
        """Espera o form ficar sem mutações por 'quieto's (Benner faz postbacks
        assíncronos que reconstroem o DOM -> stale)."""
        try:
            self.driver.execute_script("""
                if(!window.__mutObs){ window.__lastMut=Date.now();
                    window.__mutObs=new MutationObserver(function(){window.__lastMut=Date.now();});
                    window.__mutObs.observe(document.body,{childList:true,subtree:true,attributes:true});
                }else{ window.__lastMut=Date.now(); }""")
        except Exception:
            time.sleep(quieto); return True
        fim = time.time()+timeout
        while time.time()<fim:
            try:
                gap=self.driver.execute_script("return (Date.now()-(window.__lastMut||0))/1000.0;")
            except Exception:
                gap=quieto
            if gap is not None and gap>=quieto:
                return True
            time.sleep(0.3)
        return False

    def _set_valor_por_id(self, fieldname, id_interno, texto):
        """MOTOR DETERMINÍSTICO (v79 - SEM triggers select2/jQuery que CRASHAM).
        O Console provou: os triggers do select2 chamam Search.js do Benner, que
        quebra com 'ResizeObserver' e corrompe o viewstate. O servidor lê o HIDDEN
        *_VALUE no Save; então usamos só eventos NATIVOS + atualização cosmética."""
        try:
            r = self.driver.execute_script(r"""
                var fn=arguments[0], id=arguments[1], text=arguments[2];
                var sel=document.querySelector('select[data-fieldname="'+fn+'"]');
                if(!sel) return 'nosel';
                var achou=false;
                for(var i=0;i<sel.options.length;i++){
                    if((''+sel.options[i].value)===(''+id)){achou=true;break;} }
                if(!achou){ var o=document.createElement('option');
                    o.value=id;o.text=text;o.selected=true;sel.appendChild(o); }
                sel.value=''+id;
                var hid=sel.getAttribute('data-inputhiddenid');
                if(hid){ var h=document.getElementById(hid);
                    if(h){ h.value=JSON.stringify({ReadOnly:false,View:false,
                        SelectedItems:[{id:parseInt(id,10),text:text}]});
                        h.dispatchEvent(new Event('input',{bubbles:true}));
                        h.dispatchEvent(new Event('change',{bubbles:true})); } }
                // eventos NATIVOS apenas (NAO usar jQuery/select2 -> evita crash
                // do Search.js/ResizeObserver que corrompe o viewstate):
                sel.dispatchEvent(new Event('change',{bubbles:true}));
                // atualizar o TEXTO visivel do select2 direto no DOM (cosmetico):
                try{
                    var cont=sel.nextElementSibling;
                    while(cont && !(cont.classList && cont.classList.contains('select2')))
                        cont=cont.nextElementSibling;
                    if(cont){
                        var rend=cont.querySelector('.select2-selection__rendered');
                        if(rend){ rend.setAttribute('title',text);
                            var item=rend.querySelector('.selected-item, xmp');
                            if(item){ item.textContent=text; } else { rend.textContent=text; }
                        }
                    }
                }catch(e){}
                return 'ok';""", fieldname, id_interno, texto)
            if r == "ok":
                self._log(f"    [set-id] fn={fieldname} id={id_interno} ('{str(texto)[:25]}') OK.")
                return True
            self._log(f"    [set-id] fn={fieldname}: {r}.")
            return False
        except Exception as e:
            self._log(f"    [set-id] fn={fieldname}: erro {e}")
            return False

    def _selecionar_fixo(self, fieldname):
        par = IDS_FIXOS.get(fieldname)
        if not par: return False
        return self._set_valor_por_id(fieldname, par[0], par[1])

    def _selecionar_por_opcao_embutida(self, fieldname, texto):
        """Fallback: se o <select> já tem uma <option> cujo texto casa com
        'texto', seta por ela (id embutido). Útil p/ FILIAL, que costuma trazer
        a(s) opção(ões) já no HTML."""
        try:
            item = self.driver.execute_script(r"""
                var fn=arguments[0], alvo=(arguments[1]||'').toLowerCase().trim();
                var sel=document.querySelector('select[data-fieldname="'+fn+'"]');
                if(!sel) return null;
                var toks=alvo.replace(/-/g,' ').split(/\s+/).filter(function(w){return w.length>=3;});
                function casa(t){ t=(t||'').toLowerCase();
                    if(!t) return false;
                    if(t.indexOf(alvo)>=0||alvo.indexOf(t)>=0) return true;
                    if(toks.length){for(var k=0;k<toks.length;k++){if(t.indexOf(toks[k])<0)return false;}return true;}
                    return false; }
                for(var i=0;i<sel.options.length;i++){
                    var o=sel.options[i];
                    if(o.value && casa(o.text)) return {id:o.value, text:o.text};
                }
                return null;""", fieldname, texto)
            if item and item.get("id") is not None:
                return self._set_valor_por_id(fieldname, item["id"], item.get("text", texto))
        except Exception as e:
            self._log(f"    [opcao-embutida] fn={fieldname}: erro {e}")
        return False

    def _blocos_participantes(self):
        """Retorna, em ORDEM do DOM, as listas de data-fieldname dos selects de
        Participante e de Condição EXISTENTES no form. A numeração (1/2/3) VARIA
        entre carregamentos, então detectamos dinamicamente. Ignora *SEMLOTE*.
        Ex.: participantes=['PARTICIPANTE2','PARTICIPANTE3'],
             condicoes=['CONDICAO2','CONDICAO3']."""
        try:
            res = self.driver.execute_script("""
                function coletar(prefixo, exclui){
                    var out=[];
                    var sels=document.querySelectorAll('select[data-fieldname]');
                    for(var i=0;i<sels.length;i++){
                        var fn=sels[i].getAttribute('data-fieldname')||'';
                        if(fn.indexOf(prefixo)===0 && fn.indexOf(exclui)<0){
                            out.push(fn);
                        }
                    }
                    return out;
                }
                return {
                    parts: coletar('PARTICIPANTE','SEMLOTE'),
                    conds: coletar('CONDICAO','LOTE')
                };
            """) or {}
            return (res.get("parts", []), res.get("conds", []))
        except Exception as e:
            self._log(f"    [participantes] erro ao detectar: {e}")
            return ([], [])

    def _api_search(self, fieldname, query="", depend=None, timeout=15):
        """Chama a API /api/search DENTRO do navegador (usa cookies/sessão) com o
        data-searchcontext (sc) do próprio <select>. Retorna lista [{id,text}]."""
        try:
            js = r"""
              var fn=arguments[0], query=arguments[1], depend=arguments[2];
              var done=arguments[arguments.length-1];
              var sel=document.querySelector('select[data-fieldname="'+fn+'"]');
              if(!sel){ done({err:'nosel'}); return; }
              var sc=sel.getAttribute('data-searchcontext')||'';
              // base da API: mesma origem + /JURIDICO/api/search
              var m=location.pathname.match(/^(.*?\/JURIDICO)\//i);
              var base=(m?m[1]:'/JURIDICO')+'/api/search';
              var body='query='+encodeURIComponent(query)
                      +'&sc='+encodeURIComponent(sc)
                      +'&dependValueList='+encodeURIComponent(depend||'')
                      +'&maxRows=50&startRow=0&fieldsJson=';
              fetch(base,{method:'POST',
                  headers:{'Content-Type':'application/x-www-form-urlencoded'},
                  body:body, credentials:'include'})
                .then(function(r){return r.text();})
                .then(function(t){ try{ done({ok:JSON.parse(t)}); }
                    catch(e){ done({err:'parse',raw:(t||'').slice(0,200)}); } })
                .catch(function(e){ done({err:String(e)}); });
            """
            self.driver.set_script_timeout(timeout)
            res = self.driver.execute_async_script(js, fieldname, query, depend or "")
            payload = None
            if isinstance(res, dict) and "ok" in res:
                payload = res["ok"]
            else:
                self._log(f"    [api] fn={fieldname} query='{query}': {str(res)[:150]}")
                return []
            # normalizar payload para lista de dicts {id,text}
            return self._normalizar_itens(payload)
        except Exception as e:
            self._log(f"    [api] fn={fieldname}: erro {e}")
            return []

    def _normalizar_itens(self, payload):
        """Normaliza a resposta da API para uma lista de dicts {id,text},
        tratando formatos variados (lista de dicts, lista de strings, objeto
        com chave items/results/data, etc.)."""
        # se veio um dict, procurar a lista dentro dele
        if isinstance(payload, dict):
            for k in ("items", "results", "data", "rows", "list", "Items", "Results"):
                if isinstance(payload.get(k), list):
                    payload = payload[k]
                    break
            else:
                # dict único {id,text}?
                if "id" in payload or "text" in payload or "Id" in payload:
                    payload = [payload]
                else:
                    return []
        if not isinstance(payload, list):
            return []
        out = []
        for it in payload:
            if isinstance(it, dict):
                _id = it.get("id", it.get("Id", it.get("value", it.get("Value"))))
                _tx = it.get("text", it.get("Text", it.get("nome", it.get("descricao", ""))))
                if _id is not None:
                    out.append({"id": _id, "text": str(_tx or "")})
            elif isinstance(it, (list, tuple)) and len(it) >= 2:
                out.append({"id": it[0], "text": str(it[1])})
            elif isinstance(it, str):
                # string pura: sem id -> usar o próprio texto como id (fallback)
                out.append({"id": it, "text": it})
        return out

    def _match_opcao(self, itens, texto):
        """Escolhe o melhor item {id,text} para 'texto' (exato > contido > tokens
        > prefixo). Blindado contra itens que não sejam dict."""
        itens = self._normalizar_itens(itens) if not (
            itens and isinstance(itens[0], dict) and "id" in itens[0]) else itens
        if not itens:
            return None
        def _t(it):
            return (it.get("text") if isinstance(it, dict) else str(it)) or ""
        alvo = (texto or "").strip().lower()
        toks = [w for w in alvo.replace("-", " ").split() if len(w) >= 3]
        for it in itens:
            if _t(it).strip().lower() == alvo:
                return it
        for it in itens:
            t = _t(it).lower()
            if alvo and (alvo in t or t in alvo):
                return it
        if toks:
            for it in itens:
                t = _t(it).lower()
                if all(tk in t for tk in toks):
                    return it
        if len(itens) == 1:
            return itens[0]
        return None

    def _queries_para(self, texto):
        """Gera queries em ordem de eficácia para a API (busca por PREFIXO):
        texto inteiro, 1ª palavra, 1ª palavra[:4], e vazio (lista tudo)."""
        t = (texto or "").strip()
        palavras = [w for w in t.replace("-", " ").split() if w]
        qs = []
        if t:
            qs.append(t[:20])
        if palavras:
            qs.append(palavras[0][:12])
            qs.append(palavras[0][:4])
        qs.append("")   # vazio = lista tudo -> match por 'contém'
        # remover duplicados preservando ordem
        vis = set(); out = []
        for q in qs:
            if q not in vis:
                vis.add(q); out.append(q)
        return out

    def _buscar_item_api(self, fieldname, texto, depend=None):
        """Busca na API o item {id,text} que casa com 'texto'. Retorna dict ou None.
        Tenta várias queries porque a API do Benner busca por PREFIXO."""
        for q in self._queries_para(texto):
            itens = self._api_search(fieldname, query=q, depend=depend)
            it = self._match_opcao(itens, texto)
            if it and isinstance(it, dict) and it.get("id") is not None:
                return it
        return None

    def _selecionar_via_api(self, fieldname, texto, depend=None):
        """DETERMINÍSTICO p/ campos variáveis: chama /api/search, acha o id e
        seta via _set_valor_por_id. Substitui o clique/AJAX frágil."""
        if not texto:
            return True
        it = self._buscar_item_api(fieldname, texto, depend=depend)
        if it:
            return self._set_valor_por_id(fieldname, it["id"], it.get("text", texto))
        self._log(f"    [api] fn={fieldname}: nenhum match para '{texto}'.")
        return False

    def _selecionar_por_fieldname(self, fieldname, texto, essencial_nome=None):
        """Seleciona uma opção num <select2> localizado por data-fieldname.
        Estratégia: (1) FORÇAR direto no <select> (para opções embutidas, como
        TIPO/CAUSA/RITO/PEDIDO/RISCO/CONDIÇÃO); (2) se falhar (listas AJAX como
        PESSOAS/ADVOGADOS), abrir o dropdown, digitar e clicar."""
        if not texto:
            return True
        desc = f"fn={fieldname}"
        s = self._achar_select_por_fieldname(fieldname)
        if s is None:
            self._log(f"    Select ({desc}) NÃO encontrado no DOM.")
            return False
        # TODOS os campos do form real são select2 AJAX (data-searchcontext) com
        # ZERO opções embutidas. Portanto o CAMINHO PRINCIPAL é o clique REAL
        # (ActionChains) que abre o dropdown, dispara o AJAX e clica na opção.
        for tentativa in range(3):
            try:
                if self._click_real_opcao_fieldname(fieldname, texto):
                    return True
            except StaleElementReferenceException:
                self._log(f"    (tentativa {tentativa+1}) stale em {desc}; DOM mudou, refazendo...")
            self._log(f"    (tentativa {tentativa+1}) clique real falhou em {desc}; repetindo...")
            self._fechar_dropdown()
            time.sleep(0.8)
        # Último recurso: force-JS (re-localiza 's' FRESCO p/ evitar stale)
        try:
            s_fresh = self._achar_select_por_fieldname(fieldname)
        except Exception:
            s_fresh = None
        if s_fresh is not None and self._forcar_select_por_texto(
                s_fresh, texto, descricao=desc):
            return True
        self._log(f"    NÃO consegui selecionar '{texto}' em {desc}.")
        return False

    def _preencher_texto_por_fieldname(self, fieldname, valor):
        """Preenche input/textarea VISÍVEL: por span[data-field=...] OU por id contendo o fieldname."""
        elem = None
        seletores = (
            f'span[data-field="{fieldname}"] input:not([type="hidden"])',
            f'span[data-field="{fieldname}"] textarea',
            f'input[id*="{fieldname}"]:not([type="hidden"])',      # <<< NOVO fallback por id
            f'textarea[id*="{fieldname}"]',
        )
        for sel in seletores:
            try:
                for e in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    try:
                        if e.is_displayed():
                            elem = e
                            break
                    except Exception:
                        continue
            except Exception:
                pass
            if elem:
                break
        if elem is None:
            self._log(f"    Campo texto (fn={fieldname}) não encontrado.")
            return False
        try:
            self.driver.execute_script("""
                var e=arguments[0],v=arguments[1];
                if(e.readOnly){e.readOnly=false;}
                e.value='';e.value=v;
                e.dispatchEvent(new Event('input',{bubbles:true}));
                e.dispatchEvent(new Event('change',{bubbles:true}));
                e.dispatchEvent(new Event('blur',{bubbles:true}));""", elem, valor)
            return True
        except Exception as e:
            self._log(f"    Falha preencher texto fn={fieldname}: {e}")
            return False

    def _selecionar_lookup_id(self, chave, texto):
        """SELECT2 AJAX por ID EXATO (preferencial). Se o ID não existir (layout
        mudou), cai no fallback por label."""
        if not texto:
            return True
        element_id = FIELD_IDS.get(chave, chave)
        s = None
        try:
            s = self.driver.find_element(By.ID, element_id)
        except Exception:
            s = None
        if s is not None:
            # (1) técnica vencedora: forçar direto no <select> + params.data
            if self._forcar_select_por_texto(s, texto, descricao=f"id={chave}"):
                return True
            # (2) fallback: abrir dropdown, digitar e clicar
            if self._operar_select2(s, texto, descricao=f"id={chave}"):
                return True
            self._log(f"    (id={chave}) falhou; tentando fallback por label...")
        lbl = _FALLBACK_LABEL.get(chave)
        if lbl:
            # tentar forçar pelo elemento localizado por label também
            s2 = self._achar_select2_por_label(lbl, _FALLBACK_OCC.get(chave, 1))
            if s2 is not None and self._forcar_select_por_texto(
                    s2, texto, descricao=f"label='{lbl}'"):
                return True
            return self._selecionar_lookup(lbl, texto,
                                           ocorrencia=_FALLBACK_OCC.get(chave, 1))
        self._log(f"    Select (id={chave}) não encontrado e sem fallback.")
        return False

    def _preencher_texto_por_id(self, chave, valor):
        """Preenche input/textarea por ID exato (com fallback por label)."""
        element_id = FIELD_IDS.get(chave, chave)
        elem = None
        try:
            elem = self.driver.find_element(By.ID, element_id)
        except Exception:
            elem = None
        if elem is None:
            lbl = _FALLBACK_LABEL.get(chave)
            if lbl:
                return self._preencher_texto_js(lbl, valor,
                                                ocorrencia=_FALLBACK_OCC.get(chave, 1))
            self._log(f"    Campo texto (id={chave}) não encontrado.")
            return False
        try:
            self.driver.execute_script("""
                var e=arguments[0],v=arguments[1];
                if(e.readOnly){e.readOnly=false;}
                e.value='';e.value=v;
                e.dispatchEvent(new Event('input',{bubbles:true}));
                e.dispatchEvent(new Event('change',{bubbles:true}));
                e.dispatchEvent(new Event('blur',{bubbles:true}));""", elem, valor)
            return True
        except Exception as e:
            self._log(f"    Falha preencher id={chave}: {e}")
            return False

    def _trocar_para_janela_do_formulario(self, timeout=30):
        """Passa a operar na aba do seletor de categoria / formulário
        (URL com CadastroRapidoManual / CADASTRORAPIDOPASTA / form.aspx)."""
        fim = time.time() + timeout
        marcadores = ("cadastrorapidomanual", "cadastrorapidopasta", "form.aspx")
        while time.time() < fim:
            try:
                handles = self.driver.window_handles
            except Exception:
                handles = []
            for h in handles:
                try:
                    self.driver.switch_to.window(h)
                    self.driver.switch_to.default_content()
                except Exception:
                    continue
                url = (self.driver.current_url or "").lower()
                if any(m in url for m in marcadores):
                    return True
            time.sleep(1)
        return False

    def _ir_para_janela_form_real(self, timeout=120):
        """Após o OK da categoria, o form real (~33 campos) costuma abrir em
        NOVA ABA. Vai para a aba com MAIS selects (categoria=1; form real=~33),
        priorizando a URL PR_CADASTRORAPIDOPASTA."""
        fim = time.time() + timeout
        ultimo_log = 0
        while time.time() < fim:
            melhor_handle = None
            melhor_n = -1
            try:
                handles = self.driver.window_handles
            except Exception:
                handles = []
            for h in handles:
                try:
                    self.driver.switch_to.window(h)
                    self.driver.switch_to.default_content()
                except Exception:
                    continue
                url = (self.driver.current_url or "").lower()
                try:
                    n = self.driver.execute_script(
                        "return document.querySelectorAll("
                        "'select.select2-hidden-accessible, "
                        "select.benner-search').length;")
                except Exception:
                    n = 0
                if "cadastrorapidopasta" in url:
                    n += 1000
                if n > melhor_n:
                    melhor_n = n
                    melhor_handle = h
            if melhor_handle is not None and (melhor_n % 1000) >= 10:
                self.driver.switch_to.window(melhor_handle)
                self.driver.switch_to.default_content()
                self._log(f"    Form real ativo (selects={melhor_n % 1000}).")
                return True
            agora = time.time()
            if agora - ultimo_log >= 5:
                self._log(f"    Aguardando FORM REAL abrir "
                          f"(melhor selects={max(melhor_n, 0) % 1000})...")
                ultimo_log = agora
            time.sleep(1.5)
        self._log("    Timeout aguardando o FORM REAL (nova aba).")
        return False

    def _fechar_janelas_extras(self):
        """Fecha todas as abas exceto a principal (Benner) e volta o foco a ela."""
        try:
            principal = getattr(self, "_janela_principal", None)
            handles = self.driver.window_handles
            if not principal or principal not in handles:
                principal = handles[0] if handles else None
            for h in list(handles):
                if principal and h != principal:
                    try:
                        self.driver.switch_to.window(h)
                        self.driver.close()
                    except Exception:
                        pass
            if principal:
                self.driver.switch_to.window(principal)
                self.driver.switch_to.default_content()
        except Exception as e:
            self._log(f"    (aviso) Falha ao fechar abas extras: {e}")

    def _preencher_texto_js(self, label_text, valor, ocorrencia=1):
        elem = self._achar_input_por_label(label_text, ocorrencia)
        if elem is None:
            self._log(f"    Campo texto (label='{label_text}') não encontrado.")
            return False
        try:
            self.driver.execute_script("""
                var e=arguments[0],v=arguments[1];
                if(e.readOnly){e.readOnly=false;}
                e.value='';e.value=v;
                e.dispatchEvent(new Event('input',{bubbles:true}));
                e.dispatchEvent(new Event('change',{bubbles:true}));
                e.dispatchEvent(new Event('blur',{bubbles:true}));""", elem, valor)
            return True
        except Exception as e:
            self._log(f"    Falha preencher '{label_text}': {e}")
            return False

    # ==========================================================================
    def _reforcar_campos_criticos_pre_save(self, numero_cnj):
        """Re-preenche campos de texto que os postbacks dos radios costumam apagar.
        Chamar DEPOIS de todos os radios/limpezas e ANTES do precheck/save."""
        try:
            self._estabilizar_form(timeout=20, quieto=2.0)
        except Exception:
            pass
        hoje = date.today().strftime("%d/%m/%Y")
        self._ultimo_numero_cnj = numero_cnj   # guardado p/ o precheck re-tentar
        self._preencher_texto_por_fieldname("NUMERODISTRIBUICAO", numero_cnj)
        self._preencher_texto_por_fieldname("DATAANDAMENTO1", hoje)
        self._preencher_texto_por_fieldname("DATADISTRIBUICAO", hoje)
        self._log("    [reforço] número/datas re-preenchidos pós-radios.")

    def _marcar_radio_por_pergunta(self, termos, label_desejado):
        """Marca 'label_desejado' (ex.: 'Sim'/'Não') no grupo de radio cuja
        PERGUNTA (label da seção) contém TODOS os 'termos'. Robusto quando o
        NAME do grupo é desconhecido. Ex.: termos=['adverso','cadastrado']."""
        try:
            r = self.driver.execute_script(r"""
                var termos = arguments[0].map(function(t){return t.toLowerCase();});
                var alvoLabel = (arguments[1]||'').toLowerCase().trim();
                function norm(s){ return (s||'').toLowerCase()
                    .normalize('NFD').replace(/[\u0300-\u036f]/g,''); }
                function temTodos(t){ t=norm(t);
                    for(var i=0;i<termos.length;i++){ if(t.indexOf(norm(termos[i]))<0) return false; }
                    return true; }
                // achar o nó da pergunta
                var nos=document.querySelectorAll('label,div,span,td,legend,.tab-label,.label-title');
                var perg=null;
                for(var i=0;i<nos.length;i++){
                    var tx=(nos[i].innerText||nos[i].textContent||'').trim();
                    if(tx.length>0 && tx.length<120 && temTodos(tx)){ perg=nos[i]; break; }
                }
                if(!perg) return 'sem-pergunta';
                // subir até um container e achar o radio cujo label == alvoLabel
                var cont=perg.closest('.form-group,.field,.control-group,tr,.row,fieldset')||perg.parentElement;
                for(var k=0;k<7 && cont;k++){
                    var radios=cont.querySelectorAll("input[type='radio']");
                    for(var j=0;j<radios.length;j++){
                        var rr=radios[j], lab='';
                        if(rr.id){var l=document.querySelector('label[for="'+rr.id+'"]');
                            if(l) lab=l.innerText||l.textContent||'';}
                        if(!lab){var n=rr.nextElementSibling;
                            if(n&&n.tagName==='LABEL') lab=n.innerText||n.textContent||'';}
                        if(norm(lab).trim()===norm(alvoLabel)){
                            rr.checked=true;
                            rr.dispatchEvent(new Event('click',{bubbles:true}));
                            rr.dispatchEvent(new Event('change',{bubbles:true}));
                            if(window.jQuery){try{jQuery(rr).trigger('click').trigger('change');}catch(e){}}
                            return 'ok:'+(rr.id||'radio');
                        }
                    }
                    cont=cont.parentElement;
                }
                return 'sem-radio';
            """, list(termos), label_desejado)
            self._log(f"    [radio-pergunta] {termos}='{label_desejado}': {r}")
            return str(r).startswith("ok")
        except Exception as e:
            self._log(f"    [radio-pergunta] erro: {e}")
            return False

    def _marcar_ja_distribuido_nao(self):
        """Marca 'Não' no grupo 'Já distribuído judicialmente'. A numeração/nome
        do grupo varia, então tentamos: (1) nomes de grupo prováveis; (2) achar
        o radio pelo texto da PERGUNTA na seção (contém 'distribu' + 'judicial')."""
        # (1) tentar nomes de grupo prováveis
        for grupo in ("DISTRIBUIDOJUDICIALMENTE", "JADISTRIBUIDOJUDICIALMENTE",
                      "JADISTRIBUIDO", "DISTRIBUIDOJUDICIAL", "DISTRIBUIDO"):
            try:
                if self._marcar_radio_grupo(grupo, "Não"):
                    self._log(f"    [distribuido] grupo '{grupo}' = Não (OK).")
                    return True
            except Exception:
                pass
        # (2) fallback por texto da pergunta -> clicar o radio 'Não' mais próximo
        try:
            r = self.driver.execute_script(r"""
                function norm(s){return (s||'').toLowerCase()
                    .normalize('NFD').replace(/[\u0300-\u036f]/g,'');}
                // achar um nó de texto/label com 'distribu' e 'judicial'
                var todos=document.querySelectorAll('label,div,span,td,legend');
                var alvo=null;
                for(var i=0;i<todos.length;i++){
                    var t=norm(todos[i].innerText||todos[i].textContent||'');
                    if(t.indexOf('distribu')>=0 && t.indexOf('judicial')>=0){
                        alvo=todos[i]; break;
                    }
                }
                if(!alvo) return 'sem-pergunta';
                // procurar o container e, nele, o radio cujo label = 'Não'
                var cont=alvo.closest('.form-group,.field,.control-group,tr,.row,fieldset')||alvo.parentElement;
                for(var k=0;k<6 && cont;k++){
                    var radios=cont.querySelectorAll("input[type='radio']");
                    for(var j=0;j<radios.length;j++){
                        var rr=radios[j], lab='';
                        if(rr.id){var l=document.querySelector('label[for="'+rr.id+'"]');
                            if(l) lab=l.innerText||l.textContent||'';}
                        if(!lab){var n=rr.nextElementSibling;
                            if(n&&n.tagName==='LABEL') lab=n.innerText||n.textContent||'';}
                        if(norm(lab).trim()==='nao'){
                            rr.checked=true;
                            rr.dispatchEvent(new Event('click',{bubbles:true}));
                            rr.dispatchEvent(new Event('change',{bubbles:true}));
                            if(window.jQuery){try{jQuery(rr).trigger('click').trigger('change');}catch(e){}}
                            return 'ok:'+(rr.id||'radio');
                        }
                    }
                    cont=cont.parentElement;
                }
                return 'sem-radio';
            """)
            self._log(f"    [distribuido] fallback por pergunta: {r}")
            return str(r).startswith("ok")
        except Exception as e:
            self._log(f"    [distribuido] erro: {e}")
            return False

    def _marcar_radio_grupo(self, grupo, label_desejado):
        """Marca uma opção num grupo de radios do Benner, localizando pelo NAME
        (GroupRadioButton_<grupo>) + texto do label (independe de ctlNN).
        Ex.: grupo='TIPOPROCESSO', label='Ativo' ; grupo='PROCESSORELEVANTE','Não'."""
        try:
            r = self.driver.execute_script(r"""
                var grupo=arguments[0], alvo=(arguments[1]||'').toLowerCase().trim();
                var radios=document.querySelectorAll(
                    "input[type='radio'][name*='GroupRadioButton_"+grupo+"']");
                function norm(s){ return (s||'').toLowerCase()
                    .normalize('NFD').replace(/[\u0300-\u036f]/g,'').trim(); }
                function labelDe(el){
                    if(el.id){ var l=document.querySelector('label[for="'+el.id+'"]');
                        if(l) return (l.innerText||l.textContent||'').trim(); }
                    var n=el.nextElementSibling;
                    if(n && n.tagName==='LABEL') return (n.innerText||'').trim();
                    var p=el.parentElement;
                    if(p){ var l2=p.querySelector('label'); if(l2) return (l2.innerText||'').trim(); }
                    return '';
                }
                var alvoN=norm(alvo);
                for(var i=0;i<radios.length;i++){
                    if(norm(labelDe(radios[i]))===alvoN){
                        radios[i].checked=true;
                        radios[i].dispatchEvent(new Event('click',{bubbles:true}));
                        radios[i].dispatchEvent(new Event('change',{bubbles:true}));
                        if(window.jQuery){ try{ jQuery(radios[i]).trigger('click').trigger('change'); }catch(e){} }
                        return 'ok:'+radios[i].id;
                    }
                }
                return 'nao-achou ('+radios.length+' radios)';
            """, grupo, label_desejado)
            self._log(f"    [radio] {grupo}='{label_desejado}': {r}")
            return str(r).startswith("ok")
        except Exception as e:
            self._log(f"    [radio] {grupo} erro: {e}")
            return False

    def _marcar_numero_unico_nao(self):
        """Compatibilidade: marca 'Não' no grupo Número único."""
        return self._marcar_radio_grupo("NUMEROUNICO", "Não")

    def _limpar_documentos_inicial(self):
        """LIMPA os selects de Tipo Documento que vêm preenchidos com 'Inicial'
        (e afins), pois o Benner recusa o save se a seção de Documentos estiver
        preenchida. Esvazia: <select>, hidden *_VALUE (SelectedItems=[]), e clica
        no × (select2-selection__clear) quando existir. Também limpa Nome/Data doc."""
        try:
            r = self.driver.execute_script(r"""
                var limpos=[];
                // 1) selects TIPODOCUMENTO (inclui DISTRIBUIDO e ARQUIVO) com valor
                var sels=document.querySelectorAll('select[data-fieldname]');
                for(var i=0;i<sels.length;i++){
                    var fn=sels[i].getAttribute('data-fieldname')||'';
                    if(fn.indexOf('TIPODOCUMENTOARQUIVO')<0) continue; // só ARQUIVO; preserva TIPODOCUMENTO obrigatório
                    var sel=sels[i];
                    // esvaziar <select>
                    try{ sel.selectedIndex=-1; sel.value=''; }catch(e){}
                    // hidden *_VALUE -> SelectedItems vazio
                    var hid=sel.getAttribute('data-inputhiddenid');
                    if(hid){ var h=document.getElementById(hid);
                        if(h){ h.value=JSON.stringify({ReadOnly:false,View:false,SelectedItems:[]});
                            h.dispatchEvent(new Event('change',{bubbles:true})); } }
                    sel.dispatchEvent(new Event('change',{bubbles:true}));
                    // NAO usar jQuery trigger (crasha Search.js). Limpar o texto
                    // visivel do select2 diretamente no DOM:
                    try{
                        var contd=sel.nextElementSibling;
                        while(contd && !(contd.classList&&contd.classList.contains('select2')))
                            contd=contd.nextElementSibling;
                        if(contd){ var rendd=contd.querySelector('.select2-selection__rendered');
                            if(rendd){ rendd.removeAttribute('title');
                                var itd=rendd.querySelector('.selected-item, xmp');
                                if(itd) itd.textContent=''; else rendd.textContent=''; } }
                    }catch(e){}
                    limpos.push(fn);
                }
                // 2) clicar nos × (clear) visíveis do select2 (garante limpeza visual)
                var clears=document.querySelectorAll('.select2-selection__clear');
                for(var j=0;j<clears.length;j++){
                    try{ clears[j].dispatchEvent(new MouseEvent('mousedown',{bubbles:true}));
                         clears[j].click(); }catch(e){}
                }
                // 3) limpar inputs de Nome do arquivo e Data do documento
                var texts=document.querySelectorAll(
                    "input[id*='NOMEARQUIVO'],input[id*='DATADOCUMENTO']");
                for(var k=0;k<texts.length;k++){
                    if(texts[k].type==='hidden') continue;
                    try{ texts[k].value=''; 
                        texts[k].dispatchEvent(new Event('input',{bubbles:true}));
                        texts[k].dispatchEvent(new Event('change',{bubbles:true})); }catch(e){}
                }
                return 'limpos='+limpos.join(',')+' | clears='+clears.length;
            """)
            self._log(f"    [documentos] limpeza: {r}")
            return True
        except Exception as e:
            self._log(f"    [documentos] erro ao limpar: {e}")
            return False

    def _depurar_pos_preenchimento(self, motivo="diag"):
        """DEPURAÇÃO PROFUNDA (pós-preenchimento / no erro de save).
        Captura e SALVA em <pasta>/diagnostico_pos_save/<timestamp>/:
          - screenshot.png
          - pagina.html (DOM completo)
          - campos.json         -> TODOS os campos (select/input/textarea/radio)
                                   com value, hidden *_VALUE, required, visível.
          - obrigatorios.json   -> campos marcados como obrigatórios e o estado
                                   (preenchido/vazio) + validações do ASP.NET.
          - viewstate.json      -> tamanho do __VIEWSTATE/EVENTVALIDATION, e o
                                   __EVENTTARGET/__EVENTARGUMENT atuais.
          - erros_tela.json     -> mensagens de erro/validação visíveis + o texto
                                   do painel de erro do Benner (renderização).
          - devtools.json       -> console + erros JS + rede (fetch/XHR do save).
          - resumo.txt          -> leitura rápida do que provavelmente bloqueia.
        Retorna o caminho da pasta gerada.
        """
        base = self.arquivo_excel.parent / "diagnostico_pos_save"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pasta = base / ts
        pasta.mkdir(parents=True, exist_ok=True)
        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass

        # -- screenshot --
        try:
            self.driver.save_screenshot(str(pasta / "screenshot.png"))
        except Exception as e:
            self._log(f"    [DEPURAÇÃO] screenshot falhou: {e}")

        # -- HTML completo --
        try:
            (pasta / "pagina.html").write_text(self.driver.page_source, encoding="utf-8")
        except Exception as e:
            self._log(f"    [DEPURAÇÃO] html falhou: {e}")

        # -- TODOS os campos (select/input/textarea/radio) + hidden *_VALUE --
        campos = []
        try:
            campos = self.driver.execute_script(r"""
                function val(el){ try{return el.value;}catch(e){return null;} }
                var out=[];
                var nodes=document.querySelectorAll(
                    'select[data-fieldname], input[data-field], textarea[data-field], '
                    +'input[type=radio], input[type=checkbox], '
                    +"input[id*='_VALUE'], input.form-control, textarea");
                for(var i=0;i<nodes.length;i++){
                    var el=nodes[i];
                    var fn=el.getAttribute('data-fieldname')||el.getAttribute('data-field')||'';
                    var req = el.getAttribute('data-required')
                              || el.getAttribute('aria-required')
                              || (el.className.indexOf('required')>=0?'true':'');
                    var item={
                        tag:el.tagName.toLowerCase(),
                        id:el.id||'', name:el.name||'', fieldname:fn,
                        type:el.getAttribute('type')||'', required:req||'',
                        value:(val(el)||'').toString().slice(0,300),
                        checked:(el.type==='radio'||el.type==='checkbox')?el.checked:null,
                        visivel: !!(el.offsetParent!==null),
                        classe:(el.className||'').slice(0,120)
                    };
                    out.push(item);
                }
                return out;
            """) or []
            (pasta / "campos.json").write_text(
                json.dumps(campos, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            self._log(f"    [DEPURAÇÃO] campos falhou: {e}")

        # -- Campos OBRIGATÓRIOS e estado (vazio/preenchido) --
        try:
            obrig = self.driver.execute_script(r"""
                var res={obrigatorios:[], validacao_aspnet:[]};
                // 1) elementos marcados como required (aria/data/classe .required)
                var reqs=document.querySelectorAll(
                    "[aria-required='true'], [data-required='true'], .required, "
                    +".field-required, label .required");
                var vistos={};
                for(var i=0;i<reqs.length;i++){
                    var host=reqs[i];
                    // subir até o container do campo
                    var cont=host.closest('.form-group,.field,.control-group,tr,.row')||host;
                    var lbl=cont.querySelector('label, .label-title');
                    var ctrl=cont.querySelector('select,input,textarea');
                    var fn = ctrl?(ctrl.getAttribute('data-fieldname')
                             ||ctrl.getAttribute('data-field')||ctrl.id||''):'';
                    if(fn && vistos[fn]) continue; if(fn) vistos[fn]=1;
                    var v = ctrl?(ctrl.value||''):'';
                    // se houver hidden *_VALUE associado, olhar o SelectedItems
                    var preenchido = !!(v && (''+v).trim()!=='');
                    if(ctrl){
                        var hid=ctrl.getAttribute&&ctrl.getAttribute('data-inputhiddenid');
                        if(hid){ var h=document.getElementById(hid);
                            if(h && h.value){ try{ var j=JSON.parse(h.value);
                                preenchido = !!(j.SelectedItems && j.SelectedItems.length
                                    && j.SelectedItems[0].id && j.SelectedItems[0].id!=-1);
                            }catch(e){ preenchido = h.value.indexOf('"id"')>=0
                                && h.value.indexOf('-1')<0; } } }
                    }
                    res.obrigatorios.push({
                        label:(lbl?(lbl.innerText||'').trim():'').slice(0,60),
                        fieldname:fn, preenchido:preenchido,
                        valor:(''+v).slice(0,80)
                    });
                }
                // 2) validações do ASP.NET (spans de validator visíveis)
                var vals=document.querySelectorAll(
                    "span[id*='Valid'], .field-validation-error, .validation-summary-errors li");
                for(var k=0;k<vals.length;k++){
                    var t=(vals[k].innerText||'').trim();
                    if(t && vals[k].offsetParent!==null) res.validacao_aspnet.push(t);
                }
                return res;
            """) or {}
            (pasta / "obrigatorios.json").write_text(
                json.dumps(obrig, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            obrig = {}
            self._log(f"    [DEPURAÇÃO] obrigatorios falhou: {e}")

        # -- ViewState / EventValidation / EventTarget --
        try:
            vs = self.driver.execute_script(r"""
                function tam(id){var e=document.getElementById(id);return e?(''+e.value).length:0;}
                function val(id){var e=document.getElementById(id);return e?(''+e.value).slice(0,120):'';}
                return {
                    viewstate_len: tam('__VIEWSTATE'),
                    eventvalidation_len: tam('__EVENTVALIDATION'),
                    viewstategenerator: val('__VIEWSTATEGENERATOR'),
                    eventtarget: val('__EVENTTARGET'),
                    eventargument: val('__EVENTARGUMENT'),
                    url: location.href
                };
            """) or {}
            (pasta / "viewstate.json").write_text(
                json.dumps(vs, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            self._log(f"    [DEPURAÇÃO] viewstate falhou: {e}")

        # -- Erros na tela (painel de renderização + toasts + modais) --
        erros = {}
        try:
            erros = self.driver.execute_script(r"""
                function texto(sel){ var a=[];
                    document.querySelectorAll(sel).forEach(function(e){
                        var t=(e.innerText||'').trim();
                        if(t && e.offsetParent!==null) a.push(t.slice(0,500)); });
                    return a; }
                return {
                    painel_erro: texto('.error, .alert-danger, .notification.error, '
                        +'.bootstrap-dialog-message, .modal-body, #ctl00_Main_ERROR, '
                        +"[class*='error']"),
                    toasts: texto('.toast, .toast-message, .growl'),
                    body_topo: (document.body.innerText||'').slice(0,1500)
                };
            """) or {}
            (pasta / "erros_tela.json").write_text(
                json.dumps(erros, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            self._log(f"    [DEPURAÇÃO] erros_tela falhou: {e}")

        # -- DevTools: console + erros + rede (fetch/XHR do save) --
        try:
            dev = self._coletar_devtools() if hasattr(self, "_coletar_devtools") else {}
            (pasta / "devtools.json").write_text(
                json.dumps(dev, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            self._log(f"    [DEPURAÇÃO] devtools falhou: {e}")

        # -- RESUMO legível (o que provavelmente bloqueia) --
        try:
            faltando = [o for o in obrig.get("obrigatorios", []) if not o.get("preenchido")]
            linhas = []
            linhas.append(f"MOTIVO: {motivo}")
            linhas.append(f"URL: {self.driver.current_url}")
            linhas.append("")
            linhas.append("=== CAMPOS OBRIGATÓRIOS NÃO PREENCHIDOS ===")
            if faltando:
                for o in faltando:
                    linhas.append(f"  [FALTA] {o.get('label','?')} "
                                  f"(fn={o.get('fieldname','?')}) valor='{o.get('valor','')}'")
            else:
                linhas.append("  (nenhum campo obrigatório detectado como vazio)")
            linhas.append("")
            linhas.append("=== VALIDAÇÕES ASP.NET VISÍVEIS ===")
            for v in obrig.get("validacao_aspnet", []) or ["(nenhuma)"]:
                linhas.append(f"  {v}")
            linhas.append("")
            linhas.append("=== PAINEL DE ERRO DA TELA ===")
            for e in (erros.get("painel_erro", []) or ["(nenhum)"]):
                linhas.append(f"  {e}")
            linhas.append("")
            linhas.append("Arquivos: screenshot.png, pagina.html, campos.json, "
                          "obrigatorios.json, viewstate.json, erros_tela.json, devtools.json")
            (pasta / "resumo.txt").write_text("\n".join(linhas), encoding="utf-8")
            # ecoar o resumo no log do robô
            self._log("    [DEPURAÇÃO] " + " | ".join(
                [f"obrig_vazios={len(faltando)}",
                 f"validacoes={len(obrig.get('validacao_aspnet', []))}",
                 f"campos={len(campos)}"]))
            for o in faltando[:12]:
                self._log(f"        [obrig VAZIO] {o.get('label','?')} (fn={o.get('fieldname','?')})")
        except Exception as e:
            self._log(f"    [DEPURAÇÃO] resumo falhou: {e}")

        return str(pasta)

    def _handle_registro_atual(self):
        """Lê o HANDLE do registro (muda quando o Benner salva e gera a pasta)."""
        for hid in ("ctl00_Main_WIDGET_CADASTRO_RAPIDO_HANDLE_HiddenField",
                    "ctl00_Main_DOCUMENTOS_HANDLE_HiddenField"):
            try:
                v = self.driver.find_element(By.ID, hid).get_attribute("value")
                if v:
                    return v
            except Exception:
                continue
        return None

    def _confirmar_dialogos_pos_save(self, max_dialogos=3):
        """Após Salvar, o Benner pode abrir diálogo(s) (ex.: 'Adverso já
        cadastrado' Não/Sim). Regra do usuário: clicar SIM para reaproveitar a
        pessoa existente. Trata até 'max_dialogos' e reenvia o Save."""
        for _i in range(max_dialogos):
            time.sleep(1.2)
            try:
                r = self.driver.execute_script(r"""
                    function visivel(el){ if(!el) return false;
                        var s=window.getComputedStyle(el);
                        return s && s.display!=='none' && s.visibility!=='hidden'
                            && el.offsetParent!==null; }
                    var seletores=[
                        '.bootstrap-dialog-footer button',
                        '.modal.in .modal-footer button, .modal.show .modal-footer button',
                        '.ui-dialog-buttonset button',
                        '.jconfirm-buttons button',
                        '.swal2-actions button',
                        '.notification button, .toast button'];
                    var botoes=[];
                    for(var s=0;s<seletores.length;s++){
                        var nl=document.querySelectorAll(seletores[s]);
                        for(var j=0;j<nl.length;j++) botoes.push(nl[j]); }
                    var todos=document.querySelectorAll('button, a');
                    for(var k=0;k<todos.length;k++){
                        var t=(todos[k].innerText||todos[k].textContent||'').trim().toLowerCase();
                        if(t==='sim') botoes.push(todos[k]); }
                    for(var b=0;b<botoes.length;b++){
                        var el=botoes[b];
                        var t=(el.innerText||el.textContent||'').trim().toLowerCase();
                        if(t==='sim' && visivel(el)){
                            el.scrollIntoView({block:'center'}); el.click();
                            return 'sim-clicado'; } }
                    return 'sem-dialogo';
                """)
                if r == 'sim-clicado':
                    self._log("    [dialogo] 'Adverso já cadastrado' -> clicado SIM.")
                    self._aguardar_carregamento(); time.sleep(1.2)
                    try:
                        self.driver.execute_script(f"{BTN_SALVAR_JS};")
                        self._log("    [dialogo] Save reenviado após confirmar SIM.")
                        self._aguardar_carregamento()
                    except Exception:
                        pass
                    continue
                else:
                    break
            except Exception as e:
                self._log(f"    [dialogo] erro ao tratar: {e}")
                break

    def _precheck_essenciais_antes_save(self):
        """
        Precheck RELAXADO:
        - Não bloqueia apenas por hidden sem SelectedItems.
        - Considera válido se select.value estiver preenchido.
        - Retorna warnings; bloqueio só para texto crítico vazio.
        """
        warnings = []
        bloqueios = []

        essenciais_fn = ["FILIAL", "TIPO", "RITO", "ORGAO", "UF", "EVENTO1", "PEDIDO1", "RISCOPEDIDO1"]

        try:
            status = self.driver.execute_script(r"""
                var fns = arguments[0];
                var out = [];
                for (var i=0;i<fns.length;i++){
                    var fn = fns[i];
                    var sel = document.querySelector('select[data-fieldname="'+fn+'"]');
                    if(!sel){ out.push({fn:fn, ok:false, modo:'nao_encontrado'}); continue; }
                    var v = (sel.value||'').trim();
                    var hidOk = false;
                    var hid = sel.getAttribute('data-inputhiddenid');
                    if(hid){
                        var h = document.getElementById(hid);
                        if(h && h.value){
                            try{
                                var j = JSON.parse(h.value);
                                hidOk = !!(j && j.SelectedItems && j.SelectedItems.length);
                            }catch(e){ hidOk = h.value.length > 5; }
                        }
                    }
                    out.push({fn:fn, ok: !!(v || hidOk), modo: (v?'value':(hidOk?'hidden':'vazio'))});
                }
                return out;
            """, essenciais_fn) or []

            for it in status:
                if not it.get("ok"):
                    warnings.append(f"{it.get('fn')} vazio")
                elif it.get("modo") == "value":
                    warnings.append(f"{it.get('fn')} sem hidden confirmado (seguindo por value)")

        except Exception as e:
            warnings.append(f"precheck_js_erro: {e}")

        # bloqueio só para textos realmente críticos.
        # Buscar o Número por span[data-field] OU por id contendo NUMERODISTRIBUICAO
        # (consistente com o fix1). Como o reforço re-preenche pós-radios, aqui
        # tentamos re-preencher também se estiver vazio antes de bloquear.
        try:
            numero_ok = bool(self.driver.execute_script("""
                var e=document.querySelector('span[data-field="NUMERODISTRIBUICAO"] input:not([type="hidden"])')
                     || document.querySelector('input[id*="NUMERODISTRIBUICAO"]:not([type="hidden"])');
                return !!(e && (e.value||'').trim());
            """))
            if not numero_ok:
                # última tentativa: re-preencher com o número guardado no objeto
                num = getattr(self, "_ultimo_numero_cnj", "") or ""
                if num:
                    self._preencher_texto_por_fieldname("NUMERODISTRIBUICAO", num)
                    numero_ok = bool(self.driver.execute_script("""
                        var e=document.querySelector('span[data-field="NUMERODISTRIBUICAO"] input:not([type="hidden"])')
                             || document.querySelector('input[id*="NUMERODISTRIBUICAO"]:not([type="hidden"])');
                        return !!(e && (e.value||'').trim());
                    """))
            if not numero_ok:
                bloqueios.append("NUMERODISTRIBUICAO vazio")
        except Exception:
            bloqueios.append("NUMERODISTRIBUICAO não validado")

        return (len(bloqueios) == 0, warnings, bloqueios)

    def _clicar_botao_salvar(self):
        """
        Clique robusto no botão Salvar.
        ORDEM CORRETA:
          1) ActionChains (humano)
          2) click nativo
          3) click JS
          4) Ctrl+Enter
          5) __doPostBack (ÚLTIMO fallback)
        """
        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass

        alvo = None

        # localizar botão principal
        seletores = [
            "a.btn.blue.btn-save",
            "a.btn-save.command-action",
            "a.btn.blue[title*='Ctrl']",
        ]
        for sel in seletores:
            try:
                for e in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if e.is_displayed() and e.is_enabled():
                        alvo = e
                        break
            except Exception:
                pass
            if alvo:
                break

        if alvo is None:
            try:
                for e in self.driver.find_elements(By.XPATH, "//a[contains(normalize-space(),'Salvar')]"):
                    if e.is_displayed() and e.is_enabled():
                        alvo = e
                        break
            except Exception:
                pass

        # 1) ActionChains
        if alvo is not None:
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
                time.sleep(0.25)
                ActionChains(self.driver).move_to_element(alvo).pause(0.15).click().perform()
                self._log("    [salvar] clique real (ActionChains) OK.")
                self._aguardar_carregamento()
                time.sleep(1.0)
                return True
            except Exception as e:
                self._log(f"    [salvar] ActionChains falhou: {e}")

            # 2) click nativo
            try:
                alvo.click()
                self._log("    [salvar] click nativo OK.")
                self._aguardar_carregamento()
                time.sleep(1.0)
                return True
            except Exception as e:
                self._log(f"    [salvar] click nativo falhou: {e}")

            # 3) click JS
            try:
                self.driver.execute_script("arguments[0].click();", alvo)
                self._log("    [salvar] click JS no elemento OK.")
                self._aguardar_carregamento()
                time.sleep(1.0)
                return True
            except Exception as e:
                self._log(f"    [salvar] click JS no elemento falhou: {e}")

        # 4) Ctrl+Enter
        try:
            ActionChains(self.driver).key_down(Keys.CONTROL).send_keys(Keys.ENTER).key_up(Keys.CONTROL).perform()
            self._log("    [salvar] fallback Ctrl+Enter enviado.")
            self._aguardar_carregamento()
            time.sleep(1.0)
            return True
        except Exception as e:
            self._log(f"    [salvar] Ctrl+Enter falhou: {e}")

        # 5) postback (último recurso)
        try:
            self.driver.execute_script("__doPostBack('ctl00$Main$WIDGET_CADASTRO_RAPIDO','Save');")
            self._log("    [salvar] fallback final __doPostBack acionado.")
            self._aguardar_carregamento()
            time.sleep(1.0)
            return True
        except Exception as e:
            self._log(f"    [salvar] postback final falhou: {e}")
            return False

    def _mensagem_validacao_benner(self):
        """Captura mensagens de validação/erro exibidas pelo Benner (campos
        obrigatórios etc.), para diferenciar 'não salvou' de 'salvou'."""
        try:
            txts = []
            for sel in (".validation-summary-errors", ".alert-danger",
                        ".field-validation-error", ".notification.error",
                        ".toast-error", "[class*='error']"):
                for e in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    try:
                        t = (e.text or "").strip()
                        if t and e.is_displayed():
                            txts.append(t)
                    except Exception:
                        continue
            if txts:
                return " | ".join(dict.fromkeys(txts))[:200]
        except Exception:
            pass
        return ""

    def _aguardar_confirmacao_save(self, handle_antes, url_antes, timeout=75):
        """
        Confirma salvamento por:
          (a) HANDLE mudou
          (b) URL mudou
          (c) identificador visível (CIV... / span IDENTIFICADOR)
        """
        padrao = re.compile(r"\b([A-ZÇ]{2,6}[\.\-/]?\s?\d{3,}(?:[\./-]\d{2,4})?)\b")
        fim = time.time() + timeout
        ultimo_log = 0

        while time.time() < fim:
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

            # c1) span IDENTIFICADOR
            try:
                els = self.driver.find_elements(By.XPATH, "//span[@data-field='IDENTIFICADOR']")
                for el in els:
                    t = (el.text or "").strip()
                    if t:
                        self._log(f"    [save-confirm] IDENTIFICADOR detectado: {t}")
                        return True, t
            except Exception:
                pass

            # a) handle mudou
            try:
                h_agora = self._handle_registro_atual()
                if handle_antes and h_agora and h_agora != handle_antes:
                    ident = self._buscar_identificador_civ(padrao)
                    self._log(f"    [save-confirm] HANDLE mudou: {handle_antes} -> {h_agora}")
                    return True, ident or h_agora
            except Exception:
                pass

            # b) URL mudou
            try:
                u = self.driver.current_url or ""
                if url_antes and u and u != url_antes:
                    ident = self._buscar_identificador_civ(padrao)
                    self._log("    [save-confirm] URL mudou após salvar.")
                    return True, ident
            except Exception:
                pass

            # c2) padrão CIV no texto
            try:
                ident = self._buscar_identificador_civ(padrao)
                if ident:
                    self._log(f"    [save-confirm] identificador em texto: {ident}")
                    return True, ident
            except Exception:
                pass

            # Tratar MODAIS GENUÍNOS de confirmação (bootstrap-dialog/modal com
            # botão 'Sim'). NÃO confundir com o radio "Adverso já cadastrado",
            # que é campo do form (tratado no preenchimento).
            try:
                clicou = self.driver.execute_script(r"""
                    function vis(el){ if(!el) return false;
                        var s=window.getComputedStyle(el);
                        return s && s.display!=='none' && s.visibility!=='hidden'
                            && el.offsetParent!==null; }
                    // só botões dentro de um MODAL/dialog visível
                    var sel='.bootstrap-dialog-footer button, .modal.in .modal-footer button, '
                           +'.modal.show .modal-footer button, .ui-dialog-buttonset button, '
                           +'.jconfirm-buttons button, .swal2-actions button';
                    var bts=document.querySelectorAll(sel);
                    for(var i=0;i<bts.length;i++){
                        var el=bts[i];
                        var t=(el.innerText||el.textContent||'').trim().toLowerCase();
                        if(t==='sim' && vis(el)){ el.scrollIntoView({block:'center'});
                            el.click(); return 'sim'; }
                    }
                    return '';
                """)
                if clicou == 'sim':
                    self._log("    [save-confirm] modal de confirmação -> SIM clicado.")
                    self._aguardar_carregamento(); time.sleep(1.5)
                    continue
            except Exception:
                pass

            # erro de validação / renderização REAL
            msg = self._mensagem_validacao_benner()
            if msg:
                self._log(f"    [save-confirm] validação/erro detectado: {msg[:180]}")
                return False, ""

            if time.time() - ultimo_log >= 5:
                self._log("    [save-confirm] aguardando confirmação...")
                ultimo_log = time.time()

            time.sleep(1.0)

        return False, ""


    def _cdp_network_start(self):
        """Liga Network do CDP e limpa buffer local."""
        self._net_events = []
        try:
            self.driver.execute_cdp_cmd("Network.enable", {})
            self._log("    [CDP] Network.enable ON")
        except Exception as e:
            self._log(f"    [CDP] falha enable: {e}")

    def _cdp_network_collect(self):
        """
        Coleta logs de performance (Network.*) e guarda em memória.
        Funciona em Edge/Chrome Chromium quando get_log('performance') está disponível.
        """
        try:
            logs = self.driver.get_log("performance")
        except Exception:
            return

        for entry in logs:
            try:
                msg = json.loads(entry["message"])["message"]
                m = msg.get("method")
                p = msg.get("params", {})
                if m in ("Network.requestWillBeSent", "Network.responseReceived", "Network.loadingFinished"):
                    self._net_events.append(msg)
            except Exception:
                continue

    def _cdp_network_dump(self, nome="rede_dump"):
        """Salva JSON de rede para análise posterior."""
        pasta = self.arquivo_excel.parent / "diagnostico_rede"
        pasta.mkdir(parents=True, exist_ok=True)
        arq = pasta / f"{time.strftime('%Y%m%d_%H%M%S')}_{nome}.json"
        arq.write_text(json.dumps(self._net_events, ensure_ascii=False, indent=2), encoding="utf-8")
        self._log(f"    [CDP] dump rede salvo: {arq}")
        return str(arq)

    def _instalar_hook_rede(self):
        """Captura POSTs (fetch/xhr) direto no browser, independente de
        get_log('performance'). Grava tudo em window.__netcap.events."""
        js = r"""
        (function(){
          if (window.__netcap && window.__netcap.installed) return 'ja';
          window.__netcap = {installed:true, events:[]};

          function push(ev){
            try{
              window.__netcap.events.push(ev);
              if(window.__netcap.events.length>3000) window.__netcap.events.shift();
            }catch(e){}
          }

          // FETCH
          const _fetch = window.fetch;
          window.fetch = async function(input, init){
            const t0 = Date.now();
            let url = ''; let method = 'GET'; let body = '';
            try{
              url = (typeof input === 'string') ? input : (input && input.url ? input.url : '');
              method = (init && init.method) ? String(init.method).toUpperCase() : 'GET';
              body = (init && init.body) ? String(init.body).slice(0,20000) : '';
            }catch(e){}
            try{
              const resp = await _fetch.apply(this, arguments);
              push({kind:'fetch', ts:t0, url:url, method:method, body:body,
                    status: resp ? resp.status : null});
              return resp;
            }catch(err){
              push({kind:'fetch', ts:t0, url:url, method:method, body:body, error:String(err)});
              throw err;
            }
          };

          // XHR
          const _open = XMLHttpRequest.prototype.open;
          const _send = XMLHttpRequest.prototype.send;
          XMLHttpRequest.prototype.open = function(method, url){
            this.__netcap = this.__netcap || {};
            this.__netcap.method = String(method||'GET').toUpperCase();
            this.__netcap.url = String(url||'');
            this.__netcap.t0 = Date.now();
            return _open.apply(this, arguments);
          };
          XMLHttpRequest.prototype.send = function(body){
            this.__netcap = this.__netcap || {};
            this.__netcap.body = body ? String(body).slice(0,20000) : '';
            const self = this;
            function done(){
              try{
                push({kind:'xhr', ts:self.__netcap.t0||Date.now(),
                      url:self.__netcap.url||'', method:self.__netcap.method||'GET',
                      body:self.__netcap.body||'', status:self.status});
              }catch(e){}
            }
            this.addEventListener('loadend', done);
            return _send.apply(this, arguments);
          };

          // HOOK para FORM SUBMIT (postback SÍNCRONO do ASP.NET / __doPostBack)
          try {
            if (!window.__netcap.__formHooked) {
              window.__netcap.__formHooked = true;
              const _submit = HTMLFormElement.prototype.submit;
              HTMLFormElement.prototype.submit = function(){
                try {
                  const formData = new FormData(this);
                  const obj = {};
                  formData.forEach(function(v,k){ obj[k] = v.toString().slice(0,1000); });
                  push({kind:'form-submit', ts:Date.now(),
                        url:this.action||location.href,
                        method:(this.method||'POST').toUpperCase(), body:obj});
                } catch(e){}
                return _submit.apply(this, arguments);
              };
            }
          } catch(e){}

          return 'ok';
        })();
        """
        try:
            r = self.driver.execute_script(js)
            self._log(f"    [NETHOOK] instalação: {r}")
            return True
        except Exception as e:
            self._log(f"    [NETHOOK] falha instalação: {e}")
            return False

    def _dump_posts_hook(self, nome="posts_hook"):
        """Exporta window.__netcap.events para um JSON em diagnostico_rede/."""
        try:
            events = self.driver.execute_script(
                "return (window.__netcap && window.__netcap.events) "
                "? window.__netcap.events : [];")
            pasta = self.arquivo_excel.parent / "diagnostico_rede"
            pasta.mkdir(parents=True, exist_ok=True)
            arq = pasta / f"{time.strftime('%Y%m%d_%H%M%S')}_{nome}.json"
            arq.write_text(json.dumps(events, ensure_ascii=False, indent=2),
                           encoding="utf-8")
            self._log(f"    [NETHOOK] dump salvo: {arq} (eventos={len(events)})")
            return events
        except Exception as e:
            self._log(f"    [NETHOOK] dump falhou: {e}")
            return []

    def _filtrar_posts_save(self, events):
        """Filtra os POSTs relevantes (submit do save / api/search) da captura."""
        crit = []
        for e in events or []:
            m = (e.get("method") or "").upper()
            u = (e.get("url") or "").lower()
            b = (e.get("body") or "").lower()
            if m == "POST" and (
                "/api/search" in u or
                "form.aspx" in u or
                "__eventtarget" in b or
                "widget_cadastro_rapido" in b or
                "tv_cadastrorapidomanual_form" in b
            ):
                crit.append(e)
        self._log(f"    [NETHOOK] posts críticos encontrados: {len(crit)}")
        return crit

    def _extrair_posts_criticos(self):
        """
        Retorna lista resumida dos POSTs relevantes:
        - /api/search
        - form.aspx / postback de save
        """
        out = []
        req_by_id = {}
        res_by_id = {}

        for ev in getattr(self, "_net_events", []):
            m = ev.get("method")
            p = ev.get("params", {})
            if m == "Network.requestWillBeSent":
                r = p.get("request", {})
                rid = p.get("requestId")
                req_by_id[rid] = r
            elif m == "Network.responseReceived":
                rid = p.get("requestId")
                res_by_id[rid] = p.get("response", {})

        for rid, req in req_by_id.items():
            method = req.get("method", "")
            url = req.get("url", "")
            if method != "POST":
                continue
            if ("/api/search" in url) or ("form.aspx" in url.lower()) or ("__dopostback" in str(req).lower()):
                out.append({
                    "requestId": rid,
                    "url": url,
                    "method": method,
                    "headers": req.get("headers", {}),
                    "postData": req.get("postData", "")[:5000],
                    "response_status": res_by_id.get(rid, {}).get("status"),
                    "response_mime": res_by_id.get(rid, {}).get("mimeType"),
                })

        pasta = self.arquivo_excel.parent / "diagnostico_rede"
        pasta.mkdir(parents=True, exist_ok=True)
        arq = pasta / f"{time.strftime('%Y%m%d_%H%M%S')}_posts_criticos.json"
        arq.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        self._log(f"    [CDP] posts críticos: {arq}")
        return out

    def _comparar_saves(self, bom="save_bom.json", ruim="save_ruim.json"):
        """PARSER AUTOMÁTICO: compara dois dumps de POST crítico (save_bom.json =
        submit que FUNCIONOU; save_ruim.json = submit que FALHOU) e aponta
        exatamente quais fieldname/_VALUE diferem — isto é, o que está quebrando
        o submit. Lê os arquivos da pasta diagnostico_rede/ (ou caminhos dados).
        Gera diagnostico_rede/comparacao_saves.txt e ecoa no log."""
        import json as _json
        from urllib.parse import parse_qs, unquote

        pasta = self.arquivo_excel.parent / "diagnostico_rede"

        def _achar(nome):
            p = Path(nome)
            if p.is_absolute() and p.exists():
                return p
            cand = pasta / nome
            if cand.exists():
                return cand
            # pegar o mais recente que contenha o prefixo (save_bom / save_ruim)
            base = nome.replace(".json", "")
            achados = sorted(pasta.glob(f"*{base}*.json"))
            # preferir dumps do HOOK (ignorar os *_cdp que podem vir vazios)
            hook_only = [a for a in achados if "_cdp" not in a.name]
            achados = hook_only or achados
            return achados[-1] if achados else None

        def _postdata_do_arquivo(caminho):
            """Extrai o postData do submit do form.aspx de um dump de posts
            críticos (lista de dicts com url/postData) OU de um JSON já com
            postData direto."""
            try:
                data = _json.loads(Path(caminho).read_text(encoding="utf-8"))
            except Exception as e:
                self._log(f"    [comparar] falha lendo {caminho}: {e}")
                return None
            # caso 1: lista (posts críticos OU eventos do hook fetch/XHR)
            if isinstance(data, list):
                def _corpo(it):
                    return it.get("postData") or it.get("body") or ""
                # priorizar o POST de form.aspx (o submit) com corpo de viewstate
                alvo = None
                for it in data:
                    u = (it.get("url") or "").lower()
                    b = (_corpo(it) or "").lower()
                    if ("form.aspx" in u or "widget_cadastro_rapido" in b
                            or "__eventtarget" in b) and _corpo(it):
                        alvo = it; break
                if alvo is None:
                    for it in data:
                        if _corpo(it):
                            alvo = it; break
                return _corpo(alvo) if alvo else None
            # caso 2: dict único
            if isinstance(data, dict):
                return data.get("postData") or data.get("post_data")
            return None

        def _campos_value(postdata):
            """Do corpo urlencoded, extrai só os campos *_VALUE (os hidden do
            select2) e demais fieldnames relevantes -> dict {campo: valor}."""
            campos = {}
            if not postdata:
                return campos
            try:
                q = parse_qs(postdata, keep_blank_values=True)
            except Exception:
                # fallback simples
                q = {}
                for par in postdata.split("&"):
                    if "=" in par:
                        k, v = par.split("=", 1)
                        q.setdefault(k, []).append(v)
            for k, vals in q.items():
                v = unquote(vals[0]) if vals else ""
                # focar nos hidden _VALUE (select2) e campos de negócio
                if k.endswith("_VALUE") or "_VALUE" in k or any(
                        t in k for t in ("FILIAL","TIPO","RITO","ORGAO","UF","EVENTO",
                        "PEDIDO","RISCO","PARTICIPANTE","CONDICAO","ASSUNTO","DESDOBRA",
                        "INSTANCIA","FASE","DEPARTAMENTO","DIVISAO","TIPODOCUMENTO",
                        "NUMERODISTRIBUICAO","DATA")):
                    # nome curto (fieldname) a partir do id
                    m = k.rsplit("_", 2)
                    nome = k
                    for parte in reversed(k.split("_")):
                        if parte.isupper() and len(parte) >= 3:
                            nome = parte; break
                    campos[k] = {"fieldname": nome, "valor": v[:200]}
            return campos

        pb = _achar(bom)
        pr = _achar(ruim)
        if not pb or not pr:
            self._log(f"    [comparar] arquivos não encontrados "
                      f"(bom={pb}, ruim={pr}). Gere um save_bom.json e um "
                      f"save_ruim.json em diagnostico_rede/.")
            return None

        cbom = _campos_value(_postdata_do_arquivo(pb))
        cruim = _campos_value(_postdata_do_arquivo(pr))

        linhas = []
        linhas.append("COMPARAÇÃO save_bom vs save_ruim (campos *_VALUE / fieldname)")
        linhas.append(f"  bom  = {pb}")
        linhas.append(f"  ruim = {pr}")
        linhas.append("")

        todas = sorted(set(cbom) | set(cruim))
        divergentes = []
        for k in todas:
            vb = cbom.get(k, {}).get("valor", "<ausente>")
            vr = cruim.get(k, {}).get("valor", "<ausente>")
            fn = (cbom.get(k) or cruim.get(k) or {}).get("fieldname", k)
            if vb != vr:
                divergentes.append((fn, k, vb, vr))

        if not divergentes:
            linhas.append(">>> NENHUMA divergência de campo *_VALUE. O problema NÃO")
            linhas.append("    está nos valores dos campos (provável ViewState/token).")
        else:
            linhas.append(f">>> {len(divergentes)} CAMPO(S) DIVERGENTE(S) — prováveis")
            linhas.append("    responsáveis por quebrar o submit:")
            for fn, k, vb, vr in divergentes:
                linhas.append(f"  [DIFERE] fieldname={fn}")
                linhas.append(f"           id={k}")
                linhas.append(f"           BOM = {vb!r}")
                linhas.append(f"           RUIM= {vr!r}")

        # também comparar tokens ASP.NET (viewstate/eventvalidation/eventtarget)
        def _token(postdata, nome):
            if not postdata:
                return ""
            try:
                q = parse_qs(postdata, keep_blank_values=True)
                return (q.get(nome, [""])[0] or "")
            except Exception:
                return ""
        pdb = _postdata_do_arquivo(pb); pdr = _postdata_do_arquivo(pr)
        linhas.append("")
        linhas.append("=== Tokens ASP.NET (tamanho) ===")
        for tk in ("__EVENTTARGET", "__EVENTARGUMENT", "__VIEWSTATE", "__EVENTVALIDATION"):
            tb = _token(pdb, tk); tr = _token(pdr, tk)
            linhas.append(f"  {tk}: bom(len={len(tb)}) ruim(len={len(tr)}) "
                          f"{'IGUAL' if tb==tr else 'DIFERE'}")

        pasta.mkdir(parents=True, exist_ok=True)
        saida = pasta / "comparacao_saves.txt"
        saida.write_text("\n".join(linhas), encoding="utf-8")
        self._log(f"    [comparar] relatório: {saida}")
        for ln in linhas[:40]:
            self._log("      " + ln)
        return str(saida)

    def _save_transacional(self):
        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass

        try:
            self._estabilizar_form(timeout=30, quieto=2.5)
        except Exception:
            pass
        time.sleep(0.4)

        ok_pre, warnings, bloqueios = self._precheck_essenciais_antes_save()
        for w in warnings[:20]:
            self._log(f"    [precheck-warning] {w}")

        if not ok_pre:
            detalhe = "Precheck bloqueante: " + "; ".join(bloqueios)
            self._log(f"    [save] {detalhe}")
            try:
                self._depurar_pos_preenchimento(motivo=detalhe)
            except Exception:
                pass
            return False, "", detalhe

        handle_antes = self._handle_registro_atual()
        url_antes = self.driver.current_url or ""

        # >>> CAPTURA DE REDE (hook fetch/XHR): reinstalar por segurança e ligar
        #     o CDP como reforço, ANTES de clicar Salvar. <<<
        try:
            self._instalar_hook_rede()   # reinstala após postbacks
        except Exception as e:
            self._log(f"    [NETHOOK] reinstalar falhou: {e}")
        try:
            self._cdp_network_start()    # reforço (se performance log existir)
            time.sleep(0.3)
        except Exception:
            pass

        if not self._clicar_botao_salvar():
            return False, "", "Falha ao acionar botão Salvar."

        try:
            self._confirmar_dialogos_pos_save(max_dialogos=3)
        except Exception as e:
            self._log(f"    [save] aviso dialogo: {e}")

        # coletar por ~12s (o hook grava sozinho; aqui só damos tempo + CDP reforço)
        try:
            for _ in range(12):
                time.sleep(1)
                self._cdp_network_collect()
        except Exception:
            pass

        ok_save, ident = self._aguardar_confirmacao_save(handle_antes, url_antes, timeout=75)

        # DUMP do hook nomeado conforme resultado + posts críticos + comparação
        try:
            nome = "save_bom" if ok_save else "save_ruim"
            events = self._dump_posts_hook(nome)          # <- fonte confiável
            self._filtrar_posts_save(events)               # loga qtd de POSTs de save
            self._cdp_network_dump(nome + "_cdp")          # reforço (pode vir vazio)
            self._extrair_posts_criticos()
            self._comparar_saves()                         # compara bom vs ruim
        except Exception as e:
            self._log(f"    [NETHOOK] dump/extração falhou: {e}")

        if ok_save:
            if not ident:
                ident = self._capturar_id_pasta() or ""
            return True, ident, "OK"

        msg = self._mensagem_validacao_benner() or "Não confirmou salvamento."
        try:
            self._depurar_pos_preenchimento(motivo=msg)
        except Exception:
            pass
        return False, "", msg

    def _buscar_identificador_civ(self, padrao):
        """Varre o corpo/breadcrumb procurando um identificador de pasta que
        comece com letras (ex.: CIV, CÍV) seguido de número."""
        try:
            corpo = self.driver.find_element(By.TAG_NAME, "body").text or ""
        except Exception:
            return ""
        for linha in corpo.splitlines():
            up = linha.strip().upper()
            if up.startswith("CIV") or " CIV" in up or "CÍV" in up:
                m = padrao.search(linha.strip())
                if m:
                    return m.group(1)
        return ""

    def _capturar_id_pasta(self):
        try:
            el = self.driver.find_element(By.XPATH, "//span[@data-field='IDENTIFICADOR']")
            t = (el.text or "").strip()
            if t:
                return t
        except Exception:
            pass
        try:
            h = self.driver.find_element(By.ID, ID_PASTA_HANDLE)
            v = h.get_attribute("value")
            if v:
                return v
        except Exception:
            pass
        try:
            # pega o pst= do FORM REAL (PR_CADASTRORAPIDOPASTA)
            u = self.driver.current_url or ""
            if "PR_CADASTRORAPIDOPASTA" in u or "form.aspx" in u:
                m = re.search(r"(pst=\w+)", u)
                if m:
                    return m.group(1)
        except Exception:
            pass
        return ""

    def _screenshot_erro(self, prefixo="erro"):
        try:
            nome = f"{prefixo}_{time.strftime('%Y%m%d_%H%M%S')}.png"
            caminho = str(self.arquivo_excel.parent / nome)
            self.driver.save_screenshot(caminho)
            return caminho
        except Exception:
            return ""

    def _exibir_erro_e_parar(self, mensagem):
        self._log("PROCESSO PARALISADO - ERRO DETECTADO")
        self._log(mensagem)
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, mensagem, "Cadastro Pastas Benner - PARALISADO", 0x10)
        except Exception:
            pass

    def _aguardar_carregamento(self):
        try:
            self.wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
        except TimeoutException:
            pass
        time.sleep(WAIT_AFTER_CLICK)

    def _buscar_elemento_por_texto(self, tag, texto):
        try:
            for elem in self.driver.find_elements(By.TAG_NAME, tag):
                try:
                    if texto.lower() in (elem.text or "").lower():
                        return elem
                except StaleElementReferenceException:
                    continue
        except Exception:
            pass
        return None

    @staticmethod
    def _formatar_cpf(cpf_raw):
        cpf = _so_digitos(cpf_raw)
        if not cpf:
            return ""
        if len(cpf) > 11:
            return cpf
        cpf = cpf.zfill(11)
        return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:11]}"

    def gerar_relatorio(self):
        self.carregar_planilha()
        ws = self.ws
        last_row = ws.max_row
        st = {"pend": 0, "cad": 0, "err": 0, "dup": 0, "ja": 0, "agr": 0}
        for row in range(2, last_row + 1):
            s = str(ws.cell(row, COL_STATUS).value or "").strip().upper()
            if s == "PENDENTE": st["pend"] += 1
            elif s.startswith("CADASTRADO"): st["cad"] += 1
            elif s == "NÃO CADASTRAR": st["dup"] += 1
            elif s.startswith("JÁ CADASTRADO"): st["ja"] += 1
            elif s == "AGRUPADO": st["agr"] += 1
            elif "ERRO" in s: st["err"] += 1
        self._log(f"""
{'='*50}
 RELATÓRIO DE STATUS
{'='*50}
 Pendentes:    {st['pend']}
 Cadastradas:  {st['cad']}
 Agrupadas:    {st['agr']}
 Duplicatas:   {st['dup']}
 Já no Benner: {st['ja']}
 Erros:        {st['err']}
{'='*50}""")


# ==============================================================================
def main():
    import sys
    DIR = Path(r"K:\BennerData\CadastraPastas")
    arquivo = sys.argv[1] if len(sys.argv) > 1 else \
        str(DIR / "Ajuizamento+2024+2+parte+ (2) -Planilha original.xlsx")
    if not Path(arquivo).exists():
        print(f"ERRO: Arquivo não encontrado: {arquivo}")
        sys.exit(1)

    cadastro = CadastroPastasBenner(arquivo)
    print("\n" + "=" * 60)
    print(" CADASTRO DE PASTAS BENNER - PREVI JURÍDICO")
    print(" Parecer PAR.0000871/26")
    print("=" * 60)
    while True:
        print("""
Opções:
  0 - Etapa 0: Gravação/diagnóstico manual (você nomeia cada passo no terminal)
  1 - Etapa 1: Análise prévia de duplicidades
  2 - Etapa 2: Verificar no Benner (Nome > Pastas > Dívida Previdenciária)
  3 - Etapa 3: Cadastrar pastas
  4 - Relatório
  5 - Tudo (1+2+3)
  9 - Sair
""")
        op = input("Escolha: ").strip()
        if op == "0":
            cadastro.etapa_zero_gravacao_manual(); cadastro.fechar()
        elif op == "1":
            cadastro.analise_previa_duplicidades()
        elif op == "2":
            cadastro.verificar_no_benner(); cadastro.fechar()
        elif op == "3":
            cadastro.cadastrar_pastas(); cadastro.fechar()
        elif op == "4":
            cadastro.gerar_relatorio()
        elif op == "5":
            cadastro.analise_previa_duplicidades()
            cadastro.verificar_no_benner()
            cadastro.cadastrar_pastas()
            cadastro.fechar()
        elif op == "9":
            cadastro.fechar(); break
        else:
            print("Opção inválida.")


if __name__ == "__main__":
    main()