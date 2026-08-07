#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
CADASTRO DE PASTAS BENNER (PREVI JURÍDICO) — versão PLAYWRIGHT
Reescrita da Version90 (Selenium) incorporando PLAYWRIGHT com INTERAÇÃO NATIVA.
================================================================================
Parecer PAR.0000871/26 - Ajuizamento dívidas prev. 2024 Parte 2

POR QUE PLAYWRIGHT:
  O Selenium injetava os hidden *_VALUE via JS, mas o servidor ASP.NET WebForms
  RESETAVA o form no Save (id:-1) — provado por captura de rede — porque os
  valores não passavam pelo ciclo de POSTBACKS nativo. O Playwright CLICA e
  DIGITA de verdade, disparando os postbacks (UpdatePanel) que POPULAM o
  ViewState no servidor. `wait_for_load_state('networkidle')` espera cada
  postback assíncrono terminar. É a abordagem que respeita o WebForms.

MENU (igual à V90):
  1 - Etapa 1: Análise prévia de duplicidades (Python puro, sem navegador)
  3 - Etapa 3: Cadastrar pastas (PLAYWRIGHT — attach ao Edge debug 9222)
  4 - Relatório
  9 - Sair

MODO DE CONEXÃO (attach, igual à V90):
  1) Feche o Edge. 2) Rode abrir_edge_debug.bat (porta 9222). 3) Faça LOGIN no
  Benner. 4) Rode: py cadastro_pastas_benner_playwright.py  -> opção 3.

PRÉ-REQUISITO:  py -m pip install playwright openpyxl
================================================================================
"""
import asyncio
import json
import os
import random
import re
import time
from datetime import date, datetime
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

CDP_URL = os.environ.get("BENNER_CDP_URL", "http://127.0.0.1:9222")
MARK_CATEGORIA = "cadastrorapidomanual"      # aba do seletor de categoria
MARK_FORM      = "pr_cadastrorapidopasta"    # aba do formulário real



# ==============================================================================
URL_BENNER = "https://previ.bennercloud.com.br/JURIDICO/jur/e/PREVI.aspx?i=K9_INICIOPREVI&m=MAIN"

CATEGORIA = "Cível"
CAUSA_PEDIR = "Previdencial"
CAUSA_RAIZ = "Produto"
RITO = "Ordinário"
ANDAMENTO = "PEDIDO DE AJUIZAMENTO DE AÇÃO"
PEDIDO = "Dívida Previdenciária"
RISCO = "Possível"
GERENCIA_FALLBACK = "GEPAB"
COND_AUTOR = "Autor"
COND_REU = "Réu"
OBSERVACAO = "Cadastrado conforme Parecer PAR.0000871/26 - Ajuizamento dívidas previdenciárias (Ano 2024 Parte 2)"

LBL_CATEGORIA = "Categoria"
LBL_FILIAL = "Filial"
LBL_GERENCIA = "Gerência"
LBL_CAUSA_PEDIR = "Causa de Pedir"
LBL_CAUSA_RAIZ = "Causa Raiz"
LBL_RITO = "Rito"
LBL_ORGAO = "Orgão"
LBL_UF = "UF"
LBL_ANDAMENTO = "Andamento"
LBL_PARTICIPANTE = "Participante"
LBL_CONDICAO = "Condição"
LBL_ADV_INTERNO = "Advogado interno"
LBL_ADV_EXTERNO = "Advogado externo"
LBL_PEDIDO = "Pedido"
LBL_RISCO = "Risco"
LBL_OBSERVACOES = "Observações"
LBL_NUMERO = "Número"
LBL_DATA_ANDAMENTO = "Data andamento"
LBL_VALOR_PEDIDO = "Valor pedido"

# ------------------------------------------------------------------------------
# MAPA DE CAMPOS DO FORMULÁRIO REAL (PR_CADASTRORAPIDOPASTA)
# IDs confirmados via diagnóstico (Etapa 0) do HTML/JSON reais do Benner.
# Preencher por ID é MUITO mais confiável do que por label+ocorrência, porque
# o form tem vários campos com o MESMO label (Andamento x2, Participante x3,
# Condição x5, Pedido x2, Risco x2). Se o layout mudar e os IDs "ctlNN" mudarem,
# o código cai automaticamente no fallback por label.
# ------------------------------------------------------------------------------
_PREF = "ctl00_Main_WIDGET_CADASTRO_RAPIDO_PageControl_GERAL_GERAL_"
FIELD_IDS = {
    "FILIAL":        _PREF + "ctl11_ctl01_select",
    "DIRETORIA":     _PREF + "ctl18_ctl01_select",
    "GERENCIA":      _PREF + "ctl22_ctl01_select",
    "TIPO":          _PREF + "ctl30_ctl01_select",
    "CAUSA_PEDIR":   _PREF + "ctl34_ctl01_select",
    "CAUSA_RAIZ":    _PREF + "ctl43_ctl01_select",
    "RITO":          _PREF + "ctl87_ctl01_select",
    "ORGAO":         _PREF + "ctl95_ctl01_select",
    "UF":            _PREF + "ctl99_ctl01_select",
    "ANDAMENTO_1":   _PREF + "ctl122_ctl01_select",
    # Participantes: bloco RÉU (pessoa da planilha) e bloco AUTOR (PREVI)
    "PART_REU":      _PREF + "POSSUIPESSOAADVERSO_2_ctl04_ctl01_select",
    "COND_REU":      _PREF + "POSSUIPESSOAADVERSO_2_ctl13_ctl01_select",
    "PART_AUTOR":    _PREF + "ctl163_ctl01_select",
    "COND_AUTOR":    _PREF + "ctl172_ctl01_select",
    "ADV_INTERNO":   _PREF + "ctl218_ctl01_select",
    "ADV_EXTERNO":   _PREF + "ctl222_ctl01_select",
    "PEDIDO_1":      _PREF + "ctl229_ctl01_select",
    "RISCO_1":       _PREF + "ctl244_ctl01_select",
    # Campos de texto (IDs semânticos estáveis)
    "NUMERO":           _PREF + "NUMEROUNICO_1_NUMERODISTRIBUICAO",
    "DATA_ANDAMENTO_1": _PREF + "DATAANDAMENTO1_DATE",
    "VALOR_PEDIDO_1":   _PREF + "VALORPEDIDO1",
    "OBSERVACOES":      _PREF + "OBSERVACOES",
}

# ------------------------------------------------------------------------------
# MAPA POR data-fieldname (ESTÁVEL) — descoberto no diagnóstico do form real.
# Os IDs "ctlNN" MUDAM entre sessões do Benner; já o data-fieldname (e o
# data-field dos spans de texto) é ESTÁVEL. Este é o mapeamento oficial.
# ------------------------------------------------------------------------------
# SELECTS (localizados por select[data-fieldname="..."]):
FN = {
    "FILIAL": "FILIAL",
    "DIRETORIA": "DEPARTAMENTO",
    "GERENCIA": "DIVISAO",
    "TIPO": "TIPO",
    "CAUSA_PEDIR": "ASSUNTO",           # cascata: só popula após TIPO
    "CAUSA_RAIZ": "CAUSARAIZ",
    "RITO": "RITO",
    "ORGAO": "ORGAO",
    "UF": "UF",
    "ANDAMENTO_1": "EVENTO1",
    # Participantes (CONFIRMADO pelo usuário - Opção A):
    #   Bloco 1 = pessoa da planilha (Réu) ; Bloco 2 = PREVI (Autor)
    "PART_REU": "PARTICIPANTE1",
    "COND_REU": "CONDICAO1",
    "PART_AUTOR": "PARTICIPANTE2",
    "COND_AUTOR": "CONDICAO2",
    "ADV_INTERNO": "ADVOGADOINTERNO",
    "ADV_EXTERNO": "ADVOGADOEXTERNO",
    "PEDIDO_1": "PEDIDO1",
    "RISCO_1": "RISCOPEDIDO1",
}
# CAMPOS DE TEXTO (localizados por span[data-field="..."] input/textarea):
TEXT_FN = {
    "NUMERO": "NUMERODISTRIBUICAO",
    "DATA_ANDAMENTO_1": "DATAANDAMENTO1",
    "VALOR_PEDIDO_1": "VALORPEDIDO1",
    "OBSERVACOES": "OBSERVACOES",
}

# ------------------------------------------------------------------------------
# IDs INTERNOS FIXOS (via DevTools). Campos com valor IGUAL em todas as pastas.
# ------------------------------------------------------------------------------
# Padrões dos campos que o usuário definiu (Instância/Fase/Distribuição):
INSTANCIA_PADRAO = "1º Grau"
FASE_PADRAO = "Preliminar"      # confirmado pelo usuário (NÃO é "Conhecimento")

IDS_FIXOS = {
    "TIPO":         (978,  "COBRANÇA"),
    "DESDOBRAMENTO": (139, "Cobrança"),   # campo "Processo" = Cobrança
    "INSTANCIA":    (1,    "1º grau"),     # Instância = 1º Grau (id fixo)
    "ASSUNTO":      (5447, "PREVIDENCIAL"),
    "CAUSARAIZ":    (2,    "Produto"),
    "RITO":         (7,    "Ordinário"),
    "EVENTO1":      (4415, "Pedido de ajuizamento de ação"),
    "PEDIDO1":      (411,  "DÍVIDA PREVIDENCIÁRIA"),
    "RISCOPEDIDO1": (1,    "Possível"),
}
COND_ID = {"Réu": 2, "Reu": 2, "Autor": 1}
PREVI_ID = 137650
PREVI_TEXTO = "PREVI - CAIXA DE PREVIDÊNCIA DOS FUNCIONÁRIOS DO BANCO DO BRASIL"

# Cascatas conhecidas: campo dependente -> (fieldname do PAI)
CASCATA_PAI = {
    "DIVISAO": "DEPARTAMENTO",   # Gerência depende de Diretoria
}
# UF (ESTADOS) - tabela de IDs (do diagnóstico /api/search)
UF_ID = {
    "AC":8,"AL":9,"AM":10,"AP":11,"BA":12,"CE":13,"DF":14,"ES":15,"GO":16,
    "MA":17,"MG":18,"MS":19,"MT":20,"PA":21,"PB":22,"PE":23,"PI":24,"PR":3,
    "RJ":25,"RN":26,"RO":27,"RR":28,"RS":7,"SC":1,"SE":29,"SP":6,"TO":30,
}
# FILIAL (id interno) - tabela fixa (a API de FILIAL não responde à busca):
FILIAL_ID = {
    "PLANO DE BENEFICIOS 1": 3,
    # "PLANO PREVI FUTURO": <id> (se aparecer, o robô tenta a API/opções embutidas)
}

# Fallback por LABEL (usado só se o ID exato não existir mais no DOM):
_FALLBACK_LABEL = {
    "FILIAL": LBL_FILIAL, "DIRETORIA": "Diretoria", "GERENCIA": LBL_GERENCIA,
    "TIPO": "Tipo", "CAUSA_PEDIR": LBL_CAUSA_PEDIR, "CAUSA_RAIZ": LBL_CAUSA_RAIZ,
    "RITO": LBL_RITO, "ORGAO": LBL_ORGAO, "UF": LBL_UF,
    "ANDAMENTO_1": LBL_ANDAMENTO,
    "PART_REU": LBL_PARTICIPANTE, "COND_REU": LBL_CONDICAO,
    "PART_AUTOR": LBL_PARTICIPANTE, "COND_AUTOR": LBL_CONDICAO,
    "ADV_INTERNO": LBL_ADV_INTERNO, "ADV_EXTERNO": LBL_ADV_EXTERNO,
    "PEDIDO_1": LBL_PEDIDO, "RISCO_1": LBL_RISCO,
    "NUMERO": LBL_NUMERO, "DATA_ANDAMENTO_1": LBL_DATA_ANDAMENTO,
    "VALOR_PEDIDO_1": LBL_VALOR_PEDIDO, "OBSERVACOES": LBL_OBSERVACOES,
}
# Ocorrência do label no fallback (bloco Réu = 1ª; Autor = 2ª).
_FALLBACK_OCC = {
    "PART_REU": 1, "COND_REU": 1, "PART_AUTOR": 2, "COND_AUTOR": 2,
    "PEDIDO_1": 1, "RISCO_1": 1, "ANDAMENTO_1": 1,
    "DATA_ANDAMENTO_1": 1, "VALOR_PEDIDO_1": 1,
}

# Valores fixos confirmados no cadastro manual (ground truth) e pelo usuário:
TIPO = "Cobrança"           # campo "Tipo" = COBRANÇA (vale p/ todas as linhas)
PREVI_AUTOR = "PREVI"       # participante Autor (PREVI - CAIXA DE PREVIDÊNCIA...)

BTN_OK_CATEGORIA_JS = "__doPostBack('ctl00$Main$TV_CADASTRORAPIDOMANUAL_FORM','Save')"
BTN_SALVAR_JS = "__doPostBack('ctl00$Main$WIDGET_CADASTRO_RAPIDO','Save')"
ID_PASTA_HANDLE = "ctl00_Main_PASTA_VIEW_HANDLE_HiddenField"

COL_PLANO = 1
COL_NOME = 4
COL_CONTRATO = 6
COL_VALOR_DIVIDA = 15
COL_GERENCIA = 17
COL_UF = 20
COL_CPF = 23
COL_BENNER_FLAG = 28
COL_ANALISE = 30
COL_STATUS = 31
COL_CNJ = 32
COL_PLANO_DESC = 33
COL_PESQUISA_BENNER = 34
COL_ID_PASTA = 35
COL_VALOR_PEDIDO = 36

ADVOGADOS_INTERNOS = [
    "EDSON EDUARDO AGUIAR AVELAR",
    "MICHELLE CERQUEIRA NUNEZ",
    "DOMINIQUE DE SOUZA MACHADO",
]
ADVOGADOS_EXTERNOS = [
    "Aldrigues Cândido Advocacia",
    "Bicudo, Matos, e Moraes Sociedade de Advogados",
    "Dannemann Siemsen Advogados",
    "Queiroga, Vieira, Queiroz & Ramos Advocacia",
    "Wambier, Yamasaki, Bevervanço & Lobo Advocacia",
]
UF_ORGAO = {
    "AC": "do Acre", "AL": "de Alagoas", "AP": "do Amapá", "AM": "do Amazonas",
    "BA": "da Bahia", "CE": "do Ceará", "DF": "do Distrito Federal", "ES": "do Espírito Santo",
    "GO": "de Goiás", "MA": "do Maranhão", "MT": "de Mato Grosso", "MS": "de Mato Grosso do Sul",
    "MG": "de Minas Gerais", "PA": "do Pará", "PB": "da Paraíba", "PR": "do Paraná",
    "PE": "de Pernambuco", "PI": "do Piauí", "RJ": "do Rio de Janeiro",
    "RN": "do Rio Grande do Norte", "RS": "do Rio Grande do Sul", "RO": "de Rondônia",
    "RR": "de Roraima", "SC": "de Santa Catarina", "SP": "de São Paulo",
    "SE": "de Sergipe", "TO": "do Tocantins",
}

WAIT_TIMEOUT = 30
WAIT_AFTER_CLICK = 2
WAIT_BETWEEN_RECORDS = 3

# ------------------------------------------------------------------------------
# CONEXÃO (attach ao Edge debug) — Playwright via CDP.
# 1) Feche o Edge. 2) Abra em modo debug (abrir_edge_debug.bat, porta 9222).
# 3) Faça LOGIN no Benner. 4) Rode este script -> opção 3.
# A porta pode ser trocada por variável de ambiente BENNER_CDP_URL.
# ------------------------------------------------------------------------------


def _parse_valor_br(txt) -> float:
    if txt is None or txt == "":
        return 0.0
    if isinstance(txt, (int, float)):
        return float(txt)
    s = str(txt).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _so_digitos(s) -> str:
    return re.sub(r"\D", "", str(s or ""))


# ==============================================================================

# ==============================================================================
# HELPERS PLAYWRIGHT (async)
# ==============================================================================
def _termo_busca(texto):
    esp = {
        "pedido de ajuizamento de ação": "ajuizamento",
        "dívida previdenciária": "previden",
        "plano de beneficios 1": "beneficios 1",
        "previdencial": "previden", "produto": "produto", "ordinário": "ordin",
        "possível": "poss", "diretoria de seguridade": "seguridade", "cível": "cível",
    }
    low = (texto or "").lower().strip()
    if low in esp:
        return esp[low]
    pal = (texto or "").split()
    return pal[0][:8] if pal else (texto or "")[:8]


class CadastroPastasBennerPW:
    """Cadastro de pastas via Playwright (interação nativa)."""

    def __init__(self, arquivo_excel, sheet_name=None):
        self.arquivo_excel = Path(arquivo_excel)
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        self._log_path = str(self.arquivo_excel.parent /
                             f"log_execucao_{time.strftime('%Y%m%d_%H%M%S')}.txt")

    # ---------- utilidades (reaproveitadas da V90) ----------
    def _log(self, msg):
        linha = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
        print(linha)
        try:
            with open(self._log_path, "a", encoding="utf-8") as f:
                f.write(linha + "\n")
        except Exception:
            pass

    def carregar_planilha(self):
        self.wb = openpyxl.load_workbook(str(self.arquivo_excel))
        if self.sheet_name and self.sheet_name in self.wb.sheetnames:
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb[self.wb.sheetnames[0]]
        self._log(f"Planilha carregada: {self.arquivo_excel} [{self.ws.title}]")

    def salvar_planilha(self):
        self.wb.save(str(self.arquivo_excel))
        self._log("Planilha salva.")

    # ---------- ETAPA 1 (Python puro, reaproveitada da V90) ----------
    def analise_previa_duplicidades(self):
        self.carregar_planilha()
        ws = self.ws
        last_row = ws.max_row

        ws.cell(1, COL_ANALISE, "ANÁLISE DUPLICIDADE / PESQUISA BENNER (Pasta)")
        ws.cell(1, COL_STATUS, "STATUS CADASTRO")
        ws.cell(1, COL_CNJ, "NÚMERO CNJ")
        ws.cell(1, COL_PLANO_DESC, "PLANO DESCRIÇÃO (Filial)")
        ws.cell(1, COL_PESQUISA_BENNER, "PESQUISA BENNER")
        ws.cell(1, COL_ID_PASTA, "ID PASTA BENNER")
        ws.cell(1, COL_VALOR_PEDIDO, "VALOR PEDIDO")

        for row in range(2, last_row + 1):
            for col in (COL_ANALISE, COL_STATUS, COL_CNJ, COL_PLANO_DESC,
                        COL_PESQUISA_BENNER, COL_ID_PASTA, COL_VALOR_PEDIDO):
                ws.cell(row, col, None)

        def _linha_vazia(r):
            return not str(ws.cell(r, COL_NOME).value or "").strip()

        participante_linhas = {}
        for row in range(2, last_row + 1):
            if _linha_vazia(row):
                continue
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            participante_linhas.setdefault(nome, []).append(row)

        chaves_vistas = set()
        duplicatas_exatas = set()
        for row in range(2, last_row + 1):
            if _linha_vazia(row):
                continue
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            contrato = str(ws.cell(row, COL_CONTRATO).value or "")
            valor = _parse_valor_br(ws.cell(row, COL_VALOR_DIVIDA).value)
            chave = f"{nome}|{contrato}|{valor}"
            if chave in chaves_vistas:
                duplicatas_exatas.add(row)
            chaves_vistas.add(chave)

        total = 0
        for row in range(2, last_row + 1):
            if _linha_vazia(row):
                continue
            total += 1
            nome = str(ws.cell(row, COL_NOME).value or "").strip().upper()
            contrato = str(ws.cell(row, COL_CONTRATO).value or "")

            plano = ws.cell(row, COL_PLANO).value
            ws.cell(row, COL_PLANO_DESC,
                    "PLANO PREVI FUTURO" if str(plano).strip() == "2" else "PLANO DE BENEFICIOS 1")

            benner_flag = str(ws.cell(row, COL_BENNER_FLAG).value or "").strip()

            if row in duplicatas_exatas:
                ws.cell(row, COL_ANALISE, "DUPLICATA EXATA - REMOVER")
                ws.cell(row, COL_STATUS, "NÃO CADASTRAR")
                ws.cell(row, COL_CNJ, f"DP{contrato}")
                continue

            if benner_flag:
                ws.cell(row, COL_ANALISE, f"ANÁLISE MANUAL: {benner_flag}")

            linhas_participante = participante_linhas.get(nome, [])
            linhas_validas = [r for r in linhas_participante if r not in duplicatas_exatas]

            if len(linhas_validas) > 1:
                contratos = [str(ws.cell(r, COL_CONTRATO).value or "") for r in linhas_validas]
                contratos_unicos = list(dict.fromkeys(contratos))
                numero_combinado = "DP" + "/".join(contratos_unicos)
                valor_total = sum(_parse_valor_br(ws.cell(r, COL_VALOR_DIVIDA).value) for r in linhas_validas)

                if row == linhas_validas[0]:
                    ws.cell(row, COL_ANALISE,
                            (str(ws.cell(row, COL_ANALISE).value or "") +
                             f" | MESMO PARTICIPANTE - {len(linhas_validas)} OPERAÇÕES (AGRUPADO)").strip(" |"))
                    ws.cell(row, COL_STATUS, "PENDENTE")
                    ws.cell(row, COL_CNJ, numero_combinado)
                    ws.cell(row, COL_VALOR_PEDIDO, round(valor_total, 2))
                else:
                    ws.cell(row, COL_ANALISE,
                            (str(ws.cell(row, COL_ANALISE).value or "") +
                             f" | AGRUPADO COM LINHA {linhas_validas[0]} - PASTA ÚNICA").strip(" |"))
                    ws.cell(row, COL_STATUS, "AGRUPADO")
                    ws.cell(row, COL_CNJ, numero_combinado)
            else:
                if not ws.cell(row, COL_ANALISE).value:
                    ws.cell(row, COL_ANALISE, "OK")
                ws.cell(row, COL_STATUS, "PENDENTE")
                ws.cell(row, COL_CNJ, f"DP{contrato}")
                ws.cell(row, COL_VALOR_PEDIDO,
                        round(_parse_valor_br(ws.cell(row, COL_VALOR_DIVIDA).value), 2))

        self.salvar_planilha()
        self._log(f"Etapa 1 concluída. Operações válidas: {total}.")
        return total

    # ==========================================================================
    # ETAPA 2
    # ==========================================================================

    # ---------- ETAPA 4 (relatório, reaproveitada da V90) ----------
    def gerar_relatorio(self):
        self.carregar_planilha()
        ws = self.ws
        last_row = ws.max_row
        st = {"pend": 0, "cad": 0, "err": 0, "dup": 0, "ja": 0, "agr": 0}
        for row in range(2, last_row + 1):
            s = str(ws.cell(row, COL_STATUS).value or "").strip().upper()
            if s == "PENDENTE": st["pend"] += 1
            elif s.startswith("CADASTRADO"): st["cad"] += 1
            elif s == "NÃO CADASTRAR": st["dup"] += 1
            elif s.startswith("JÁ CADASTRADO"): st["ja"] += 1
            elif s == "AGRUPADO": st["agr"] += 1
            elif "ERRO" in s: st["err"] += 1
        self._log(f"""
{'='*50}
 RELATÓRIO DE STATUS
{'='*50}
 Pendentes:    {st['pend']}
 Cadastradas:  {st['cad']}
 Agrupadas:    {st['agr']}
 Duplicatas:   {st['dup']}
 Já no Benner: {st['ja']}
 Erros:        {st['err']}
{'='*50}""")

    # ==========================================================================
    # ETAPA 3 — CADASTRO VIA PLAYWRIGHT (interação nativa)
    # ==========================================================================
    async def _instalar_hook_rede(self, page):
        """Injeta hook fetch/XHR + form-submit em window.__netcap (igual V90)."""
        js = r"""
        (function(){
          if (window.__netcap && window.__netcap.installed) return 'ja';
          window.__netcap = {installed:true, events:[]};
          function push(ev){ try{ window.__netcap.events.push(ev);
            if(window.__netcap.events.length>3000) window.__netcap.events.shift(); }catch(e){} }
          const _fetch = window.fetch;
          window.fetch = async function(input, init){
            const t0=Date.now(); let url='',method='GET',body='';
            try{ url=(typeof input==='string')?input:(input&&input.url?input.url:'');
              method=(init&&init.method)?String(init.method).toUpperCase():'GET';
              body=(init&&init.body)?String(init.body).slice(0,20000):''; }catch(e){}
            try{ const r=await _fetch.apply(this,arguments);
              push({kind:'fetch',ts:t0,url:url,method:method,body:body,status:r?r.status:null});
              return r; }catch(err){ push({kind:'fetch',ts:t0,url:url,method:method,body:body,error:String(err)}); throw err; }
          };
          const _open=XMLHttpRequest.prototype.open, _send=XMLHttpRequest.prototype.send;
          XMLHttpRequest.prototype.open=function(m,u){ this.__nc={m:String(m||'GET').toUpperCase(),u:String(u||''),t0:Date.now()}; return _open.apply(this,arguments); };
          XMLHttpRequest.prototype.send=function(b){ this.__nc=this.__nc||{}; this.__nc.body=b?String(b).slice(0,20000):'';
            const s=this; this.addEventListener('loadend',function(){ try{ push({kind:'xhr',ts:s.__nc.t0||Date.now(),url:s.__nc.u||'',method:s.__nc.m||'GET',body:s.__nc.body||'',status:s.status}); }catch(e){} });
            return _send.apply(this,arguments); };
          try{ if(!window.__netcap.__formHooked){ window.__netcap.__formHooked=true;
            const _sub=HTMLFormElement.prototype.submit;
            HTMLFormElement.prototype.submit=function(){ try{ const fd=new FormData(this); const o={};
              fd.forEach(function(v,k){o[k]=v.toString().slice(0,1000);});
              push({kind:'form-submit',ts:Date.now(),url:this.action||location.href,method:(this.method||'POST').toUpperCase(),body:o}); }catch(e){}
              return _sub.apply(this,arguments); }; } }catch(e){}
          return 'ok';
        })();
        """
        try:
            r = await page.evaluate(js)
            self._log(f"    [NETHOOK] instalação: {r}")
            return True
        except Exception as e:
            self._log(f"    [NETHOOK] falha: {e}")
            return False

    async def _dump_hook(self, page, nome="save_ruim"):
        """Exporta window.__netcap.events para diagnostico_rede/<ts>_<nome>.json."""
        try:
            events = await page.evaluate(
                "() => (window.__netcap && window.__netcap.events) ? window.__netcap.events : []")
            pasta = self.arquivo_excel.parent / "diagnostico_rede"
            pasta.mkdir(parents=True, exist_ok=True)
            arq = pasta / f"{time.strftime('%Y%m%d_%H%M%S')}_{nome}.json"
            arq.write_text(json.dumps(events, ensure_ascii=False, indent=2), encoding="utf-8")
            self._log(f"    [NETHOOK] dump: {arq} (eventos={len(events)})")
            return events
        except Exception as e:
            self._log(f"    [NETHOOK] dump falhou: {e}")
            return []

    async def _esperar_rede(self, page, t=30000):
        try:
            await page.wait_for_load_state("networkidle", timeout=t)
        except PWTimeout:
            pass

    async def _achar_page(self, ctx, marca, timeout=90, precisa_select=None):
        import time as _t
        fim = _t.time() + timeout
        while _t.time() < fim:
            for pg in ctx.pages:
                try:
                    u = (pg.url or "").lower()
                except Exception:
                    u = ""
                if marca in u:
                    if precisa_select:
                        try:
                            await pg.wait_for_selector(
                                f'select[data-fieldname="{precisa_select}"]',
                                state="attached", timeout=3000)
                            return pg
                        except Exception:
                            continue
                    return pg
            await asyncio.sleep(1)
        return None

    async def _sel2(self, page, fieldname, texto, espera=True, timeout=15000):
        """Select2 nativo (corrigido): abre o container do PROPRIO campo, e usa o
        input de busca do DROPDOWN ABERTO (select2-container--open) - evita o
        'strict mode: 3 elements' porque ha varios select2 na pagina."""
        if not texto:
            return False
        try:
            sel = page.locator(f'select[data-fieldname="{fieldname}"]').first
            await sel.wait_for(state="attached", timeout=timeout)
            
            # **CORREÇÃO AQUI**: Verificar se o elemento container está no DOM e visível ANTES de interagir
            cont = sel.locator('xpath=following-sibling::*[contains(@class,"select2")][1]')
            try:
                # Espera o container ficar anexado e visível
                await cont.wait_for(state="attached", timeout=timeout)
                await cont.wait_for(state="visible", timeout=timeout)
                await cont.scroll_into_view_if_needed()
            except PWTimeout:
                 self._log(f"    [select2] container para {fieldname} não ficou visível a tempo.")
                 return False
            except Exception as e:
                 self._log(f"    [select2] Erro ao aguardar container para {fieldname}: {e}")
                 return False

            # fechar qualquer dropdown aberto antes
            try:
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(200) # Pequena pausa para garantir o fechamento
            except Exception:
                pass
            
            # clicar no "selection" (a caixa visivel) para abrir o dropdown
            selection = cont.locator(".select2-selection").first
            try:
                 if await selection.count() and await selection.is_visible():
                     await selection.click()
                 else:
                     await cont.click()
            except Exception as e:
                self._log(f"    [select2] Erro ao clicar no container/selection para {fieldname}: {e}")
                return False

            await page.wait_for_timeout(300)
            # o input de busca do dropdown ABERTO: dentro de .select2-dropdown
            # (que e unico quando ha um dropdown aberto) OU o container--open.
            busca = page.locator(".select2-dropdown input.select2-search__field").first
            if await busca.count() == 0:
                busca = page.locator(
                    ".select2-container--open input.select2-search__field").first
            if await busca.count() == 0:
                # fallback: o ultimo input de busca (o recem-aberto)
                busca = page.locator("input.select2-search__field").last
            await busca.wait_for(state="visible", timeout=6000)
            await busca.fill("")
            await busca.type(_termo_busca(texto), delay=40)
            await page.wait_for_timeout(700)
            try:
                await page.wait_for_selector(
                    "li.select2-results__option:not(.loading-results)", timeout=timeout)
            except PWTimeout:
                pass
            # opcoes SEMPRE dentro do dropdown aberto
            base = page.locator(".select2-results").last
            opt = base.locator("li.select2-results__option", has_text=texto).first
            if await opt.count() == 0:
                pref = texto.split()[0][:8] if texto.split() else texto[:8]
                opt = base.locator("li.select2-results__option", has_text=pref).first
            if await opt.count() == 0:
                opt = base.locator(
                    "li.select2-results__option:not(.loading-results)").first
            
            try:
                 await opt.wait_for(state="attached", timeout=timeout)
                 await opt.wait_for(state="visible", timeout=timeout)
                 await opt.scroll_into_view_if_needed()
                 await opt.click()
            except Exception as e:
                 self._log(f"    [select2] Erro ao clicar na opção para {fieldname}: {e}")
                 return False

            if espera:
                await self._esperar_rede(page)
            self._log(f"    [select2] {fieldname} = '{texto[:30]}' OK.")
            return True
        except Exception as e:
            self._log(f"    [select2] {fieldname} FALHOU: {str(e)[:120]}")
            try:
                await page.keyboard.press("Escape")
            except Exception:
                pass
            return False

    async def _texto(self, page, fieldname, valor):
        if valor is None:
            return False
        for s in (f'span[data-field="{fieldname}"] input:not([type=hidden])',
                  f'span[data-field="{fieldname}"] textarea',
                  f'input[id*="{fieldname}"]:not([type=hidden])',
                  f'textarea[id*="{fieldname}"]'):
            loc = page.locator(s).first
            try:
                if await loc.count() and await loc.is_visible():
                    await loc.scroll_into_view_if_needed()
                    await loc.fill(str(valor))
                    await loc.press("Tab")
                    self._log(f"    [texto] {fieldname} = '{str(valor)[:25]}' OK.")
                    return True
            except Exception:
                continue
        self._log(f"    [texto] {fieldname} não encontrado.")
        return False

    async def _radio(self, page, termos, label):
        js = """(a)=>{const T=a.termos.map(t=>t.toLowerCase());
          const A=(a.label||'').toLowerCase().trim();
          const N=s=>(s||'').toLowerCase().normalize('NFD').replace(/[̀-ͯ]/g,'');
          const todos=t=>{t=N(t);return T.every(x=>t.indexOf(N(x))>=0);};
          const nos=document.querySelectorAll('label,div,span,td,legend,.tab-label,.label-title');
          let P=null;for(const n of nos){const tx=(n.innerText||n.textContent||'').trim();
            if(tx.length>0&&tx.length<120&&todos(tx)){P=n;break;}}
          if(!P)return'sem-pergunta';
          let c=P.closest('.form-group,.field,.control-group,tr,.row,fieldset')||P.parentElement;
          for(let k=0;k<7&&c;k++){const rs=c.querySelectorAll("input[type='radio']");
            for(const r of rs){let l='';if(r.id){const e=document.querySelector('label[for="'+r.id+'"]');if(e)l=e.innerText||e.textContent||'';}
              if(!l){const nx=r.nextElementSibling;if(nx&&nx.tagName==='LABEL')l=nx.innerText||nx.textContent||'';}
              if(N(l).trim()===N(A))return r.id||'RID';}c=c.parentElement;}
          return'sem-radio';}"""
        try:
            rid = await page.evaluate(js, {"termos": list(termos), "label": label})
            if rid in ("sem-pergunta", "sem-radio", "RID"):
                self._log(f"    [radio] {termos}='{label}': {rid}")
                return False
            esc = re.sub(r'([^a-zA-Z0-9_\-])', r'\\', rid)
            r = page.locator(f'#{esc}')
            await r.scroll_into_view_if_needed()
            # clicar no LABEL (mais confiavel que check() no ASP.NET)
            try:
                lab = page.locator(f'label[for="{esc}"]').first
                if await lab.count():
                    await lab.click()
                else:
                    await r.click(force=True)
            except Exception:
                await r.click(force=True)
            await self._esperar_rede(page)
            self._log(f"    [radio] {termos}='{label}': OK.")
            return True
        except Exception as e:
            self._log(f"    [radio] {termos}='{label}' erro: {e}")
            return False

    async def _blocos_part(self, page):
        js = """()=>{const c=(p,e)=>{const o=[];document.querySelectorAll('select[data-fieldname]').forEach(s=>{
          const f=s.getAttribute('data-fieldname')||'';if(f.indexOf(p)===0&&f.indexOf(e)<0)o.push(f);});return o;};
          return{parts:c('PARTICIPANTE','SEMLOTE'),conds:c('CONDICAO','LOTE')};}"""
        r = await page.evaluate(js)
        return r.get("parts", []), r.get("conds", [])

    async def _abrir_e_categoria(self, home, ctx):
        """Abre 'Cadastro rápido', vai à aba da categoria, Cível+OK, retorna a
        page do FORM REAL (ou None)."""
        try:
            await home.locator("#sidebar_novoItem").click(timeout=6000)
            await home.wait_for_timeout(1000)
        except Exception:
            pass
        for txt in ("Cadastro rápido de pasta", "Cadastro rapido de pasta"):
            try:
                link = home.get_by_text(txt, exact=False).first
                if await link.count():
                    await link.click(timeout=5000)
                    await self._esperar_rede(home)
                    break
            except Exception:
                pass
        cat = await self._achar_page(ctx, MARK_CATEGORIA, timeout=30, precisa_select="CATEGORIA")
        if cat is None:
            try:
                await home.wait_for_selector('select[data-fieldname="CATEGORIA"]',
                                             state="attached", timeout=8000)
                cat = home
            except Exception:
                cat = None
        if cat is None:
            self._log("    [categoria] aba/seletor não apareceu.")
            return None
        await cat.bring_to_front()
        await self._sel2(cat, "CATEGORIA", CATEGORIA, espera=False)
        await cat.wait_for_timeout(800)
        # garantir que CATEGORIA commitou (hidden CATEGORIA_VALUE com id valido)
        for _tent in range(3):
            commit = await cat.evaluate(r"""
                () => { var s=document.querySelector('select[data-fieldname="CATEGORIA"]');
                    if(!s) return false;
                    var hid=s.getAttribute('data-inputhiddenid');
                    if(hid){ var h=document.getElementById(hid);
                        if(h && h.value && h.value.indexOf('"id":-1')<0 && h.value.indexOf('"id"')>=0) return true; }
                    return (s.value && s.value!=='' && s.value!=='-1'); }
            """)
            if commit:
                break
            self._log(f"    [categoria] Civel nao commitou (tent {_tent+1}); repetindo...")
            await self._sel2(cat, "CATEGORIA", CATEGORIA, espera=False)
            await cat.wait_for_timeout(800)
        # OK da categoria: o Benner usa __doPostBack (nao um link 'Ok' comum).
        # Igual a V90: __doPostBack('ctl00$Main$TV_CADASTRORAPIDOMANUAL_FORM','Save')
        try:
            await cat.evaluate(
                "() => { __doPostBack('ctl00$Main$TV_CADASTRORAPIDOMANUAL_FORM','Save'); }")
            self._log("    [categoria] OK acionado (__doPostBack).")
        except Exception as e:
            self._log(f"    [categoria] __doPostBack falhou: {e}; tentando botao visivel...")
            for nome_ok in ("Ok", "OK", "Confirmar"):
                try:
                    ok = cat.get_by_role("link", name=re.compile(rf"^\s*{nome_ok}\s*$", re.I)).first
                    if await ok.count():
                        await ok.click(); break
                except Exception:
                    pass
        await self._esperar_rede(cat)
        form = await self._achar_page(ctx, MARK_FORM, timeout=120, precisa_select="TIPO")
        if form is None:
            # talvez o form real tenha carregado na MESMA aba da categoria
            try:
                await cat.wait_for_selector('select[data-fieldname="TIPO"]',
                                            state="attached", timeout=8000)
                form = cat
            except Exception:
                form = None
        if form is None:
            self._log("    [form] aba do formulário real não apareceu.")
            # DUMP de rede para diagnostico mesmo sem salvar
            try:
                await self._dump_hook(cat, "form_nao_apareceu")
            except Exception:
                pass
            return None
        await form.bring_to_front()
        self._log(f"    [conn] form real: {form.url[:60]}")
        await self._instalar_hook_rede(form)   # captura de rede p/ diagnostico
        return form

    async def _preencher(self, page, dados):
        nome = dados["nome"]; uf = dados["uf"]
        numero = dados["numero_cnj"]; valor = dados["valor_pedido"]
        adv_int = random.choice(ADVOGADOS_INTERNOS)
        adv_ext = random.choice(ADVOGADOS_EXTERNOS)
        self._log(f"    === Preenchendo: {nome} (UF={uf}) ===")

        await self._sel2(page, FN["DIRETORIA"], "Diretoria de Seguridade")
        await self._sel2(page, FN["GERENCIA"], "GESOP")
        fil = "PLANO DE BENEFICIOS 1"
        await self._sel2(page, FN["FILIAL"], fil)

        await self._sel2(page, "TIPO", "COBRANÇA")
        await page.wait_for_timeout(1500)
        await self._sel2(page, "ASSUNTO", "PREVIDENCIAL")
        await self._sel2(page, "CAUSARAIZ", "Produto")
        await self._sel2(page, "DESDOBRAMENTO", "Cobrança")
        await self._sel2(page, "RITO", "Ordinário")
        await self._sel2(page, "INSTANCIA", "1º grau")
        await self._sel2(page, "FASE", "Preliminar")
        await self._texto(page, TEXT_FN["OBSERVACOES"], OBSERVACAO)

        orgao = f"Tribunal de Justiça do Estado {UF_ORGAO.get(uf, '')}".strip()
        await self._sel2(page, "ORGAO", orgao)
        await self._sel2(page, "UF", uf)

        await self._radio(page, ["número", "único"], "Não")
        await self._radio(page, ["tipo", "processo"], "Ativo")
        await self._radio(page, ["processo", "relevante"], "Não")
        await self._radio(page, ["localiza"], "Física")
        await self._radio(page, ["distribu", "judicial"], "Não")
        # "Adverso já cadastrado" = Sim ANTES dos participantes (reconstrói seção)
        await self._radio(page, ["adverso", "cadastrado"], "Sim")
        await self._radio(page, ["advogado", "adverso"], "Sim")
        await page.wait_for_timeout(1200)

        await self._texto(page, TEXT_FN["NUMERO"], numero)
        await self._sel2(page, "EVENTO1", "Pedido de ajuizamento de ação")
        hoje = date.today().strftime("%d/%m/%Y")
        await self._texto(page, TEXT_FN["DATA_ANDAMENTO_1"], hoje)
        await self._texto(page, "DATADISTRIBUICAO", hoje)

        parts, conds = await self._blocos_part(page)
        self._log(f"    [participantes] parts={parts} conds={conds}")
        if len(parts) >= 1:
            await self._sel2(page, parts[0], nome)
            if len(conds) >= 1:
                await self._sel2(page, conds[0], "Réu")
        if len(parts) >= 2:
            await self._sel2(page, parts[1], "PREVI")
            if len(conds) >= 2:
                await self._sel2(page, conds[1], "Autor")

        await self._sel2(page, "ADVOGADOINTERNO", adv_int)
        await self._sel2(page, "ADVOGADOEXTERNO", adv_ext)
        await self._sel2(page, "PEDIDO1", "DÍVIDA PREVIDENCIÁRIA")
        if valor and valor > 0:
            await self._texto(page, TEXT_FN["VALOR_PEDIDO_1"],
                              f"{valor:.2f}".replace(".", ","))
        await self._sel2(page, "RISCOPEDIDO1", "Possível")

        self._log("    Salvando (clique nativo)...")
        return await self._salvar(page)

    async def _salvar(self, page):
        try:
            await self._instalar_hook_rede(page)   # reinstala apos postbacks
        except Exception:
            pass
        async def dlg():
            try:
                b = page.locator(
                    ".modal.in .modal-footer button, .modal.show .modal-footer button, "
                    ".bootstrap-dialog-footer button", has_text="Sim").first
                if await b.count() and await b.is_visible():
                    await b.click(); await self._esperar_rede(page)
                    self._log("    [salvar] diálogo -> Sim.")
            except Exception:
                pass
        try:
            btn = page.locator("a.btn.blue.btn-save, a.btn-save.command-action").first
            if await btn.count() == 0:
                btn = page.get_by_role("link", name=re.compile("Salvar", re.I)).first
            await btn.scroll_into_view_if_needed()
            await btn.click()
        except Exception as e:
            self._log(f"    [salvar] falha clique: {e}")
            return False, ""
        for _ in range(30):
            await page.wait_for_timeout(1000)
            await dlg()
            try:
                iel = page.locator("span[data-field='IDENTIFICADOR']").first
                if await iel.count():
                    t = (await iel.inner_text() or "").strip()
                    if t:
                        self._log(f"    [salvar] IDENTIFICADOR: {t}")
                        try:
                            await self._dump_hook(page, "save_bom")
                        except Exception:
                            pass
                        return True, t
            except Exception:
                pass
            try:
                corpo = (await page.locator("body").inner_text())[:2000].lower()
                if "problema ao renderizar" in corpo or "failed to load viewstate" in corpo:
                    self._log("    [salvar] ERRO renderização (viewstate).")
                    try:
                        await self._dump_hook(page, "save_ruim")
                    except Exception:
                        pass
                    return False, "viewstate"
            except Exception:
                pass
        try:
            await self._dump_hook(page, "save_ruim")   # diagnostico de rede
        except Exception:
            pass
        await self._esperar_rede(page)
        return False, "sem-confirmacao"

    async def cadastrar_pastas_async(self):
        """ETAPA 3 (async): attach ao Edge, itera pendentes e cadastra via PW."""
        self.carregar_planilha()
        ws = self.ws
        pend = [r for r in range(2, ws.max_row + 1)
                if str(ws.cell(r, COL_STATUS).value or "").strip().upper() == "PENDENTE"]
        self._log(f"Cadastrando {len(pend)} pastas (Playwright)...")
        if not pend:
            self._log("Nenhuma operação PENDENTE."); return

        async with async_playwright() as p:
            self._log(f"Conectando ao Edge debug {CDP_URL} ...")
            browser = await p.chromium.connect_over_cdp(CDP_URL)
            ctx = browser.contexts[0] if browser.contexts else await browser.new_context()
            home = None
            for pg in ctx.pages:
                if "bennercloud" in (pg.url or "").lower():
                    home = pg; break
            home = home or (ctx.pages[0] if ctx.pages else await ctx.new_page())
            self._log(f"Aba inicial: {home.url[:70]}")

            ok_t = err_t = 0
            for row in pend:
                nome = str(ws.cell(row, COL_NOME).value or "").strip()
                contrato = str(ws.cell(row, COL_CONTRATO).value or "")
                uf = str(ws.cell(row, COL_UF).value or "").strip().upper()
                numero = str(ws.cell(row, COL_CNJ).value or f"DP{contrato}")
                vp = ws.cell(row, COL_VALOR_PEDIDO).value
                valor = _parse_valor_br(vp) if vp not in (None, "")                     else _parse_valor_br(ws.cell(row, COL_VALOR_DIVIDA).value)
                try:
                    try:
                        await home.bring_to_front()
                    except Exception:
                        pass
                    form = await self._abrir_e_categoria(home, ctx)
                    if form is None:
                        self._log(f"  [linha {row}] {nome}: não chegou ao form real.")
                        err_t += 1; break
                    salvo, ident = await self._preencher(form, {
                        "nome": nome, "uf": uf, "numero_cnj": numero,
                        "valor_pedido": valor})
                    if salvo:
                        ws.cell(row, COL_STATUS, "CADASTRADO (PW)")
                        if ident:
                            ws.cell(row, COL_ID_PASTA, str(ident))
                        self.salvar_planilha()
                        self._log(f"  [linha {row}] {nome}: OK -> {ident}")
                        ok_t += 1
                    else:
                        ws.cell(row, COL_STATUS, f"ERRO PW: {ident}")
                        self.salvar_planilha()
                        self._log(f"  [linha {row}] {nome}: FALHOU ({ident})")
                        err_t += 1; break
                except Exception as e:
                    self._log(f"  [linha {row}] {nome}: EXCEÇÃO {e}")
                    err_t += 1; break
            self._log(f"Concluído! Sucesso: {ok_t}, Erros: {err_t}")

    async def _verificar_no_benner_async(self):
        """ETAPA 2 (Playwright): para cada PENDENTE, pesquisa o nome no Benner e
        anota na planilha se ja existe pasta de Divida Previdenciaria."""
        self.carregar_planilha()
        ws = self.ws
        pend = [r for r in range(2, ws.max_row + 1)
                if str(ws.cell(r, COL_STATUS).value or "").strip().upper() == "PENDENTE"]
        self._log(f"[Etapa 2] Verificando {len(pend)} nome(s) no Benner...")
        if not pend:
            self._log("Nenhuma operação PENDENTE."); return
        async with async_playwright() as p:
            browser = await p.chromium.connect_over_cdp(CDP_URL)
            ctx = browser.contexts[0] if browser.contexts else await browser.new_context()
            page = None
            for pg in ctx.pages:
                if "bennercloud" in (pg.url or "").lower():
                    page = pg; break
            page = page or (ctx.pages[0] if ctx.pages else await ctx.new_page())
            for row in pend:
                nome = str(ws.cell(row, COL_NOME).value or "").strip()
                try:
                    # busca global do Benner (lupa) por nome
                    await page.keyboard.press("Control+Space")
                    await page.wait_for_timeout(500)
                    box = page.locator("input.searcher-input, input#searcher-input, input[type='search']").first
                    if await box.count():
                        await box.fill(nome[:30])
                        await page.wait_for_timeout(1500)
                        txt = ""
                        try:
                            txt = (await page.locator("#searcher-results-items").inner_text())[:500]
                        except Exception:
                            pass
                        achou = "pasta" in txt.lower() or "processo" in txt.lower()
                        ws.cell(row, COL_PESQUISA_BENNER,
                                "POSSIVEL DUPLICIDADE" if achou else "NAO ENCONTRADO")
                        self._log(f"  [{row}] {nome[:25]}: {'DUPLIC?' if achou else 'ok'}")
                    await page.keyboard.press("Escape")
                except Exception as e:
                    self._log(f"  [{row}] {nome[:25]}: erro {e}")
            self.salvar_planilha()
            self._log("[Etapa 2] Concluída (resultados na coluna PESQUISA BENNER).")

    def verificar_no_benner(self):
        """Wrapper sincrono da Etapa 2."""
        asyncio.run(self._verificar_no_benner_async())

    def cadastrar_pastas(self):
        """Wrapper síncrono para o menu (roda o async)."""
        asyncio.run(self.cadastrar_pastas_async())

    def fechar(self):
        # attach: NÃO fechamos o Edge do usuário.
        pass


# ==============================================================================
def main():
    import sys
    DIR = Path(r"K:\BennerData\CadastraPastas")
    arquivo = sys.argv[1] if len(sys.argv) > 1 else         str(DIR / "Ajuizamento+2024+2+parte+ (2) -Planilha original.xlsx")
    if not Path(arquivo).exists():
        print(f"ERRO: Arquivo não encontrado: {arquivo}")
        sys.exit(1)

    cadastro = CadastroPastasBennerPW(arquivo)
    print("=" * 60)
    print(" CADASTRO DE PASTAS BENNER - PREVI JURÍDICO (PLAYWRIGHT)")
    print(" Parecer PAR.0000871/26")
    print("=" * 60)
    while True:
        print("""
Opções:
  1 - Etapa 1: Análise prévia de duplicidades
  2 - Etapa 2: Verificar no Benner (Nome > Pastas > Dívida Previdenciária)
  3 - Etapa 3: Cadastrar pastas (Playwright — Edge debug 9222)
  4 - Relatório
  5 - Tudo (1+2+3)
  9 - Sair
""")
        op = input("Escolha: ").strip()
        if op == "1":
            cadastro.analise_previa_duplicidades()
        elif op == "2":
            cadastro.verificar_no_benner()
        elif op == "3":
            cadastro.cadastrar_pastas(); cadastro.fechar()
        elif op == "4":
            cadastro.gerar_relatorio()
        elif op == "5":
            cadastro.analise_previa_duplicidades()
            cadastro.verificar_no_benner()
            cadastro.cadastrar_pastas()
            cadastro.fechar()
        elif op == "9":
            cadastro.fechar(); break
        else:
            print("Opção inválida.")


if __name__ == "__main__":
    main()
